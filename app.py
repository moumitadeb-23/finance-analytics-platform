from asyncio import events
from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    session,
    flash,
    jsonify,
    send_file
)
from flask_mail import Mail, Message
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import random
import string
import time
import sqlite3
import re
import os
import hashlib
from datetime import timedelta
from datetime import datetime
import pandas as pd
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.units import inch
from openpyxl import Workbook
import os
import csv
import uuid
from collections import defaultdict
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from ai.fraud_analysis import get_dataset_summary
from ai.predictor import (
    predict_transaction,
    get_dropdown_data
)
from database.db import (
    add_investment,
    get_all_investments,
    delete_investment,
    get_portfolio_analytics,
    update_investment,
    add_financial_goal,
    get_all_financial_goals,
    update_financial_goal,
    add_goal_savings,
    delete_financial_goal,
    add_notification,
    get_dataset_info,
    get_dataset_history
)

from finance_ai import FinanceAI
finance_ai = FinanceAI()

app = Flask(__name__)

app.config["MAIL_SERVER"] = "smtp.gmail.com"
app.config["MAIL_PORT"] = 587
app.config["MAIL_USE_TLS"] = True
app.config["MAIL_USERNAME"] = "moumitadeb.official@gmail.com"
app.config["MAIL_PASSWORD"] = "fquk dzpb tmku eicx"
app.config["MAIL_DEFAULT_SENDER"] = "moumitadeb.official@gmail.com"

mail = Mail(app)

# ============================================================
# AI UPLOAD CONFIGURATION
# ============================================================

UPLOAD_FOLDER = os.path.join(

    app.root_path,

    "uploads"

)

os.makedirs(

    UPLOAD_FOLDER,

    exist_ok=True

)

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

ALLOWED_EXTENSIONS = {

    "pdf",

    "xlsx",

    "xls",

    "csv",

    "png",

    "jpg",

    "jpeg"

}

# =====================================================
# DATASET CONFIGURATION
# =====================================================

DATASET_UPLOAD_FOLDER = os.path.join(
    app.root_path,
    "uploads",
    "datasets"
)

os.makedirs(DATASET_UPLOAD_FOLDER, exist_ok=True)

ALLOWED_EXTENSIONS = {"csv"}

app.config["DATASET_UPLOAD_FOLDER"] = DATASET_UPLOAD_FOLDER

def allowed_dataset(filename):

    return (

        "." in filename

        and

        filename.rsplit(".",1)[1].lower()

        in ALLOWED_EXTENSIONS

    )

# ============================================================
# CHECK FILE TYPE
# ============================================================

def allowed_file(filename):

    return (

        "." in filename

        and

        filename.rsplit(".",1)[1].lower()

        in ALLOWED_EXTENSIONS

    )

# ============================================================
# SAVE UPLOADED FILE
# ============================================================

def save_uploaded_file(file):

    if file is None:

        return None

    if file.filename=="":

        return None

    if not allowed_file(file.filename):

        return None

    extension=file.filename.rsplit(".",1)[1].lower()

    filename=f"{uuid.uuid4()}.{extension}"

    filepath=os.path.join(

        app.config["UPLOAD_FOLDER"],

        filename

    )

    file.save(filepath)

    return filepath

# ============================================================
# DELETE TEMP FILE
# ============================================================

def delete_uploaded_file(path):

    try:

        if path and os.path.exists(path):

            os.remove(path)

    except:

        pass



# ================= OTP STORAGE =================
otp_storage = {}


def calculate_health_score(income, expense, savings, investment, budget_usage):

    score = 100

    # Expense ratio
    if income > 0:
        expense_ratio = (expense / income) * 100

        if expense_ratio > 90:
            score -= 30
        elif expense_ratio > 75:
            score -= 20
        elif expense_ratio > 60:
            score -= 10

    # Savings
    if savings <= 0:
        score -= 20
    elif savings < income * 0.20:
        score -= 10

    # Investment

    if investment <= 0:
        score -= 15

    # Budget

    if budget_usage > 100:
        score -= 20
    elif budget_usage > 90:
        score -= 10

    return max(0, min(100, score))



def generate_otp():

    return str(random.randint(100000, 999999))



@app.context_processor
def inject_settings():

    if "user_id" not in session:
        return {"settings": None}

    conn = sqlite3.connect("finance.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""

        SELECT *

        FROM user_settings

        WHERE user_id=?

    """, (session["user_id"],))

    settings = cursor.fetchone()

    conn.close()

    return {"settings": settings}

def generate_notifications(user_id):

    conn = sqlite3.connect("finance.db")
    cursor = conn.cursor()

    # ===============================
    # Income
    # ===============================

    cursor.execute("""

    SELECT IFNULL(SUM(amount),0)

    FROM income

    WHERE user_id=?

    """,

    (user_id,))

    income = cursor.fetchone()[0]

    # ===============================
    # Expense
    # ===============================

    cursor.execute("""

    SELECT IFNULL(SUM(amount),0)

    FROM expense

    WHERE user_id=?

    """,

    (user_id,))

    expense = cursor.fetchone()[0]

    savings = income - expense

    notifications = []

    # ===============================
    # Budget Alert
    # ===============================

    if income > 0:

        expense_ratio = (expense / income) * 100

        if expense_ratio > 80:

            notifications.append(

                (

                    "Budget Alert",

                    "⚠ Your expenses exceed 80% of your income.",

                    "danger"

                )

            )

        elif expense_ratio > 60:

            notifications.append(

                (

                    "Budget Warning",

                    "📊 Your expenses are increasing.",

                    "warning"

                )

            )

        else:

            notifications.append(

                (

                    "Excellent",

                    "✅ Your spending is well under control.",

                    "success"

                )

            )

    # ===============================
    # Savings Alert
    # ===============================

    if savings > income * 0.30:

        notifications.append(

            (

                "Savings",

                "🎉 Excellent savings habit. Keep it up!",

                "success"

            )

        )

    elif savings > income * 0.20:

        notifications.append(

            (

                "Savings",

                "👍 Good savings. Try increasing it slightly.",

                "primary"

            )

        )

    else:

        notifications.append(

            (

                "Savings",

                "💰 Try increasing your monthly savings.",

                "warning"

            )

        )

    # ===============================
    # Store Notifications
    # ===============================

    for title, message, ntype in notifications:

        # Check whether the same notification already exists
        cursor.execute("""
            SELECT id
            FROM notifications
            WHERE user_id = ?
            AND title = ?
            AND message = ?
            AND type = ?
        """, (
            user_id,
            title,
            message,
            ntype
        ))

        existing = cursor.fetchone()

        # Insert only if it does not already exist
        if existing is None:

            cursor.execute("""
                INSERT INTO notifications
                (user_id, title, message, type)
                VALUES (?, ?, ?, ?)
            """, (
                user_id,
                title,
                message,
                ntype
            ))

    # Save all changes
    conn.commit()

    conn.close()




def get_date_months_ago(date_obj, months):
    year = date_obj.year
    month = date_obj.month - months
    while month <= 0:
        month += 12
        year -= 1
    return datetime(year, month, 1)




@app.context_processor
def inject_notifications():

    if "user_id" not in session:

        return dict(
            notification_count=0,
            latest_notifications=[]
        )

    conn = sqlite3.connect("finance.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        SELECT *
        FROM notifications
        WHERE user_id=?
        ORDER BY created_at DESC
    """,(session["user_id"],))

    notifications = cursor.fetchall()

    cursor.execute("""
        SELECT COUNT(*)
        FROM notifications
        WHERE user_id=?
        AND is_read=0
    """,(session["user_id"],))

    count = cursor.fetchone()[0]

    conn.close()

    return dict(

        notification_count=count,

        latest_notifications=notifications[:5]

    )
app.secret_key = "smartfinance123"
app.permanent_session_lifetime = timedelta(days=30)

@app.route("/")
def home():
    return render_template("landing.html")



@app.route("/ai_dashboard")
def ai_dashboard():

    if "user_id" not in session:
        return redirect(url_for("login"))

    generate_notifications(session["user_id"])

    conn = sqlite3.connect("finance.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    # ======================================
    # Income
    # ======================================

    cursor.execute("""
        SELECT IFNULL(SUM(amount),0)
        FROM income
        WHERE user_id=?
    """,(session["user_id"],))

    income = cursor.fetchone()[0]

    # ======================================
    # Expense
    # ======================================

    cursor.execute("""
        SELECT IFNULL(SUM(amount),0)
        FROM expense
        WHERE user_id=?
    """,(session["user_id"],))

    expense = cursor.fetchone()[0]

    savings = income - expense

    # ======================================
    # Prediction Statistics
    # ======================================

    cursor.execute("""
        SELECT COUNT(*)
        FROM prediction_history
        WHERE user_id=?
    """,(session["user_id"],))

    total_transactions = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COUNT(*)
        FROM prediction_history
        WHERE user_id=?
        AND prediction='Fraud'
    """,(session["user_id"],))

    fraud_transactions = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COUNT(*)
        FROM prediction_history
        WHERE user_id=?
        AND prediction='Genuine'
    """,(session["user_id"],))

    genuine_transactions = cursor.fetchone()[0]

    # ======================================
    # Fraud Percentage
    # ======================================

    if total_transactions > 0:

        fraud_percentage = round(
            (fraud_transactions / total_transactions) * 100,
            2
        )

    else:

        fraud_percentage = 0

    # ======================================
    # Amount Statistics
    # ======================================

    cursor.execute("""
        SELECT
        IFNULL(AVG(amount),0),
        IFNULL(MAX(amount),0),
        IFNULL(MIN(amount),0)
        FROM prediction_history
        WHERE user_id=?
    """,(session["user_id"],))

    average_amount, highest_amount, lowest_amount = cursor.fetchone()

    # ======================================
    # Categories
    # ======================================

    cursor.execute("""
        SELECT category,COUNT(*)
        FROM prediction_history
        WHERE user_id=?
        GROUP BY category
    """,(session["user_id"],))

    rows = cursor.fetchall()

    categories = []
    category_counts = []

    for row in rows:

        categories.append(row["category"])
        category_counts.append(row["COUNT(*)"])

    # ======================================
    # States
    # ======================================

    cursor.execute("""
        SELECT state,COUNT(*)
        FROM prediction_history
        WHERE user_id=?
        GROUP BY state
    """,(session["user_id"],))

    rows = cursor.fetchall()

    states = []
    state_counts = []

    for row in rows:

        states.append(row["state"])
        state_counts.append(row["COUNT(*)"])

    # ======================================
    # Highest Category
    # ======================================

    if categories:

        highest_category = categories[
            category_counts.index(max(category_counts))
        ]

        highest_category_count = max(category_counts)

    else:

        highest_category = "-"

        highest_category_count = 0

    # ======================================
    # Highest State
    # ======================================

    if states:

        highest_state = states[
            state_counts.index(max(state_counts))
        ]

        highest_state_count = max(state_counts)

    else:

        highest_state = "-"

        highest_state_count = 0

    # ======================================
    # Smart Notifications
    # ======================================

    cursor.execute("""
        SELECT *
        FROM notifications
        WHERE user_id=?
        ORDER BY created_at DESC
        LIMIT 5
    """,(session["user_id"],))

    smart_notifications = cursor.fetchall()

    conn.close()



    print("="*50)

    print("SMART NOTIFICATIONS")

    for n in smart_notifications:
        print(dict(n))

    print("Length =", len(smart_notifications))

    print("="*50)

    return render_template(

        "ai_dashboard.html",

        income=income,
        expense=expense,
        savings=savings,

        total_transactions=total_transactions,
        fraud_transactions=fraud_transactions,
        genuine_transactions=genuine_transactions,

        fraud_percentage=fraud_percentage,

        average_amount=round(average_amount,2),
        highest_amount=highest_amount,
        lowest_amount=lowest_amount,

        categories=categories,
        category_counts=category_counts,

        states=states,
        state_counts=state_counts,

        highest_category=highest_category,
        highest_category_count=highest_category_count,

        highest_state=highest_state,
        highest_state_count=highest_state_count,

        smart_notifications=smart_notifications

    )



@app.route("/fraud-analysis")
def fraud_analysis():

    if "user_id" not in session:

        return redirect(url_for("login"))

    conn=sqlite3.connect("finance.db")

    conn.row_factory=sqlite3.Row

    cursor=conn.cursor()

    cursor.execute("""

    SELECT *

    FROM prediction_history

    WHERE user_id=?

    ORDER BY uploaded_at DESC

    """,(session["user_id"],))

    history=cursor.fetchall()

    conn.close()

    return render_template(

        "fraud_analysis.html",

        history=history

    )



@app.route("/prediction_history")
def prediction_history():

    if "user_id" not in session:
        return redirect(url_for("login"))

    from_date = request.args.get("from_date")
    to_date = request.args.get("to_date")

    conn = sqlite3.connect("finance.db")
    cursor = conn.cursor()

    # ==========================================
    # Prediction History
    # ==========================================

    if from_date and to_date:

        cursor.execute("""

        SELECT

            id,

            merchant,

            category,

            amount,

            prediction,

            confidence,

            created_at

        FROM prediction_history

        WHERE

            user_id=?

            AND DATE(created_at) BETWEEN ? AND ?

        ORDER BY id DESC

        """,

        (

            session["user_id"],

            from_date,

            to_date

        ))

    else:

        cursor.execute("""

        SELECT

            id,

            merchant,

            category,

            amount,

            prediction,

            confidence,

            created_at

        FROM prediction_history

        WHERE user_id=?

        ORDER BY id DESC

        """,

        (

            session["user_id"],

        ))

    history = cursor.fetchall()

    # ==========================================
    # Dashboard Statistics
    # ==========================================

    cursor.execute("""

    SELECT COUNT(*)

    FROM prediction_history

    WHERE user_id=?

    """,

    (session["user_id"],))

    total_predictions = cursor.fetchone()[0]

    cursor.execute("""

    SELECT COUNT(*)

    FROM prediction_history

    WHERE

        user_id=?

        AND prediction='Fraud'

    """,

    (session["user_id"],))

    fraud_count = cursor.fetchone()[0]

    genuine_count = total_predictions - fraud_count

    cursor.execute("""

    SELECT IFNULL(AVG(confidence),0)

    FROM prediction_history

    WHERE user_id=?

    """,

    (session["user_id"],))

    average_confidence = round(cursor.fetchone()[0], 2)

    # ==========================================
    # Fraud Percentage
    # ==========================================

    if total_predictions > 0:

        fraud_percentage = round(

            (fraud_count / total_predictions) * 100,

            2

        )

        genuine_percentage = round(

            (genuine_count / total_predictions) * 100,

            2

        )

    else:

        fraud_percentage = 0

        genuine_percentage = 0

    # ==========================================
    # Top Risky Merchant
    # ==========================================

    cursor.execute("""

    SELECT

        merchant,

        COUNT(*)

    FROM prediction_history

    WHERE

        user_id=?

        AND prediction='Fraud'

    GROUP BY merchant

    ORDER BY COUNT(*) DESC

    LIMIT 1

    """,

    (session["user_id"],))

    merchant = cursor.fetchone()

    if merchant:

        top_merchant = merchant[0]

    else:

        top_merchant = "No Fraud Detected"

    # ==========================================
    # Highest Confidence
    # ==========================================

    cursor.execute("""

    SELECT MAX(confidence)

    FROM prediction_history

    WHERE user_id=?

    """,

    (session["user_id"],))

    highest_confidence = cursor.fetchone()[0]

    if highest_confidence is None:

        highest_confidence = 0

    # ==========================================
    # Lowest Confidence
    # ==========================================

    cursor.execute("""

    SELECT MIN(confidence)

    FROM prediction_history

    WHERE user_id=?

    """,

    (session["user_id"],))

    lowest_confidence = cursor.fetchone()[0]

    if lowest_confidence is None:

        lowest_confidence = 0

    # ==========================================
    # Top 5 Risky Merchants
    # ==========================================

    cursor.execute("""

    SELECT

    merchant,

    COUNT(*)

    FROM prediction_history

    WHERE

    user_id=?

    AND prediction='Fraud'

    GROUP BY merchant

    ORDER BY COUNT(*) DESC

    LIMIT 5

    """,

    (session["user_id"],))

    top_merchants = cursor.fetchall()


    # ==========================================
    # Top Fraud Categories
    # ==========================================

    cursor.execute("""

    SELECT

    category,

    COUNT(*)

    FROM prediction_history

    WHERE

    user_id=?

    AND prediction='Fraud'

    GROUP BY category

    ORDER BY COUNT(*) DESC

    LIMIT 5

    """,

    (session["user_id"],))

    top_categories = cursor.fetchall()


    # ==========================================
    # Top Fraud States
    # ==========================================

    cursor.execute("""

    SELECT

    state,

    COUNT(*)

    FROM prediction_history

    WHERE

    user_id=?

    AND prediction='Fraud'

    GROUP BY state

    ORDER BY COUNT(*) DESC

    LIMIT 5

    """,

    (session["user_id"],))

    top_states = cursor.fetchall()

    # ==========================================
    # Confidence Trend Chart
    # ==========================================

    cursor.execute("""

    SELECT

        created_at,

        confidence

    FROM prediction_history

    WHERE user_id=?

    ORDER BY id

    """,

    (session["user_id"],))

    chart = cursor.fetchall()

    dates = [row[0] for row in chart]

    confidence = [row[1] for row in chart]

    conn.close()

    return render_template(

        "prediction_history.html",

        history=history,

        total_predictions=total_predictions,

        fraud_count=fraud_count,

        genuine_count=genuine_count,

        average_confidence=average_confidence,

        fraud_percentage=fraud_percentage,

        genuine_percentage=genuine_percentage,

        top_merchant=top_merchant,

        highest_confidence=highest_confidence,

        lowest_confidence=lowest_confidence,

        dates=dates,

        confidence=confidence,

        top_merchants=top_merchants,

        top_categories=top_categories,

        top_states=top_states

    )




@app.route("/budget_advisor")
def budget_advisor():

    if "user_id" not in session:
        return redirect(url_for("login"))

    conn = sqlite3.connect("finance.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    user_id = session["user_id"]

    # =====================================================
    # TOTAL INCOME
    # =====================================================

    cursor.execute("""
        SELECT IFNULL(SUM(amount),0)
        FROM income
        WHERE user_id=?
    """, (user_id,))

    total_income = cursor.fetchone()[0]

    # =====================================================
    # TOTAL EXPENSE
    # =====================================================

    cursor.execute("""
        SELECT IFNULL(SUM(amount),0)
        FROM expense
        WHERE user_id=?
    """, (user_id,))

    total_expense = cursor.fetchone()[0]

    savings = total_income - total_expense

    # =====================================================
    # CURRENT MONTHLY BUDGET
    # =====================================================

    cursor.execute("""
        SELECT monthly_budget
        FROM budget
        WHERE user_id=?
    """, (user_id,))

    row = cursor.fetchone()

    if row:
        current_budget = row["monthly_budget"]
    else:
        current_budget = 0

    # =====================================================
    # RECOMMENDED CATEGORY BUDGETS
    # =====================================================

    recommended_budget = {
        "Food": round(total_income * 0.25,2),
        "Transportation": round(total_income * 0.10,2),
        "Travel": round(total_income * 0.10,2),
        "Shopping": round(total_income * 0.10,2),
        "Entertainment": round(total_income * 0.10,2),
        "Bills": round(total_income * 0.25,2),
        "Bills & Utilities": round(total_income * 0.25,2),
        "Healthcare": round(total_income * 0.05,2),
        "Savings": round(total_income * 0.30,2)
    }

    # =====================================================
    # CATEGORY EXPENSES
    # =====================================================

    cursor.execute("""
        SELECT
            category,
            SUM(amount) AS total
        FROM expense
        WHERE user_id=?
        GROUP BY category
    """, (user_id,))

    rows = cursor.fetchall()

    current_expense = {}

    for row in rows:
        current_expense[row["category"]] = row["total"]

    # =====================================================
    # CURRENT VS RECOMMENDED TABLE
    # =====================================================

    comparison = []

    recommendations = []

    labels = []

    current_values = []

    recommended_values = []

    all_categories = sorted(
        set(current_expense.keys()) |
        set(recommended_budget.keys())
    )

    for category in all_categories:

        current = current_expense.get(category,0)

        recommended = recommended_budget.get(category,0)

        difference = current - recommended

        if category == "Savings":

            if current < recommended:

                status = "Increase Savings"

                recommendations.append(
                    f"Increase savings by ₹{recommended-current:,.2f}"
                )

            else:

                status = "Excellent"

        else:

            if current > recommended:

                status = "Over Budget"

                recommendations.append(
                    f"Reduce {category} spending by ₹{difference:,.2f}"
                )

            elif current == recommended:

                status = "On Track"

            else:

                status = "Good"

        comparison.append({

            "category": category,

            "current": current,

            "recommended": recommended,

            "difference": difference,

            "status": status

        })

        labels.append(category)

        current_values.append(current)

        recommended_values.append(recommended)

    # =====================================================
    # FINANCIAL HEALTH SCORE
    # =====================================================

    if total_income > 0:

        saving_ratio = round((savings/total_income)*100,2)

        expense_ratio = round((total_expense/total_income)*100,2)

    else:

        saving_ratio = 0

        expense_ratio = 0

    if saving_ratio >= 40:

        health_score = 95
        health = "Excellent"

    elif saving_ratio >= 25:

        health_score = 80
        health = "Good"

    elif saving_ratio >= 10:

        health_score = 65
        health = "Average"

    else:

        health_score = 40
        health = "Needs Improvement"

    # =====================================================
    # BUDGET UTILIZATION
    # =====================================================

    if current_budget > 0:

        budget_used = round(
            (total_expense/current_budget)*100,
            1
        )

    else:

        budget_used = 0

    conn.close()

    return render_template(

        "budget_advisor.html",

        total_income=total_income,

        total_expense=total_expense,

        savings=savings,

        current_budget=current_budget,

        budget_used=budget_used,

        health_score=health_score,

        health=health,

        expense_ratio=expense_ratio,

        saving_ratio=saving_ratio,

        comparison=comparison,

        recommendations=recommendations,

        labels=labels,

        current_values=current_values,

        recommended_values=recommended_values,

        food_budget=recommended_budget.get("Food",0),

        travel_budget=recommended_budget.get("Travel",recommended_budget.get("Transportation",0)),

        shopping_budget=recommended_budget.get("Shopping",0),

        bills_budget=recommended_budget.get("Bills",recommended_budget.get("Bills & Utilities",0)),

        savings_budget=recommended_budget.get("Savings",0),

    )




@app.route("/delete_prediction/<int:id>")
def delete_prediction(id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    conn = sqlite3.connect("finance.db")
    cursor = conn.cursor()

    cursor.execute("""

    DELETE FROM prediction_history

    WHERE id=?

    AND user_id=?

    """,

    (

        id,

        session["user_id"]

    ))

    conn.commit()
    conn.close()

    return redirect(url_for("prediction_history"))





@app.route("/register", methods=["GET", "POST"])
def register():

    if request.method == "POST":

        name = request.form["name"].strip()
        email = request.form["email"].strip().lower()
        password = request.form["password"]
        confirm_password = request.form["confirm_password"]

        # ---------------- Name Validation ----------------
        if not re.fullmatch(r"[A-Za-z ]{3,50}", name):
            return render_template(
                "register.html",
                error="Name should contain only letters and spaces (3-50 characters)."
            )

        # ---------------- Email Validation ----------------
        email_pattern = r"^[^\s@]+@[^\s@]+\.[^\s@]+$"

        if not re.fullmatch(email_pattern, email):
            return render_template(
                "register.html",
                error="Invalid email address."
            )

        # ---------------- Password Match ----------------
        if password != confirm_password:
            return render_template(
                "register.html",
                error="Passwords do not match."
            )

        # ---------------- Strong Password ----------------

        if len(password) < 8:
            return render_template(
                "register.html",
                error="Password must be at least 8 characters."
            )

        if not re.search(r"[A-Z]", password):
            return render_template(
                "register.html",
                error="Password must contain at least one uppercase letter."
            )

        if not re.search(r"[a-z]", password):
            return render_template(
                "register.html",
                error="Password must contain at least one lowercase letter."
            )

        if not re.search(r"\d", password):
            return render_template(
                "register.html",
                error="Password must contain at least one number."
            )

        if not re.search(r"[!@#$%^&*(),.?\":{}|<>]", password):
            return render_template(
                "register.html",
                error="Password must contain at least one special character."
            )

        # ---------------- Database ----------------

        conn = sqlite3.connect("finance.db")
        cursor = conn.cursor()

        cursor.execute(
            "SELECT id FROM users WHERE email=?",
            (email,)
        )

        existing_user = cursor.fetchone()

        if existing_user:

            conn.close()

            return render_template(
                "register.html",
                error="Email already registered. Please login."
            )

        hashed_password = generate_password_hash(password)

        cursor.execute(
            """
            INSERT INTO users(name,email,password)
            VALUES(?,?,?)
            """,
            (
                name,
                email,
                hashed_password
            )
        )
        conn.commit()
        conn.close()

        return redirect(url_for("login"))

    return render_template("register.html")



@app.route("/login", methods=["GET", "POST"])
def login():

    if request.method == "POST":

        email = request.form["email"].strip().lower()
        password = request.form["password"]

        remember = request.form.get("remember")

        # ---------------- Email Validation ----------------

        email_pattern = r"^[^\s@]+@[^\s@]+\.[^\s@]+$"

        if not re.fullmatch(email_pattern, email):

            return render_template(
                "login.html",
                error="Please enter a valid email address."
            )

        # ---------------- Database ----------------

        conn = sqlite3.connect("finance.db")
        cursor = conn.cursor()

        cursor.execute(
            "SELECT * FROM users WHERE email=?",
            (email,)
        )

        user = cursor.fetchone()

        conn.close()

        if user is None:

            return render_template(
                "login.html",
                error="Email is not registered."
            )

        stored_password = user[3]

        # If the password is already hashed
        if stored_password.startswith("pbkdf2:") or stored_password.startswith("scrypt:"):

            if not check_password_hash(stored_password, password):
                return render_template(
                    "login.html",
                    error="Incorrect password."
                )

        # Old plain-text password
        else:

            if stored_password != password:
                return render_template(
                    "login.html",
                    error="Incorrect password."
                )
        # ---------------- Remember Me ----------------

        if remember:
            session.permanent = True
        else:
            session.permanent = False

        session["user_id"] = user[0]
        session["name"] = user[1]
        session["email"] = user[2]

        return render_template(
            "login_success.html",
            name=user[1]
        )

    return render_template("login.html")





@app.route("/test_mail", methods=["GET", "POST"])
def test_mail():

    if request.method == "POST":

        email = request.form["email"]

        try:

            msg = Message(
                subject="Finance Analytics Platform - Test Email",
                recipients=[email]
            )

            msg.body = f"""
Hello,

This is a test email from Finance Analytics Platform.

If you received this email, your SMTP configuration is working perfectly.

Have a great day!

Finance Analytics Platform
"""

            mail.send(msg)

            return f"Email successfully sent to {email}"

        except Exception as e:

            return str(e)

    return """
    <form method="POST">
        <h2>SMTP Email Test</h2>

        <input
            type="email"
            name="email"
            placeholder="Enter any email address"
            required
        >

        <br><br>

        <button type="submit">
            Send Test Email
        </button>
    </form>
    """




@app.route("/forgot_password", methods=["GET", "POST"])
def forgot_password():

    if request.method == "POST":

        email = request.form["email"].strip().lower()

        conn = sqlite3.connect("finance.db")
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute(
            "SELECT * FROM users WHERE email=?",
            (email,)
        )

        user = cursor.fetchone()

        conn.close()

        if not user:

            return render_template(
                "forgot_password.html",
                error="No account found with this email."
            )

        otp = generate_otp()

        otp_storage[email] = {
            "otp": otp,
            "time": time.time()
        }

        try:

            msg = Message(
                subject="Finance Analytics Platform - Password Reset OTP",
                recipients=[email]
            )

            msg.html = f"""
            <!DOCTYPE html>

            <html>

            <body style="background:#f4f7fb;padding:40px;font-family:Arial;">

            <div style="
            max-width:600px;
            margin:auto;
            background:white;
            border-radius:15px;
            overflow:hidden;
            box-shadow:0 5px 20px rgba(0,0,0,.15);
            ">

            <div style="
            background:#0d6efd;
            padding:25px;
            text-align:center;
            color:white;
            ">

            <h1>💰 Finance Analytics Platform</h1>

            <p>Password Reset Verification</p>

            </div>

            <div style="padding:35px;">

            <h2>Hello!</h2>

            <p>

            We received a request to reset your password.

            </p>

            <p>

            Use the following OTP:

            </p>

            <div style="
            text-align:center;
            margin:30px;
            ">

            <span style="
            font-size:40px;
            font-weight:bold;
            letter-spacing:12px;
            color:#0d6efd;
            ">

            {otp}

            </span>

            </div>

            <p>

            This OTP will expire in

            <b>5 minutes.</b>

            </p>

            <p>

            If you didn't request this,

            please ignore this email.

            </p>

            <hr>

            <p style="text-align:center;color:gray;">

            © 2026 Finance Analytics Platform

            </p>

            </div>

            </div>

            </body>

            </html>
            """

            mail.send(msg)

            session["reset_email"] = email

            flash("OTP sent successfully to your registered email.", "success")

            return redirect(url_for("verify_otp"))

        except Exception as e:

            return render_template(
                "forgot_password.html",
                error=str(e)
            )

    return render_template("forgot_password.html")



@app.route("/verify_otp", methods=["GET", "POST"])
def verify_otp():

    if "reset_email" not in session:
        return redirect(url_for("forgot_password"))

    email = session["reset_email"]

    if request.method == "POST":

        entered_otp = request.form["otp"].strip()

        if email not in otp_storage:
            return render_template(
                "verify_otp.html",
                error="OTP expired. Please request a new OTP."
            )

        saved_otp = otp_storage[email]["otp"]
        if "attempts" not in otp_storage[email]:
            otp_storage[email]["attempts"] = 0

        otp_storage[email]["attempts"] += 1

        if otp_storage[email]["attempts"] > 5:

            del otp_storage[email]

            session.pop("reset_email",None)

            return render_template(
                "verify_otp.html",
                error="Maximum OTP attempts exceeded."
            )
        otp_time = otp_storage[email]["time"]

        # OTP valid for 5 minutes
        OTP_EXPIRY = 300

        if time.time() - otp_time > OTP_EXPIRY:
            del otp_storage[email]

            return render_template(
                "verify_otp.html",
                error="OTP has expired."
            )

        if entered_otp != saved_otp:

            return render_template(
                "verify_otp.html",
                error="Invalid OTP."
            )

        # OTP verified successfully
        session["otp_verified"] = True

        flash(
        "OTP Verified Successfully!",
        "success"
        )

        del otp_storage[email]

        return redirect(url_for("reset_password"))

    return render_template("verify_otp.html")



@app.route("/resend_otp")
def resend_otp():

    if "reset_email" not in session:
        return redirect(url_for("forgot_password"))

    email=session["reset_email"]

    otp=generate_otp()

    otp_storage[email]={
        "otp":otp,
        "time":time.time(),
        "attempts":0
    }

    msg=Message(
        subject="New Finance Analytics Platform OTP",
        recipients=[email]
    )

    msg.body=f"""

Your new OTP is:

{otp}

Valid for 5 minutes.

"""

    mail.send(msg)

    flash("New OTP Sent Successfully.","success")

    return redirect(url_for("verify_otp"))





@app.route("/reset_password", methods=["GET", "POST"])
def reset_password():

    if "reset_email" not in session or "otp_verified" not in session:
        return redirect(url_for("forgot_password"))

    email = session["reset_email"]

    if request.method == "POST":

        password = request.form["password"]
        import re

        if not re.search(r"[A-Z]",password):

            return render_template(
                "reset_password.html",
                error="Password must contain one uppercase letter."
            )

        if not re.search(r"[a-z]",password):

            return render_template(
                "reset_password.html",
                error="Password must contain one lowercase letter."
            )

        if not re.search(r"[0-9]",password):

            return render_template(
                "reset_password.html",
                error="Password must contain one number."
            )

        if not re.search(r"[!@#$%^&*(),.?\":{}|<>]",password):

            return render_template(
                "reset_password.html",
                error="Password must contain one special character."
            )
        confirm_password = request.form["confirm_password"]

        if password != confirm_password:

            return render_template(
                "reset_password.html",
                error="Passwords do not match."
            )

        if len(password) < 8:

            return render_template(
                "reset_password.html",
                error="Password must contain at least 8 characters."
            )

        hashed_password = generate_password_hash(password)

        conn = sqlite3.connect("finance.db")
        cursor = conn.cursor()

        cursor.execute(
            """
            UPDATE users
            SET password=?
            WHERE email=?
            """,
            (
                hashed_password,
                email
            )
        )

        conn.commit()
        conn.close()

        session.pop("reset_email", None)
        session.pop("otp_verified", None)

        flash(
            "Password reset successfully. Please login.",
            "success"
        )

        return redirect(url_for("login"))

    return render_template("reset_password.html")









# ==========================================
# SETTINGS PAGE
# ==========================================

@app.route("/settings", methods=["GET", "POST"])
def settings():

    if "user_id" not in session:
        flash("Please login first.", "warning")
        return redirect(url_for("login"))

    conn = sqlite3.connect("finance.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    user_id = session["user_id"]

    # ==========================================
    # Create default settings if not available
    # ==========================================

    cursor.execute("""
        SELECT *
        FROM user_settings
        WHERE user_id=?
    """, (user_id,))

    settings = cursor.fetchone()

    if settings is None:

        cursor.execute("""
            INSERT INTO user_settings
            (
                user_id,
                theme,
                currency,
                remember_login,
                two_factor,
                email_alert,
                ai_level,
                notifications
            )
            VALUES
            (?, 'light', 'INR', 1, 0, 1, 'Balanced', 1)
        """, (user_id,))

        conn.commit()

    # ==========================================
    # Update Profile
    # ==========================================

    if request.method == "POST":

        name = request.form.get("name")
        email = request.form.get("email")
        phone = request.form.get("phone")
        occupation = request.form.get("occupation")

        monthly_income = request.form.get("monthly_income") or 0

        date_of_birth = request.form.get("date_of_birth")
        gender = request.form.get("gender")
        address = request.form.get("address")
        city = request.form.get("city")
        state = request.form.get("state")
        country = request.form.get("country")
        pincode = request.form.get("pincode")

        theme = request.form.get("theme", "light")
        currency = request.form.get("currency", "INR")

        remember_login = 1 if request.form.get("remember_login") else 0
        two_factor = 1 if request.form.get("two_factor") else 0
        email_alert = 1 if request.form.get("email_alert") else 0
        notifications = 1 if request.form.get("notifications") else 0

        ai_level = request.form.get("ai_level", "Balanced")
        from werkzeug.utils import secure_filename
        import os
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


        photo = request.files.get("profile_photo")

        photo_path = None

        if photo and photo.filename != "":

            filename = secure_filename(photo.filename)

            upload_folder = os.path.join(
                app.root_path,
                "static",
                "uploads"
            )

            os.makedirs(upload_folder, exist_ok=True)

            filepath = os.path.join(upload_folder, filename)

            photo.save(filepath)

            photo_path = f"uploads/{filename}"

        # --------------------------
        # Update users table
        # --------------------------

        cursor.execute("""
            UPDATE users
            SET
            name=?,
            email=?,
            phone=?,
            occupation=?,
            monthly_income=?,
            profile_photo=?,
            date_of_birth=?,
            gender=?,
            address=?,
            city=?,
            state=?,
            country=?,
            pincode=?,
            updated_at=?
            WHERE id=?
        """,
        (
            (
                name,
                email,
                phone,
                occupation,
                monthly_income,
                photo_path,
                date_of_birth,
                gender,
                address,
                city,
                state,
                country,
                pincode,
                now,
                user_id
)
        ))

        # --------------------------
        # Update settings table
        # --------------------------

        cursor.execute("""
            UPDATE user_settings
            SET
                theme=?,
                currency=?,
                remember_login=?,
                two_factor=?,
                email_alert=?,
                ai_level=?,
                notifications=?
            WHERE user_id=?
        """,
        (
            theme,
            currency,
            remember_login,
            two_factor,
            email_alert,
            ai_level,
            notifications,
            user_id
        ))

        conn.commit()
        conn.close()

        flash("Settings updated successfully!", "success")

        return redirect(url_for("settings"))

    # ==========================================
    # Load User Information
    # ==========================================

    cursor.execute("""
        SELECT

            id,
            name,
            email,
            phone,
            occupation,
            monthly_income,
            profile_photo,
            date_of_birth,
            gender,
            address,
            city,
            state,
            country,
            pincode,
            created_at,
            updated_at

        FROM users

        WHERE id=?

    """, (user_id,))

    user = cursor.fetchone()

    # ==========================================
    # Load Settings
    # ==========================================

    cursor.execute("""
        SELECT *
        FROM user_settings
        WHERE user_id=?
    """, (user_id,))

    settings = cursor.fetchone()

    conn.close()

    return render_template(
        "settings.html",
        user=user,
        settings=settings
    )






@app.route("/save_theme",methods=["POST"])
def save_theme():

    if "user_id" not in session:

        return jsonify({"success":False})

    data=request.get_json()

    theme=data.get("theme","light")

    conn=sqlite3.connect("finance.db")

    cursor=conn.cursor()

    cursor.execute("""

    UPDATE user_settings

    SET theme=?

    WHERE user_id=?

    """,(theme,session["user_id"]))

    conn.commit()

    conn.close()

    return jsonify({"success":True})



@app.route("/logout")
def logout():

    session.clear()

    return redirect(url_for("home"))



@app.route("/savings_tips")
def savings_tips():

    if "user_id" not in session:
        return redirect(url_for("login"))

    conn = sqlite3.connect("finance.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    # ===============================
    # Total Income
    # ===============================

    cursor.execute("""
        SELECT IFNULL(SUM(amount),0)
        FROM income
        WHERE user_id=?
    """,(session["user_id"],))

    income = cursor.fetchone()[0]

    # ===============================
    # Total Expense
    # ===============================

    cursor.execute("""
        SELECT IFNULL(SUM(amount),0)
        FROM expense
        WHERE user_id=?
    """,(session["user_id"],))

    expense = cursor.fetchone()[0]

    savings = income - expense

    # ===============================
    # Savings Rate
    # ===============================

    if income > 0:

        savings_rate = round(
            (savings / income) * 100,
            1
        )

    else:

        savings_rate = 0

    # ===============================
    # AI Tip
    # ===============================

    if savings_rate >= 40:

        ai_tip = "🎉 Excellent! Your savings habit is outstanding. Keep investing consistently."

    elif savings_rate >= 20:

        ai_tip = "👍 Good savings! Try increasing your monthly savings by another 10%."

    else:

        ai_tip = "⚠ Your savings are low. Reduce unnecessary spending and save before spending."

    conn.close()

    return render_template(

        "savings_tips.html",

        income=income,

        expense=expense,

        savings=savings,

        savings_rate=savings_rate,

        ai_tip=ai_tip

    )



@app.route("/profile")
def profile():

    if "user_id" not in session:
        return redirect(url_for("login"))
    
    conn = sqlite3.connect("finance.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute("PRAGMA table_info(investments)")
    print("INVESTMENTS SCHEMA:")
    for r in cursor.fetchall():
        print(dict(r))

    cursor.execute("PRAGMA table_info(budget)")
    print("BUDGET SCHEMA:")
    for r in cursor.fetchall():
        print(dict(r))
    user_id = session["user_id"]

    # ==========================================
    # USER DETAILS
    # ==========================================

    cursor.execute(
        "SELECT * FROM users WHERE id=?",
        (user_id,)
    )
    user = cursor.fetchone()

    # ==========================================
    # TOTAL INCOME
    # ==========================================

    cursor.execute(
        """
        SELECT IFNULL(SUM(amount),0)
        FROM income
        WHERE user_id=?
        """,
        (user_id,)
    )
    total_income = cursor.fetchone()[0]

    # ==========================================
    # TOTAL EXPENSE
    # ==========================================

    cursor.execute(
        """
        SELECT IFNULL(SUM(amount),0)
        FROM expense
        WHERE user_id=?
        """,
        (user_id,)
    )
    total_expense = cursor.fetchone()[0]

    # ==========================================
    # TOTAL INVESTMENT
    # ==========================================

    try:

        print("Session User ID:", user_id)

        cursor.execute("""
            SELECT IFNULL(SUM(current_value),0)
            FROM investments
            WHERE user_id=?
        """, (user_id,))

        total_investment = cursor.fetchone()[0]

        print("Investment Total:", total_investment)

    except Exception as e:

        print("Investment Error:", e)

        total_investment = 0

    # ==========================================
    # GOAL (Financial Goals)
    # ==========================================

    try:

        cursor.execute("""
            SELECT
                IFNULL(SUM(target_amount),0) AS target_amount,
                IFNULL(SUM(saved_amount),0) AS saved_amount
            FROM financial_goals
            WHERE user_id=?
        """, (user_id,))

        goal = cursor.fetchone()

        goal_amount = goal["target_amount"] or 0
        current_goal = goal["saved_amount"] or 0

        if goal_amount > 0:
            goal_percentage = round(
                (current_goal / goal_amount) * 100,
                1
            )
        else:
            goal_percentage = 0

    except Exception as e:

        print("Goal Error:", e)
        goal_amount = 0
        goal_percentage = 0


    # ==========================================
    # MONTHLY BUDGET
    # ==========================================

    monthly_budget = 0

    try:

        cursor.execute("""
            SELECT monthly_budget
            FROM budget
            WHERE user_id=?
            ORDER BY id DESC
            LIMIT 1
        """, (user_id,))

        budget = cursor.fetchone()

        if budget:
            monthly_budget = budget["monthly_budget"]

    except Exception as e:

        print("Budget Error:", e)

    # ==========================================
    # SAVINGS
    # ==========================================

    total_savings = total_income - total_expense

    # ==========================================
    # HEALTH SCORE
    # ==========================================

    if total_income > 0:

        health_score = round(
            (total_savings / total_income) * 100
        )

        health_score = max(0, min(100, health_score))

    else:

        health_score = 0

    # ==========================================
    # LAST INCOME
    # ==========================================

    cursor.execute(
        """
        SELECT amount
        FROM income
        WHERE user_id=?
        ORDER BY id DESC
        LIMIT 1
        """,
        (user_id,)
    )

    row = cursor.fetchone()

    last_income = row["amount"] if row else 0

    # ==========================================
    # LAST EXPENSE
    # ==========================================

    cursor.execute(
        """
        SELECT amount
        FROM expense
        WHERE user_id=?
        ORDER BY id DESC
        LIMIT 1
        """,
        (user_id,)
    )

    row = cursor.fetchone()

    last_expense = row["amount"] if row else 0

    # ==========================================
    # MONTHLY CHART
    # ==========================================

    month_labels = [
        "Jan",
        "Feb",
        "Mar",
        "Apr",
        "May",
        "Jun"
    ]

    month_values = [
        0,
        0,
        0,
        0,
        0,
        total_expense
    ]

    # ==========================================
    # CATEGORY CHART
    # ==========================================

    try:

        cursor.execute("""
            SELECT category,
            SUM(amount)
            FROM expense
            WHERE user_id=?
            GROUP BY category
        """, (user_id,))

        rows = cursor.fetchall()

        category_labels = []
        category_values = []

        for row in rows:

            category_labels.append(row["category"])
            category_values.append(row["SUM(amount)"])

    except:

        category_labels = []
        category_values = []

    conn.close()

    # ==========================================
    # RENDER TEMPLATE
    # ==========================================

    return render_template(

        "profile.html",

        user=user,

        total_income=total_income,

        total_expense=total_expense,

        total_savings=total_savings,

        total_investment=total_investment,

        monthly_budget=monthly_budget,

        goal_amount=goal_amount,

        goal_percentage=goal_percentage,

        health_score=health_score,

        last_income=last_income,

        last_expense=last_expense,

        month_labels=month_labels,

        month_values=month_values,

        category_labels=category_labels,

        category_values=category_values,

        current_year=datetime.now().year

    )



@app.route("/dashboard")
def dashboard():

    if "user_id" not in session:
        return redirect(url_for("login"))

    # Generate notifications for the logged-in user
    generate_notifications(session["user_id"])

    conn = sqlite3.connect("finance.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    # ===============================
    # TOTAL INCOME
    # ===============================

    cursor.execute("""
        SELECT IFNULL(SUM(amount),0)
        FROM income
        WHERE user_id=?
    """, (session["user_id"],))

    total_income = cursor.fetchone()[0]

    # ===============================
    # TOTAL EXPENSE
    # ===============================

    cursor.execute("""
        SELECT IFNULL(SUM(amount),0)
        FROM expense
        WHERE user_id=?
    """, (session["user_id"],))

    total_expense = cursor.fetchone()[0]

    # ===============================
    # SAVINGS
    # ===============================

    savings = total_income - total_expense

    # ===============================
    # MONTHLY BUDGET
    # ===============================

    cursor.execute("""
        SELECT monthly_budget
        FROM budget
        WHERE user_id=?
    """, (session["user_id"],))

    budget_row = cursor.fetchone()

    if budget_row:
        monthly_budget = budget_row[0]
    else:
        monthly_budget = 0

    # ===============================
    # REMAINING BUDGET
    # ===============================

    remaining_budget = monthly_budget - total_expense

    if remaining_budget < 0:
        remaining_budget = 0

    # ===============================
    # BUDGET USED %
    # ===============================

    if monthly_budget > 0:

        budget_used = min(

            round(

                (total_expense / monthly_budget) * 100,

                1

            ),

            100

        )

    else:

        budget_used = 0

    # ===============================
    # EXPENSE CATEGORY DATA
    # ===============================

    cursor.execute("""
        SELECT category,
               SUM(amount)
        FROM expense
        WHERE user_id=?
        GROUP BY category
    """, (session["user_id"],))

    expense_data = cursor.fetchall()

    categories = [row[0] for row in expense_data]
    category_amounts = [row[1] for row in expense_data]

    # ===============================
    # AI SPENDING ANALYSIS
    # ===============================

    highest_category = "No Expenses"
    highest_amount = 0

    if expense_data:

        highest = max(
            expense_data,
            key=lambda x: x[1]
        )

        highest_category = highest[0]
        highest_amount = highest[1]

    # Savings Rate

    if total_income > 0:

        savings_rate = round(
            (savings / total_income) * 100,
            1
        )

    else:

        savings_rate = 0

    # Financial Status

    if savings_rate >= 50:

        financial_status = "Excellent"

    elif savings_rate >= 30:

        financial_status = "Good"

    elif savings_rate >= 10:

        financial_status = "Average"

    else:

        financial_status = "Needs Improvement"

    # AI Recommendation

    if budget_used >= 100:

        ai_message = (
            "⚠️ Budget exceeded. "
            "Reduce unnecessary spending."
        )

    elif highest_category == "Food":

        ai_message = (
            "🍔 Food is your highest expense. "
            "Consider reducing dining expenses."
        )

    elif highest_category == "Shopping":

        ai_message = (
            "🛍️ Shopping expenses are high. "
            "Try limiting impulse purchases."
        )

    elif highest_category == "Travel":

        ai_message = (
            "🚗 Travel expenses are increasing. "
            "Plan transportation wisely."
        )

    else:

        ai_message = (
            "🎉 Your financial habits look healthy. "
            "Keep saving!"
        )

    # ===============================
    # RECENT TRANSACTIONS
    # ===============================

    cursor.execute("""

    SELECT source AS title,
        amount,
        date,
        'Income'

    FROM income

    WHERE user_id=?

    UNION ALL

    SELECT category,
        amount,
        date,
        'Expense'

    FROM expense

    WHERE user_id=?

    ORDER BY date DESC

    LIMIT 5

    """,
    (
        session["user_id"],
        session["user_id"]
    ))

    transactions = cursor.fetchall()
    # ===============================
    # MONTHLY REPORT
    # ===============================

    from datetime import datetime

    current_month = datetime.now().strftime("%Y-%m")

    # Monthly Income

    cursor.execute("""

        SELECT IFNULL(SUM(amount),0)

        FROM income

        WHERE user_id=?

        AND substr(date,1,7)=?

    """,

    (

        session["user_id"],
        current_month

    ))

    monthly_income = cursor.fetchone()[0]

    # Monthly Expense

    cursor.execute("""

        SELECT IFNULL(SUM(amount),0)

        FROM expense

        WHERE user_id=?

        AND substr(date,1,7)=?

    """,

    (

        session["user_id"],
        current_month

    ))

    monthly_expense = cursor.fetchone()[0]

    monthly_savings = monthly_income - monthly_expense

    if monthly_income > 0:

        monthly_saving_rate = round(

            (monthly_savings / monthly_income) * 100,

            1

        )

    else:

        monthly_saving_rate = 0

    # ===============================
    # LAST 6 MONTHS TREND
    # ===============================

    months = []

    income_trend = []

    expense_trend = []

    saving_trend = []

    today = datetime.today()

    for i in range(5, -1, -1):

        month = get_date_months_ago(today, i)

        month_key = month.strftime("%Y-%m")

        month_name = month.strftime("%b")

        months.append(month_name)

        # Income

        cursor.execute("""

            SELECT IFNULL(SUM(amount),0)

            FROM income

            WHERE user_id=?

            AND substr(date,1,7)=?

        """,

        (

            session["user_id"],
            month_key

        ))

        income = cursor.fetchone()[0]

        # Expense

        cursor.execute("""

            SELECT IFNULL(SUM(amount),0)

            FROM expense

            WHERE user_id=?

            AND substr(date,1,7)=?

        """,

        (

            session["user_id"],
            month_key

        ))

        expense = cursor.fetchone()[0]

        income_trend.append(income)

        expense_trend.append(expense)

        saving_trend.append(income - expense)

    # ===============================
    # AI FINANCIAL PREDICTION
    # ===============================

    if income_trend:

        predicted_income = round(

            sum(income_trend) / len(income_trend),

            2

        )

    else:

        predicted_income = 0

    if expense_trend:

        predicted_expense = round(

            sum(expense_trend) / len(expense_trend),

            2

        )

    else:

        predicted_expense = 0

    predicted_savings = predicted_income - predicted_expense

    if predicted_income > 0:

        predicted_saving_rate = round(

            (predicted_savings / predicted_income) * 100,

            1

        )

    else:

        predicted_saving_rate = 0

    # ===============================
    # CONTINUE TO PART 3
    # ===============================

        # ===============================
    # FINANCIAL HEALTH SCORE
    # ===============================

    health_score = 100

    if savings_rate < 50:
        health_score -= 20

    if budget_used > 80:
        health_score -= 20

    if budget_used >= 100:
        health_score -= 20

    if total_income == 0:
        health_score = 0

    if health_score < 0:
        health_score = 0

    # Health Grade

    if health_score >= 90:
        health_grade = "Excellent"
        health_color = "success"

    elif health_score >= 75:
        health_grade = "Good"
        health_color = "primary"

    elif health_score >= 60:
        health_grade = "Average"
        health_color = "warning"

    else:
        health_grade = "Poor"
        health_color = "danger"

    # ===============================
    # SMART NOTIFICATIONS
    # ===============================

    """notifications = []

    if budget_used >= 100:

        notifications.append(
            "🚨 You have exceeded your monthly budget."
        )

    elif budget_used >= 80:

        notifications.append(
            "⚠️ You have used more than 80% of your monthly budget."
        )

    if savings_rate >= 50:

        notifications.append(
            "🎉 Excellent savings rate this month!"
        )

    if highest_category != "No Expenses":

        notifications.append(
            f"💸 Highest spending category: {highest_category}"
        )

    if len(notifications) == 0:

        notifications.append(
            "✅ Everything looks good. Keep tracking your finances."
        )"""



    # ==========================================
    # Notification Count
    # ==========================================

    cursor.execute("""

    SELECT COUNT(*)

    FROM notifications

    WHERE user_id=?

    AND is_read=0

    """,

    (session["user_id"],))

    notification_count = cursor.fetchone()[0]

    # ==========================================
    # Latest Notifications
    # ==========================================

    cursor.execute("""

    SELECT *

    FROM notifications

    WHERE user_id=?

    ORDER BY created_at DESC

    LIMIT 5

    """,

    (session["user_id"],))

    latest_notifications = cursor.fetchall()


    # ============================================================
    # GOAL SUMMARY DASHBOARD ANALYTICS
    # ============================================================

    cursor.execute("""
        SELECT
            target_amount,
            saved_amount,
            target_date
        FROM financial_goals
        WHERE user_id = ?
    """, (session["user_id"],))

    dashboard_goals = cursor.fetchall()

    total_goals = len(dashboard_goals)

    goals_on_track = 0
    goals_behind = 0

    total_goal_target = 0
    total_goal_saved = 0

    for goal in dashboard_goals:

        target_amount = float(goal[0] or 0)
        saved_amount = float(goal[1] or 0)

        total_goal_target += target_amount
        total_goal_saved += saved_amount

        if target_amount > 0:

            goal_progress = (
                saved_amount / target_amount
            ) * 100

            if goal_progress >= 50:
                goals_on_track += 1

            else:
                goals_behind += 1


    # Overall Goal Progress

    overall_goal_progress = 0

    if total_goal_target > 0:

        overall_goal_progress = round(
            (
                total_goal_saved /
                total_goal_target
            ) * 100,
            1
        )

    overall_goal_progress = min(
        overall_goal_progress,
        100
    )


    # ============================================================
    # PORTFOLIO RISK ANALYSIS
    # ============================================================

    cursor.execute("""
        SELECT
            asset_type,
            invested_amount
        FROM investments
    """)

    risk_investments = cursor.fetchall()

    asset_totals = {}

    total_portfolio_investment = 0


    for investment in risk_investments:

        asset_type = investment[0]

        invested_amount = float(
            investment[1] or 0
        )

        total_portfolio_investment += invested_amount

        if asset_type in asset_totals:

            asset_totals[asset_type] += invested_amount

        else:

            asset_totals[asset_type] = invested_amount


    # Number of asset categories

    asset_category_count = len(asset_totals)


    # ============================================================
    # DIVERSIFICATION SCORE
    # ============================================================

    diversification_score = 0


    if total_portfolio_investment > 0:

        concentration_score = 0

        for amount in asset_totals.values():

            allocation = (
                amount /
                total_portfolio_investment
            )

            concentration_score += allocation ** 2


        diversification_score = round(
            (1 - concentration_score) * 100
        )


        # Category diversification bonus

        category_bonus = min(
            asset_category_count * 8,
            30
        )


        diversification_score += category_bonus


        diversification_score = min(
            diversification_score,
            100
        )


    # ============================================================
    # RISK LEVEL CALCULATION
    # ============================================================

    high_risk_assets = [
        "Stocks",
        "Cryptocurrency"
    ]

    medium_risk_assets = [
        "Mutual Funds",
        "Real Estate"
    ]

    low_risk_assets = [
        "Fixed Deposit",
        "Gold",
        "Bonds"
    ]


    risk_points = 0


    for asset_type, amount in asset_totals.items():

        if total_portfolio_investment > 0:

            allocation_percentage = (
                amount /
                total_portfolio_investment
            ) * 100


            if asset_type in high_risk_assets:

                risk_points += (
                    allocation_percentage * 1
                )


            elif asset_type in medium_risk_assets:

                risk_points += (
                    allocation_percentage * 0.6
                )


            elif asset_type in low_risk_assets:

                risk_points += (
                    allocation_percentage * 0.3
                )


    # Risk Classification

    if risk_points >= 70:

        portfolio_risk_level = "High"

        risk_class = "risk-high"


    elif risk_points >= 40:

        portfolio_risk_level = "Moderate"

        risk_class = "risk-moderate"


    else:

        portfolio_risk_level = "Low"

        risk_class = "risk-low"



    # ===============================
    # DATABASE CLOSE
    # ===============================

    conn.close()


    # ===============================
    # RENDER TEMPLATE
    # ===============================

    return render_template(

        "dashboard.html",

        name=session["name"],

        total_income=total_income,

        total_expense=total_expense,

        savings=savings,

        monthly_budget=monthly_budget,

        remaining_budget=remaining_budget,

        budget_used=budget_used,

        transactions=transactions,

        categories=categories,

        category_amounts=category_amounts,

        highest_category=highest_category,

        highest_amount=highest_amount,

        savings_rate=savings_rate,

        financial_status=financial_status,

        ai_message=ai_message,

        monthly_income=monthly_income,

        monthly_expense=monthly_expense,

        monthly_savings=monthly_savings,

        monthly_saving_rate=monthly_saving_rate,

        months=months,

        income_trend=income_trend,

        expense_trend=expense_trend,

        saving_trend=saving_trend,

        predicted_income=predicted_income,

        predicted_expense=predicted_expense,

        predicted_savings=predicted_savings,

        predicted_saving_rate=predicted_saving_rate,

        health_score=health_score,

        health_grade=health_grade,

        health_color=health_color,

        notification_count=notification_count,

        latest_notifications=latest_notifications,

        total_goals=total_goals,

        goals_on_track=goals_on_track,

        goals_behind=goals_behind,

        overall_goal_progress=overall_goal_progress,

        portfolio_risk_level=portfolio_risk_level,

        risk_class=risk_class,

        diversification_score=diversification_score
    )


@app.route("/test_user")
def test_user():
    return str(session.get("user_id"))


def create_notification(user_id, title, message,
                        notification_type="General",
                        priority="Medium",
                        status="ACTIVE"):

    conn = sqlite3.connect("finance.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO notifications
        (
            user_id,
            title,
            message,
            type,
            is_read,
            created_at,
            priority,
            status
        )
        VALUES (?,?,?,?,?,?,?,?)
    """, (
        user_id,
        title,
        message,
        notification_type,
        0,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        priority,
        status
    ))

    conn.commit()
    conn.close()


@app.route('/mark_notification_read/<int:id>')
def mark_notification_read(id):

    conn = sqlite3.connect("finance.db")
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE notifications
        SET is_read=1
        WHERE id=?
    """,(id,))

    conn.commit()
    conn.close()

    flash("Notification marked as read.","success")

    return redirect(url_for("notification_center"))


@app.route('/mark_all_notifications_read')
def mark_all_notifications_read():

    user_id=session["user_id"]

    conn=sqlite3.connect("finance.db")
    cursor=conn.cursor()

    cursor.execute("""
        UPDATE notifications
        SET is_read=1
        WHERE user_id=?
    """,(user_id,))

    conn.commit()
    conn.close()

    flash("All notifications marked as read.","success")

    return redirect(url_for("notification_center"))


@app.route('/delete_notification/<int:id>')
def delete_notification(id):

    conn=sqlite3.connect("finance.db")
    cursor=conn.cursor()

    cursor.execute("""
        DELETE FROM notifications
        WHERE id=?
    """,(id,))

    conn.commit()
    conn.close()

    flash("Notification deleted.","success")

    return redirect(url_for("notification_center"))

@app.route('/notification-filter')
def notification_filter():

    user_id=session["user_id"]

    notification_type=request.args.get("type","All")

    conn=sqlite3.connect("finance.db")
    conn.row_factory=sqlite3.Row

    cursor=conn.cursor()

    if notification_type=="All":

        cursor.execute("""
            SELECT *
            FROM notifications
            WHERE user_id=?
            ORDER BY created_at DESC
        """,(user_id,))

    else:

        cursor.execute("""
            SELECT *
            FROM notifications
            WHERE user_id=?
            AND type=?
            ORDER BY created_at DESC
        """,(user_id,notification_type))

    notifications=cursor.fetchall()

    conn.close()

    return render_template(
        "notification.html",
        notifications=notifications
    )


# ============================================================
# MODULE 3 - FINANCIAL GOAL PLANNING
# ============================================================


@app.route("/financial_goals")
def financial_goals():

    if "user_id" not in session:
        return redirect(url_for("login"))

    # Get goals only for logged-in user

    financial_goals = get_all_financial_goals(session["user_id"])

    # ========================================================
    # TOTAL NUMBER OF GOALS
    # ========================================================

    total_goals = len(financial_goals)

    # ========================================================
    # TOTAL TARGET AMOUNT
    # ========================================================

    total_target = sum(

        goal[4]

        for goal in financial_goals

    )


    # ========================================================
    # TOTAL SAVED AMOUNT
    # ========================================================

    total_saved = sum(

        financial_goal[5]

        for financial_goal in financial_goals

    )


    # ========================================================
    # ACHIEVED GOALS
    # ========================================================

    goals_achieved = sum(

        1

        for goal in financial_goals

        if goal[5] >= goal[4]

    )


    # ========================================================
    # OVERALL PROGRESS
    # ========================================================

    if total_target > 0:

        overall_progress = round(

            (
                total_saved /
                total_target
            ) * 100,

            2

        )

        overall_progress = min(

            overall_progress,

            100

        )

    else:

        overall_progress = 0


    # ========================================================
    # REMAINING SAVINGS
    # ========================================================

    remaining_savings = max(

        total_target - total_saved,

        0

    )


    # ========================================================
    # ACTIVE GOALS
    # ========================================================

    active_goals = sum(

        1

        for goal in financial_goals

        if goal[5] < goal[4]

    )


    return render_template(
        "financial_goals.html",
        financial_goals=financial_goals,
        total_goals=total_goals,
        total_target=total_target,
        total_saved=total_saved,
        goals_achieved=goals_achieved,
        active_goals=active_goals,
        overall_progress=overall_progress,
        remaining_savings=remaining_savings
    )

# ============================================================
# CREATE FINANCIAL GOAL
# ============================================================


@app.route(
    "/add-financial-goal",
    methods=["POST"]
)
def create_financial_goal():

    if "user_id" not in session:
        return redirect(url_for("login"))

    try:

        goal_type = request.form["goal_type"]

        goal_name = request.form["goal_name"]

        target_amount = float(
            request.form["target_amount"]
        )

        saved_amount = float(

            request.form.get(
                "saved_amount",
                0
            ) or 0

        )

        target_date = request.form["target_date"]


        # ====================================================
        # VALIDATION
        # ====================================================

        if target_amount <= 0:

            flash(

                "Target amount must be greater than zero.",

                "danger"

            )

            return redirect(
                url_for("financial_goals")
            )


        if saved_amount < 0:

            flash(

                "Saved amount cannot be negative.",

                "danger"

            )

            return redirect(
                url_for("financial_goals")
            )


        # ====================================================
        # SAVE GOAL
        # ====================================================

        add_financial_goal(

            session["user_id"],

            goal_type,

            goal_name,

            target_amount,

            saved_amount,

            target_date

        )


        flash(

            "🎯 Financial Goal Created Successfully!",

            "success"

        )
        create_notification(
        session["user_id"],
        "Goal Created",
        f"Financial goal '{goal_name}' has been created.",
        "Goal",
        "Low"
    )

    except ValueError:

        flash(

            "Please enter valid amount values.",

            "danger"

        )


    return redirect(
        url_for("financial_goals")
    )


# ============================================================
# ADD SAVINGS TO GOAL
# ============================================================


@app.route(
    "/add-goal-savings/<int:id>",
    methods=["POST"]
)
def add_savings_to_goal(id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    try:

        savings_amount = float(

            request.form["savings_amount"]

        )


        if savings_amount <= 0:

            flash(

                "Savings amount must be greater than zero.",

                "danger"

            )

            return redirect(
                url_for("financial_goals")
            )


        conn = sqlite3.connect("finance.db")
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute(
            """
            SELECT goal_name
            FROM financial_goals
            WHERE id=? AND user_id=?
            """,
            (id, session["user_id"])
        )

        goal_row = cursor.fetchone()
        conn.close()

        goal_name = goal_row["goal_name"] if goal_row else "your goal"

        add_goal_savings(

            id,

            session["user_id"],

            savings_amount

        )

        create_notification(
            session["user_id"],
            "Goal Achieved 🎉",
            f"Congratulations! You achieved your goal '{goal_name}'.",
            "Goal",
            "High",
            "COMPLETED"
        )

        flash(

            "💰 Savings Added Successfully!",

            "success"

        )


    except ValueError:

        flash(

            "Please enter a valid savings amount.",

            "danger"

        )


    return redirect(
        url_for("financial_goals")
    )


# ============================================================
# EDIT FINANCIAL GOAL
# ============================================================


@app.route(
    "/edit-financial-goal/<int:id>",
    methods=["POST"]
)
def edit_financial_goal(id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    try:

        goal_type = request.form["goal_type"]

        goal_name = request.form["goal_name"]

        target_amount = float(

            request.form["target_amount"]

        )

        saved_amount = float(

            request.form["saved_amount"]

        )

        target_date = request.form["target_date"]


        # ====================================================
        # VALIDATION
        # ====================================================

        if target_amount <= 0:

            flash(

                "Target amount must be greater than zero.",

                "danger"

            )

            return redirect(
                url_for("financial_goals")
            )


        if saved_amount < 0:

            flash(

                "Saved amount cannot be negative.",

                "danger"

            )

            return redirect(
                url_for("financial_goals")
            )


        # ====================================================
        # UPDATE GOAL
        # ====================================================

        update_financial_goal(

            id,

            session["user_id"],

            goal_type,

            goal_name,

            target_amount,

            saved_amount,

            target_date

        )


        flash(

            "✏️ Financial Goal Updated Successfully!",

            "success"

        )


    except ValueError:

        flash(

            "Please enter valid amount values.",

            "danger"

        )


    return redirect(
        url_for("financial_goals")
    )


# ============================================================
# DELETE FINANCIAL GOAL
# ============================================================


@app.route(
    "/delete-financial-goal/<int:id>"
)
def remove_financial_goal(id):

    if "user_id" not in session:
        return redirect(url_for("login"))


    delete_financial_goal(

        id,

        session["user_id"]

    )


    flash(

        "🗑️ Financial Goal Deleted Successfully!",

        "danger"

    )


    return redirect(
        url_for("financial_goals")
    )



@app.route("/income", methods=["GET","POST"])
def income():

    if "user_id" not in session:
        return redirect(url_for("login"))

    if request.method=="POST":

        source = request.form["source"]
        amount = float(request.form["amount"])
        date = request.form["date"]
        conn=sqlite3.connect("finance.db")

        cursor=conn.cursor()

        cursor.execute("""

        INSERT INTO income

        (user_id,source,amount,date)

        VALUES(?,?,?,?)

        """,

        (

        session["user_id"],

        source,

        amount,

        date

        ))

        conn.commit()

        create_notification(
        session["user_id"],
        "Income Added",
        f"₹{amount} income added successfully under '{source}'.",
        "Income",
        "Low"
    )

        conn.close()


        return redirect(url_for("dashboard"))

    return render_template("income.html")


@app.route("/expense", methods=["GET", "POST"])
def expense():

    if "user_id" not in session:
        return redirect(url_for("login"))

    if request.method == "POST":

        category = request.form["category"]
        amount = float(request.form["amount"])
        description = request.form["description"]
        date = request.form["date"]

        conn = sqlite3.connect("finance.db")
        cursor = conn.cursor()

        # ---------------------------------
        # Insert Expense
        # ---------------------------------

        cursor.execute("""
            INSERT INTO expense
            (user_id, category, amount, description, date)
            VALUES (?, ?, ?, ?, ?)
        """, (
            session["user_id"],
            category,
            amount,
            description,
            date
        ))

        conn.commit()

        # ---------------------------------
        # Expense Added Notification
        # ---------------------------------

        create_notification(
        session["user_id"],
        "Expense Recorded",
        f"₹{amount} spent on '{category}'.",
        "Expense",
        "Medium"
    )

        # ---------------------------------
        # Shopping Alert
        # ---------------------------------

        if category.lower() == "shopping":

            add_notification(

                session["user_id"],

                "Shopping Alert",

                "🛍 Shopping expenses are increasing. Avoid impulse buying.",

                "warning"

            )

        # ---------------------------------
        # Food Alert
        # ---------------------------------

        elif category.lower() == "food":

            add_notification(

                session["user_id"],

                "Food Spending",

                "🍔 Food expenses are increasing this month.",

                "primary"

            )

        # ---------------------------------
        # Entertainment Alert
        # ---------------------------------

        elif category.lower() == "entertainment":

            add_notification(

                session["user_id"],

                "Entertainment",

                "🎬 Entertainment spending is getting high.",

                "warning"

            )

        # ---------------------------------
        # Travel Alert
        # ---------------------------------

        elif category.lower() == "travel":

            add_notification(

                session["user_id"],

                "Travel Expense",

                "✈ Travel expenses have increased this month.",

                "primary"

            )

        # ---------------------------------
        # Medical Alert
        # ---------------------------------

        elif category.lower() == "medical":

            add_notification(

                session["user_id"],

                "Medical Expense",

                "💊 Medical expenses recorded successfully.",

                "success"

            )

        # ---------------------------------
        # Fetch Current Budget
        # ---------------------------------

        cursor.execute("""
            SELECT monthly_budget
            FROM budget
            WHERE user_id=?
            ORDER BY id DESC
            LIMIT 1
        """,(session["user_id"],))

        budget = cursor.fetchone()

        # ---------------------------------
        # Budget Analysis
        # ---------------------------------

        if budget:

            budget_amount = float(budget[0])

            cursor.execute("""
                SELECT IFNULL(SUM(amount),0)
                FROM expense
                WHERE user_id=?
            """, (session["user_id"],))

            total_expense = float(cursor.fetchone()[0])

            usage = (total_expense / budget_amount) * 100

            # 80% Warning

        if usage >= 90 and usage < 100:

            create_notification(
                session["user_id"],
                "Budget Alert",
                f"You have used {usage:.1f}% of your monthly budget.",
                "Budget Alert",
                "High"
            )

            # Budget Exceeded

        if usage >= 100:

            create_notification(
                session["user_id"],
                "Budget Exceeded",
                "You have exceeded your monthly budget.",
                "Budget Alert",
                "High"
            )

        conn.close()

        return redirect(url_for("dashboard"))

    from datetime import date

    return render_template(

        "expense.html",

        today=date.today().isoformat()

    )


@app.route("/income_list")
def income_list():

    if "user_id" not in session:
        return redirect(url_for("login"))

    conn=sqlite3.connect("finance.db")

    cursor=conn.cursor()

    cursor.execute(

    "SELECT * FROM income WHERE user_id=?",

    (

    session["user_id"],

    )

    )

    incomes=cursor.fetchall()

    conn.close()

    return render_template(

    "income_list.html",

    incomes=incomes

    )


@app.route("/edit_income/<int:id>", methods=["GET", "POST"])
def edit_income(id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    conn = sqlite3.connect("finance.db")
    cursor = conn.cursor()

    if request.method == "POST":

        source = request.form["source"]
        amount = request.form["amount"]
        date = request.form["date"]

        cursor.execute("""

        UPDATE income

        SET

        source=?,
        amount=?,
        date=?

        WHERE id=? AND user_id=?

        """,

        (

        source,
        amount,
        date,
        id,
        session["user_id"]

        ))

        conn.commit()

        conn.close()

        return redirect(url_for("income_list"))

    cursor.execute("""

    SELECT *

    FROM income

    WHERE id=? AND user_id=?

    """,

    (

    id,
    session["user_id"]

    ))

    income = cursor.fetchone()

    conn.close()

    return render_template(

        "edit_income.html",

        income=income

    )




@app.route("/expense_list")
def expense_list():

    if "user_id" not in session:
        return redirect(url_for("login"))

    conn = sqlite3.connect("finance.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    cursor.execute(

    "SELECT * FROM expense WHERE user_id=?",

    (

    session["user_id"],

    )

    )

    expenses=cursor.fetchall()

    conn.close()

    return render_template(

    "expense_list.html",

    expenses=expenses

    )

@app.route("/delete_income/<int:id>")
def delete_income(id):
    
    if "user_id" not in session:
        return redirect(url_for("login"))

    conn=sqlite3.connect("finance.db")

    cursor=conn.cursor()

    cursor.execute("""

    DELETE FROM income

    WHERE id=?

    AND user_id=?

    """,

    (

    id,

    session["user_id"]

    ))

    conn.commit()

    conn.close()

    return redirect(url_for("income_list"))



@app.route("/delete_expense/<int:id>")
def delete_expense(id):
    if "user_id" not in session:
        return redirect(url_for("login"))
    conn=sqlite3.connect("finance.db")

    cursor=conn.cursor()

    cursor.execute("""

    DELETE FROM expense

    WHERE id=?

    AND user_id=?

    """,

    (

    id,

    session["user_id"]

    ))

    conn.commit()

    conn.close()

    return redirect(url_for("expense_list"))



# ============================================================
# REPORTS MODULE
# FILTER HELPER
# ============================================================

from datetime import datetime, timedelta


def get_report_date_filter():

    filter_type = request.args.get(
        "filter",
        "all"
    )

    start_date = request.args.get(
        "start_date",
        ""
    )

    end_date = request.args.get(
        "end_date",
        ""
    )

    today = datetime.today()

    # ---------------------------------
    # TODAY
    # ---------------------------------

    if filter_type == "today":

        start_date = today.strftime("%Y-%m-%d")

        end_date = start_date

    # ---------------------------------
    # THIS WEEK
    # ---------------------------------

    elif filter_type == "week":

        monday = today - timedelta(
            days=today.weekday()
        )

        start_date = monday.strftime("%Y-%m-%d")

        end_date = today.strftime("%Y-%m-%d")

    # ---------------------------------
    # THIS MONTH
    # ---------------------------------

    elif filter_type == "month":

        start_date = today.replace(
            day=1
        ).strftime("%Y-%m-%d")

        end_date = today.strftime("%Y-%m-%d")

    # ---------------------------------
    # LAST MONTH
    # ---------------------------------

    elif filter_type == "last_month":

        first_day = today.replace(day=1)

        last_day_previous = first_day - timedelta(days=1)

        start_date = last_day_previous.replace(
            day=1
        ).strftime("%Y-%m-%d")

        end_date = last_day_previous.strftime("%Y-%m-%d")

    # ---------------------------------
    # LAST 3 MONTHS
    # ---------------------------------

    elif filter_type == "3months":

        start_date = (

            today - timedelta(days=90)

        ).strftime("%Y-%m-%d")

        end_date = today.strftime("%Y-%m-%d")

    # ---------------------------------
    # THIS YEAR
    # ---------------------------------

    elif filter_type == "year":

        start_date = today.replace(

            month=1,

            day=1

        ).strftime("%Y-%m-%d")

        end_date = today.strftime("%Y-%m-%d")

    # ---------------------------------
    # CUSTOM
    # ---------------------------------

    elif filter_type == "custom":

        if start_date == "" or end_date == "":

            start_date = ""

            end_date = ""

    else:

        start_date = ""

        end_date = ""

    return (

        filter_type,

        start_date,

        end_date

    )


# ============================================================
# SQL FILTER HELPER
# ============================================================

def build_date_filter(
    user_id,
    start_date,
    end_date
):

    condition = ""

    params = [user_id]

    if start_date != "" and end_date != "":

        condition = """

        AND date BETWEEN ? AND ?

        """

        params.extend(

            [

                start_date,

                end_date

            ]

        )

    return (

        condition,

        params

    )


@app.route("/reports")
def reports():

    # =====================================================
    # LOGIN CHECK
    # =====================================================

    if "user_id" not in session:
        return redirect(url_for("login"))

    # =====================================================
    # DATABASE CONNECTION
    # =====================================================

    conn = sqlite3.connect("finance.db")
    conn.row_factory = sqlite3.Row

    cursor = conn.cursor()

    user_id = session["user_id"]

    # =====================================================
    # REPORT FILTER
    # =====================================================

    filter_type, start_date, end_date = get_report_date_filter()

    income_condition, income_params = build_date_filter(
        user_id,
        start_date,
        end_date
    )

    expense_condition, expense_params = build_date_filter(
        user_id,
        start_date,
        end_date
    )

    print("=" * 70)
    print("REPORT GENERATED")
    print("Filter :", filter_type)
    print("Start :", start_date)
    print("End   :", end_date)
    print("=" * 70)

    # =====================================================
    # REPORT TITLE
    # =====================================================

    report_title = "Financial Report"

    if filter_type == "today":

        report_title = "Today's Financial Report"

    elif filter_type == "week":

        report_title = "This Week Report"

    elif filter_type == "month":

        report_title = "Monthly Financial Report"

    elif filter_type == "last_month":

        report_title = "Last Month Report"

    elif filter_type == "3months":

        report_title = "Last 3 Months Report"

    elif filter_type == "year":

        report_title = "Yearly Financial Report"

    elif filter_type == "custom":

        report_title = "Custom Financial Report"

    # =====================================================
    # REPORT PERIOD
    # =====================================================

    if start_date == "" or end_date == "":

        report_period = "All Records"

    else:

        report_period = f"{start_date}  →  {end_date}"

    # =====================================================
    # TOTAL INCOME
    # =====================================================

    cursor.execute(f"""

        SELECT

            IFNULL(SUM(amount),0) AS total

        FROM income

        WHERE user_id=?

        {income_condition}

    """, income_params)

    total_income = cursor.fetchone()["total"]

    # =====================================================
    # TOTAL EXPENSE
    # =====================================================

    cursor.execute(f"""

        SELECT

            IFNULL(SUM(amount),0) AS total

        FROM expense

        WHERE user_id=?

        {expense_condition}

    """, expense_params)

    total_expense = cursor.fetchone()["total"]

    # =====================================================
    # SAVINGS
    # =====================================================

    savings = total_income - total_expense

        # =====================================================
    # MONTHLY INCOME
    # =====================================================

    cursor.execute(f"""

        SELECT

            COALESCE(SUM(amount),0) AS total

        FROM income

        WHERE user_id=?

        {income_condition}

    """, income_params)

    monthly_income = cursor.fetchone()["total"]

    # =====================================================
    # MONTHLY EXPENSE
    # =====================================================

    cursor.execute(f"""

        SELECT

            COALESCE(SUM(amount),0) AS total

        FROM expense

        WHERE user_id=?

        {expense_condition}

    """, expense_params)

    monthly_expense = cursor.fetchone()["total"]

    # =====================================================
    # CATEGORY ANALYSIS
    # =====================================================

    cursor.execute(f"""

        SELECT

            category,

            SUM(amount) AS total

        FROM expense

        WHERE user_id=?

        {expense_condition}

        GROUP BY category

        ORDER BY total DESC

    """, expense_params)

    rows = cursor.fetchall()

    categories = []

    category_amounts = []

    for row in rows:

        categories.append(

            row["category"]

        )

        category_amounts.append(

            row["total"]

        )

    # =====================================================
    # TOP SPENDING CATEGORY
    # =====================================================

    if len(categories) > 0:

        top_category = categories[0]

        top_category_amount = category_amounts[0]

    else:

        top_category = "N/A"

        top_category_amount = 0

    # =====================================================
    # MONTHLY BUDGET
    # =====================================================

    cursor.execute("""

        SELECT

            monthly_budget

        FROM budget

        WHERE user_id=?

    """, (user_id,))

    row = cursor.fetchone()

    if row:

        monthly_budget = row["monthly_budget"]

    else:

        monthly_budget = 0

    # =====================================================
    # BUDGET REPORT
    # =====================================================

    if monthly_budget > 0:

        budget_used = round(

            (total_expense / monthly_budget) * 100,

            2

        )

        budget_remaining = round(

            monthly_budget - total_expense,

            2

        )

    else:

        budget_used = 0

        budget_remaining = 0

    # =====================================================
    # BUDGET STATUS
    # =====================================================

    if budget_used >= 100:

        budget_status = "Budget Exceeded"

    elif budget_used >= 90:

        budget_status = "Budget Almost Full"

    elif budget_used >= 70:

        budget_status = "Budget Healthy"

    else:

        budget_status = "Excellent"

        # =====================================================
    # INVESTMENT REPORT
    # =====================================================

    cursor.execute("""

        SELECT

            COALESCE(SUM(invested_amount),0) AS invested,

            COALESCE(SUM(current_value),0) AS current

        FROM investments

        WHERE user_id=?

    """,(user_id,))

    investment = cursor.fetchone()

    total_invested = investment["invested"]

    current_value = investment["current"]

    investment_profit = round(

        current_value - total_invested,

        2

    )

    if total_invested > 0:

        roi = round(

            (investment_profit / total_invested) * 100,

            2

        )

    else:

        roi = 0

    # =====================================================
    # ASSET ALLOCATION
    # =====================================================

    cursor.execute("""

        SELECT

            asset_type,

            SUM(current_value) AS total

        FROM investments

        WHERE user_id=?

        GROUP BY asset_type

        ORDER BY total DESC

    """,(user_id,))

    assets = cursor.fetchall()

    asset_labels = []

    asset_values = []

    for asset in assets:

        asset_labels.append(

            asset["asset_type"]

        )

        asset_values.append(

            asset["total"]

        )

    # =====================================================
    # BEST PERFORMING ASSET
    # =====================================================

    if len(asset_labels) > 0:

        best_asset = asset_labels[0]

    else:

        best_asset = "N/A"

    # =====================================================
    # GOAL REPORT
    # =====================================================

    cursor.execute("""

        SELECT

            goal_name,

            goal_type,

            target_amount,

            saved_amount,

            target_date

        FROM financial_goals

        WHERE user_id=?

        ORDER BY target_date

    """,(user_id,))

    goals = cursor.fetchall()

    goal_data = []

    goal_labels = []

    goal_progress = []

    completed_goals = 0

    total_goal_target = 0

    total_goal_saved = 0

    for goal in goals:

        target = goal["target_amount"]

        saved = goal["saved_amount"]

        total_goal_target += target

        total_goal_saved += saved

        if target > 0:

            progress = round(

                (saved / target) * 100,

                1

            )

        else:

            progress = 0

        if progress >= 100:

            status = "Achieved"

            completed_goals += 1

        elif progress >= 75:

            status = "Almost Complete"

        elif progress >= 50:

            status = "Halfway"

        else:

            status = "In Progress"

        goal_labels.append(

            goal["goal_name"]

        )

        goal_progress.append(

            progress

        )

        goal_data.append({

            "goal_name": goal["goal_name"],

            "goal_type": goal["goal_type"],

            "target_amount": target,

            "saved_amount": saved,

            "target_date": goal["target_date"],

            "progress": progress,

            "status": status

        })

    # =====================================================
    # OVERALL GOAL COMPLETION
    # =====================================================

    if total_goal_target > 0:

        overall_goal_progress = round(

            (total_goal_saved / total_goal_target) * 100,

            2

        )

    else:

        overall_goal_progress = 0

        # =====================================================
    # SAVINGS RATE
    # =====================================================

    if total_income > 0:

        saving_rate = round(

            (savings / total_income) * 100,

            2

        )

    else:

        saving_rate = 0


    # =====================================================
    # EXPENSE RATIO
    # =====================================================

    if total_income > 0:

        expense_ratio = round(

            (total_expense / total_income) * 100,

            2

        )

    else:

        expense_ratio = 0


    # =====================================================
    # INVESTMENT RATIO
    # =====================================================

    if total_income > 0:

        investment_ratio = round(

            (total_invested / total_income) * 100,

            2

        )

    else:

        investment_ratio = 0


    # =====================================================
    # FINANCIAL HEALTH SCORE
    # =====================================================

    health_score = 100

    if expense_ratio > 90:

        health_score -= 30

    elif expense_ratio > 75:

        health_score -= 20

    elif expense_ratio > 60:

        health_score -= 10


    if saving_rate < 10:

        health_score -= 25

    elif saving_rate < 20:

        health_score -= 15

    elif saving_rate < 30:

        health_score -= 5


    if budget_used > 100:

        health_score -= 20

    elif budget_used > 90:

        health_score -= 10


    if roi < 0:

        health_score -= 15

    elif roi < 5:

        health_score -= 5


    if completed_goals == len(goal_data) and len(goal_data) > 0:

        health_score += 5


    health_score = max(

        0,

        min(

            100,

            health_score

        )

    )


    # =====================================================
    # HEALTH STATUS
    # =====================================================

    if health_score >= 90:

        health_status = "Excellent"

        health_color = "success"

    elif health_score >= 75:

        health_status = "Good"

        health_color = "primary"

    elif health_score >= 60:

        health_status = "Average"

        health_color = "warning"

    else:

        health_status = "Needs Improvement"

        health_color = "danger"


    # =====================================================
    # AI RECOMMENDATIONS
    # =====================================================

    recommendations = []


    # Savings

    if saving_rate < 20:

        recommendations.append({

            "type":"warning",

            "title":"Increase Savings",

            "message":"Try to save at least 20% of your monthly income."

        })


    # Budget

    if budget_used > 90:

        recommendations.append({

            "type":"danger",

            "title":"Budget Alert",

            "message":"Your monthly budget is almost exhausted."

        })


    # Expenses

    if expense_ratio > 80:

        recommendations.append({

            "type":"danger",

            "title":"High Expenses",

            "message":"Your expenses are consuming a large portion of your income."

        })


    # Investment

    if roi < 5:

        recommendations.append({

            "type":"info",

            "title":"Investment Advice",

            "message":"Consider reviewing or diversifying your investment portfolio."

        })


    # Goals

    if completed_goals == len(goal_data) and len(goal_data) > 0:

        recommendations.append({

            "type":"success",

            "title":"Goals Achieved",

            "message":"Congratulations! All your financial goals have been achieved."

        })


    elif overall_goal_progress >= 80:

        recommendations.append({

            "type":"success",

            "title":"Almost There",

            "message":"You are very close to completing your financial goals."

        })


    # Healthy Report

    if len(recommendations) == 0:

        recommendations.append({

            "type":"success",

            "title":"Excellent Financial Health",

            "message":"Your financial report looks healthy. Keep up the good work!"

        })


    # =====================================================
    # REPORT SUMMARY
    # =====================================================

    cursor.execute(f"""
        SELECT COUNT(*) AS count FROM income
        WHERE user_id=?
        {income_condition}
    """, income_params)
    income_count = cursor.fetchone()["count"]

    cursor.execute(f"""
        SELECT COUNT(*) AS count FROM expense
        WHERE user_id=?
        {expense_condition}
    """, expense_params)
    expense_count = cursor.fetchone()["count"]

    total_transactions = income_count + expense_count

    net_worth = current_value + savings

    report_generated_on = datetime.now().strftime(

        "%d-%m-%Y %I:%M %p"

    )

    print("="*70)

    print("REPORT GENERATED SUCCESSFULLY")

    print("Income :", total_income)

    print("Expense :", total_expense)

    print("Savings :", savings)

    print("Health Score :", health_score)

    print("ROI :", roi)

    print("Saving Rate :", saving_rate)

    print("="*70)

    # =====================================================
    # TRANSACTION REPORT
    # =====================================================

    transaction_params = [user_id]

    transaction_condition = ""

    if start_date and end_date:

        transaction_condition = """

        AND date BETWEEN ? AND ?

        """

        transaction_params.extend([

            start_date,

            end_date

        ])

    cursor.execute(f"""

        SELECT

            date,

            'Income' AS transaction_type,

            source AS category,

            source AS description,

            amount

        FROM income

        WHERE user_id=?

        {transaction_condition}

        UNION ALL

        SELECT

            date,

            'Expense' AS transaction_type,

            category,

            description,

            amount

        FROM expense

        WHERE user_id=?

        {transaction_condition}

        ORDER BY date DESC

    """,

    transaction_params + transaction_params)

    transactions = cursor.fetchall()

    # =====================================================
    # TRANSACTION SUMMARY
    # =====================================================

    transaction_count = len(transactions)

    income_transactions = 0

    expense_transactions = 0

    highest_income = 0

    highest_expense = 0

    average_income = 0

    average_expense = 0

    income_total_for_average = 0

    expense_total_for_average = 0

    for row in transactions:

        if row["transaction_type"] == "Income":

            income_transactions += 1

            income_total_for_average += row["amount"]

            if row["amount"] > highest_income:

                highest_income = row["amount"]

        else:

            expense_transactions += 1

            expense_total_for_average += row["amount"]

            if row["amount"] > highest_expense:

                highest_expense = row["amount"]

    # =====================================================
    # AVERAGES
    # =====================================================

    if income_transactions > 0:

        average_income = round(

            income_total_for_average /

            income_transactions,

            2

        )

    else:

        average_income = 0

    if expense_transactions > 0:

        average_expense = round(

            expense_total_for_average /

            expense_transactions,

            2

        )

    else:

        average_expense = 0

    # =====================================================
    # RECENT TRANSACTIONS
    # =====================================================

    recent_transactions = transactions[:10]

    # =====================================================
    # REPORT TABLE
    # =====================================================

    report_table = []

    serial = 1

    for row in transactions:

        report_table.append({

            "sl": serial,

            "date": row["date"],

            "type": row["transaction_type"],

            "category": row["category"],

            "description": row["description"],

            "amount": row["amount"]

        })

        serial += 1

    print("=" * 70)
    print("TRANSACTION REPORT")
    print("Total Transactions :", transaction_count)
    print("Income Transactions :", income_transactions)
    print("Expense Transactions :", expense_transactions)
    print("Highest Income :", highest_income)
    print("Highest Expense :", highest_expense)
    print("=" * 70)

        # =====================================================
    # MONTHLY INCOME TREND
    # =====================================================

    cursor.execute("""

        SELECT

            strftime('%Y-%m', date) AS month,

            SUM(amount) AS total

        FROM income

        WHERE user_id=?

        GROUP BY strftime('%Y-%m', date)

        ORDER BY month

    """,(user_id,))

    income_monthly = cursor.fetchall()

    income_month_labels = []

    income_month_values = []

    highest_income_month = ""

    highest_income_value = 0

    total_monthly_income = 0

    for row in income_monthly:

        income_month_labels.append(

            row["month"]

        )

        income_month_values.append(

            round(

                row["total"],

                2

            )

        )

        total_monthly_income += row["total"]

        if row["total"] > highest_income_value:

            highest_income_value = row["total"]

            highest_income_month = row["month"]

    if len(income_month_values) > 0:

        average_monthly_income = round(

            total_monthly_income /

            len(income_month_values),

            2

        )

    else:

        average_monthly_income = 0

    print("=" * 60)
    print("MONTHLY INCOME TREND")
    print("Months :", len(income_month_labels))
    print("Highest Month :", highest_income_month)
    print("Highest Income :", highest_income_value)
    print("Average Monthly Income :", average_monthly_income)
    print("=" * 60)

        # =====================================================
    # MONTHLY EXPENSE TREND
    # =====================================================

    cursor.execute("""

        SELECT

            strftime('%Y-%m', date) AS month,

            SUM(amount) AS total

        FROM expense

        WHERE user_id=?

        GROUP BY strftime('%Y-%m', date)

        ORDER BY month

    """,(user_id,))

    expense_monthly = cursor.fetchall()

    expense_month_labels = []

    expense_month_values = []

    highest_expense_month = ""

    highest_expense_value = 0

    total_monthly_expense = 0

    lowest_expense_month = ""

    lowest_expense_value = None

    for row in expense_monthly:

        expense_month_labels.append(

            row["month"]

        )

        expense_month_values.append(

            round(

                row["total"],

                2

            )

        )

        total_monthly_expense += row["total"]

        if row["total"] > highest_expense_value:

            highest_expense_value = row["total"]

            highest_expense_month = row["month"]

        if lowest_expense_value is None:

            lowest_expense_value = row["total"]

            lowest_expense_month = row["month"]

        elif row["total"] < lowest_expense_value:

            lowest_expense_value = row["total"]

            lowest_expense_month = row["month"]

    if len(expense_month_values) > 0:

        average_monthly_expense = round(

            total_monthly_expense /

            len(expense_month_values),

            2

        )

    else:

        average_monthly_expense = 0

    # =====================================================
    # EXPENSE TREND SUMMARY
    # =====================================================

    expense_growth = []

    if len(expense_month_values) > 1:

        for i in range(

            1,

            len(expense_month_values)

        ):

            growth = round(

                expense_month_values[i] -

                expense_month_values[i-1],

                2

            )

            expense_growth.append(

                growth

            )

    else:

        expense_growth = []

    print("=" * 60)
    print("MONTHLY EXPENSE TREND")
    print("Months :", len(expense_month_labels))
    print("Highest Month :", highest_expense_month)
    print("Highest Expense :", highest_expense_value)
    print("Lowest Month :", lowest_expense_month)
    print("Lowest Expense :", lowest_expense_value)
    print("Average Monthly Expense :", average_monthly_expense)
    print("=" * 60)

        # =====================================================
    # MONTHLY SAVINGS TREND
    # =====================================================

    savings_month_labels = []
    savings_month_values = []

    monthly_income_map = {}

    monthly_expense_map = {}

    for i in range(len(income_month_labels)):

        monthly_income_map[
            income_month_labels[i]
        ] = income_month_values[i]

    for i in range(len(expense_month_labels)):

        monthly_expense_map[
            expense_month_labels[i]
        ] = expense_month_values[i]

    all_months = sorted(

        set(

            income_month_labels +

            expense_month_labels

        )

    )

    best_saving_month = ""

    best_saving_value = -999999999

    worst_saving_month = ""

    worst_saving_value = 999999999

    total_savings_overall = 0

    for month in all_months:

        income = monthly_income_map.get(

            month,

            0

        )

        expense = monthly_expense_map.get(

            month,

            0

        )

        saving = round(

            income - expense,

            2

        )

        savings_month_labels.append(

            month

        )

        savings_month_values.append(

            saving

        )

        total_savings_overall += saving

        if saving > best_saving_value:

            best_saving_value = saving

            best_saving_month = month

        if saving < worst_saving_value:

            worst_saving_value = saving

            worst_saving_month = month

    if len(savings_month_values) > 0:

        average_monthly_saving = round(

            total_savings_overall /

            len(savings_month_values),

            2

        )

    else:

        average_monthly_saving = 0

    # =====================================================
    # CHART DATA
    # =====================================================

    monthly_chart = {

        "labels": all_months,

        "income": income_month_values,

        "expense": expense_month_values,

        "saving": savings_month_values

    }

    # =====================================================
    # DASHBOARD HIGHLIGHTS
    # =====================================================

    report_highlights = {

        "best_income_month": highest_income_month,

        "best_income": highest_income_value,

        "highest_expense_month": highest_expense_month,

        "highest_expense": highest_expense_value,

        "lowest_expense_month": lowest_expense_month,

        "lowest_expense": lowest_expense_value,

        "best_saving_month": best_saving_month,

        "best_saving": best_saving_value,

        "worst_saving_month": worst_saving_month,

        "worst_saving": worst_saving_value,

        "average_income": average_monthly_income,

        "average_expense": average_monthly_expense,

        "average_saving": average_monthly_saving

    }

    print("=" * 70)
    print("MONTHLY SAVINGS TREND")
    print("Best Saving Month :", best_saving_month)
    print("Best Saving :", best_saving_value)
    print("Worst Saving Month :", worst_saving_month)
    print("Worst Saving :", worst_saving_value)
    print("Average Monthly Saving :", average_monthly_saving)
    print("=" * 70)

        # =====================================================
    # YEARLY FINANCIAL TREND
    # =====================================================

    cursor.execute("""

        SELECT

            strftime('%Y', date) AS year,

            SUM(amount) AS total

        FROM income

        WHERE user_id=?

        GROUP BY strftime('%Y', date)

        ORDER BY year

    """,(user_id,))

    yearly_income = cursor.fetchall()

    cursor.execute("""

        SELECT

            strftime('%Y', date) AS year,

            SUM(amount) AS total

        FROM expense

        WHERE user_id=?

        GROUP BY strftime('%Y', date)

        ORDER BY year

    """,(user_id,))

    yearly_expense = cursor.fetchall()

    income_year_map = {}

    expense_year_map = {}

    for row in yearly_income:

        income_year_map[

            row["year"]

        ] = row["total"]

    for row in yearly_expense:

        expense_year_map[

            row["year"]

        ] = row["total"]

    all_years = sorted(

        set(

            list(income_year_map.keys()) +

            list(expense_year_map.keys())

        )

    )

    yearly_labels = []

    yearly_income_values = []

    yearly_expense_values = []

    yearly_saving_values = []

    best_year = ""

    best_year_saving = -999999999

    worst_year = ""

    worst_year_saving = 999999999

    total_yearly_income = 0

    total_yearly_expense = 0

    total_yearly_savings = 0

    for year in all_years:

        income = income_year_map.get(

            year,

            0

        )

        expense = expense_year_map.get(

            year,

            0

        )

        saving = income - expense

        yearly_labels.append(

            year

        )

        yearly_income_values.append(

            round(

                income,

                2

            )

        )

        yearly_expense_values.append(

            round(

                expense,

                2

            )

        )

        yearly_saving_values.append(

            round(

                saving,

                2

            )

        )

        total_yearly_income += income

        total_yearly_expense += expense

        total_yearly_savings += saving

        if saving > best_year_saving:

            best_year_saving = saving

            best_year = year

        if saving < worst_year_saving:

            worst_year_saving = saving

            worst_year = year

    if len(yearly_labels) > 0:

        average_yearly_income = round(

            total_yearly_income /

            len(yearly_labels),

            2

        )

        average_yearly_expense = round(

            total_yearly_expense /

            len(yearly_labels),

            2

        )

        average_yearly_saving = round(

            total_yearly_savings /

            len(yearly_labels),

            2

        )

    else:

        average_yearly_income = 0

        average_yearly_expense = 0

        average_yearly_saving = 0

    # =====================================================
    # YEARLY GROWTH
    # =====================================================

    yearly_growth = []

    if len(yearly_saving_values) > 1:

        for i in range(

            1,

            len(yearly_saving_values)

        ):

            growth = round(

                yearly_saving_values[i] -

                yearly_saving_values[i-1],

                2

            )

            yearly_growth.append(

                growth

            )

    else:

        yearly_growth = []

    # =====================================================
    # YEARLY CHART DATA
    # =====================================================

    yearly_chart = {

        "labels": yearly_labels,

        "income": yearly_income_values,

        "expense": yearly_expense_values,

        "saving": yearly_saving_values

    }

    print("=" * 70)
    print("YEARLY FINANCIAL TREND")
    print("Years :", len(yearly_labels))
    print("Best Financial Year :", best_year)
    print("Best Saving :", best_year_saving)
    print("Worst Financial Year :", worst_year)
    print("Worst Saving :", worst_year_saving)
    print("Average Yearly Income :", average_yearly_income)
    print("Average Yearly Expense :", average_yearly_expense)
    print("Average Yearly Saving :", average_yearly_saving)
    print("=" * 70)

        # =====================================================
    # REPORT STATISTICS
    # =====================================================

    report_statistics = {

        "total_income": total_income,

        "total_expense": total_expense,

        "total_savings": savings,

        "saving_rate": saving_rate,

        "expense_ratio": expense_ratio,

        "investment_ratio": investment_ratio,

        "health_score": health_score,

        "health_status": health_status,

        "budget_used": budget_used,

        "budget_remaining": budget_remaining,

        "completed_goals": completed_goals,

        "total_goals": len(goal_data),

        "roi": roi,

        "transaction_count": transaction_count

    }

    # =====================================================
    # REPORT KPI CARDS
    # =====================================================

    kpi_cards = [

        {

            "title": "Total Income",

            "value": round(total_income,2),

            "icon": "fa-wallet",

            "color": "success"

        },

        {

            "title": "Total Expense",

            "value": round(total_expense,2),

            "icon": "fa-money-bill-wave",

            "color": "danger"

        },

        {

            "title": "Savings",

            "value": round(savings,2),

            "icon": "fa-piggy-bank",

            "color": "primary"

        },

        {

            "title": "ROI",

            "value": f"{roi}%",

            "icon": "fa-chart-line",

            "color": "info"

        },

        {

            "title": "Health Score",

            "value": f"{health_score}/100",

            "icon": "fa-heart",

            "color": health_color

        }

    ]

    # =====================================================
    # CHART DATA
    # =====================================================

    chart_data = {

        "category_labels": categories,

        "category_values": category_amounts,

        "asset_labels": asset_labels,

        "asset_values": asset_values,

        "goal_labels": goal_labels,

        "goal_progress": goal_progress,

        "monthly": monthly_chart,

        "yearly": yearly_chart

    }

    # =====================================================
    # REPORT INFORMATION
    # =====================================================

    report_info = {

        "title": report_title,

        "period": report_period,

        "generated_on": report_generated_on,

        "generated_by": session.get(

            "username",

            "User"

        )

    }

    # =====================================================
    # EXPORT DATA
    # =====================================================

    export_data = {

        "summary": report_statistics,

        "transactions": report_table,

        "goals": goal_data,

        "recommendations": recommendations,

        "investments": {

            "invested": total_invested,

            "current": current_value,

            "profit": investment_profit,

            "roi": roi

        },

        "budget": {

            "monthly_budget": monthly_budget,

            "budget_used": budget_used,

            "budget_remaining": budget_remaining,

            "status": budget_status

        }

    }

    # =====================================================
    # REPORT STATUS
    # =====================================================

    if total_income == 0 and total_expense == 0:

        report_status = "No Records Found"

    else:

        report_status = "Report Generated Successfully"

    print("=" * 80)
    print("REPORT DASHBOARD READY")
    print("Status :", report_status)
    print("Charts :", len(chart_data))
    print("KPIs :", len(kpi_cards))
    print("Recommendations :", len(recommendations))
    print("Transactions :", transaction_count)
    print("=" * 80)

        # =====================================================
    # CLOSE DATABASE
    # =====================================================

    conn.close()

    # =====================================================
    # RENDER REPORT
    # =====================================================

    return render_template(

        "reports.html",

        # -------------------------------------------------
        # REPORT INFORMATION
        # -------------------------------------------------

        report_title=report_title,
        report_period=report_period,
        report_generated_on=report_generated_on,
        report_status=report_status,
        filter_type=filter_type,
        start_date=start_date,
        end_date=end_date,

        # -------------------------------------------------
        # SUMMARY
        # -------------------------------------------------

        total_income=total_income,
        total_expense=total_expense,
        savings=savings,

        monthly_income=monthly_income,
        monthly_expense=monthly_expense,

        saving_rate=saving_rate,
        expense_ratio=expense_ratio,
        investment_ratio=investment_ratio,

        # -------------------------------------------------
        # BUDGET
        # -------------------------------------------------

        monthly_budget=monthly_budget,
        budget_used=budget_used,
        budget_remaining=budget_remaining,
        budget_status=budget_status,

        # -------------------------------------------------
        # CATEGORY ANALYSIS
        # -------------------------------------------------

        categories=categories,
        category_amounts=category_amounts,
        top_category=top_category,
        top_category_amount=top_category_amount,

        # -------------------------------------------------
        # INVESTMENTS
        # -------------------------------------------------

        total_invested=total_invested,
        current_value=current_value,
        investment_profit=investment_profit,
        roi=roi,

        asset_labels=asset_labels,
        asset_values=asset_values,
        best_asset=best_asset,

        # -------------------------------------------------
        # GOALS
        # -------------------------------------------------

        goal_data=goal_data,
        goal_labels=goal_labels,
        goal_progress=goal_progress,

        completed_goals=completed_goals,
        overall_goal_progress=overall_goal_progress,

        # -------------------------------------------------
        # HEALTH
        # -------------------------------------------------

        health_score=health_score,
        health_status=health_status,
        health_color=health_color,

        # -------------------------------------------------
        # AI
        # -------------------------------------------------

        recommendations=recommendations,

        # -------------------------------------------------
        # TRANSACTIONS
        # -------------------------------------------------

        transactions=transactions,
        report_table=report_table,
        recent_transactions=recent_transactions,

        transaction_count=transaction_count,
        income_transactions=income_transactions,
        expense_transactions=expense_transactions,

        highest_income=highest_income,
        highest_expense=highest_expense,

        average_income=average_income,
        average_expense=average_expense,

        # -------------------------------------------------
        # MONTHLY TREND
        # -------------------------------------------------

        income_month_labels=income_month_labels,
        income_month_values=income_month_values,

        expense_month_labels=expense_month_labels,
        expense_month_values=expense_month_values,

        savings_month_labels=savings_month_labels,
        savings_month_values=savings_month_values,

        highest_income_month=highest_income_month,
        highest_income_value=highest_income_value,

        highest_expense_month=highest_expense_month,
        highest_expense_value=highest_expense_value,

        lowest_expense_month=lowest_expense_month,
        lowest_expense_value=lowest_expense_value,

        average_monthly_income=average_monthly_income,
        average_monthly_expense=average_monthly_expense,
        average_monthly_saving=average_monthly_saving,

        best_saving_month=best_saving_month,
        best_saving_value=best_saving_value,

        worst_saving_month=worst_saving_month,
        worst_saving_value=worst_saving_value,

        monthly_chart=monthly_chart,

        # -------------------------------------------------
        # YEARLY TREND
        # -------------------------------------------------

        yearly_labels=yearly_labels,
        yearly_income_values=yearly_income_values,
        yearly_expense_values=yearly_expense_values,
        yearly_saving_values=yearly_saving_values,

        yearly_growth=yearly_growth,

        best_year=best_year,
        best_year_saving=best_year_saving,

        worst_year=worst_year,
        worst_year_saving=worst_year_saving,

        average_yearly_income=average_yearly_income,
        average_yearly_expense=average_yearly_expense,
        average_yearly_saving=average_yearly_saving,

        yearly_chart=yearly_chart,

        # -------------------------------------------------
        # DASHBOARD
        # -------------------------------------------------

        report_statistics=report_statistics,
        report_highlights=report_highlights,
        chart_data=chart_data,
        kpi_cards=kpi_cards,
        report_info=report_info,

        # -------------------------------------------------
        # EXPORT
        # -------------------------------------------------

export_data=export_data

    )


# ============================================================
# EXPORT REPORT AS EXCEL
# ============================================================

@app.route("/export_excel")
def export_excel():

    # =====================================================
    # LOGIN CHECK
    # =====================================================

    if "user_id" not in session:
        return redirect(url_for("login"))

    # =====================================================
    # REPORT FILTER  (same filter as the reports page)
    # =====================================================

    filter_type, start_date, end_date = get_report_date_filter()

    user_id = session["user_id"]

    conn = sqlite3.connect("finance.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    income_condition, income_params = build_date_filter(
        user_id,
        start_date,
        end_date
    )

    expense_condition, expense_params = build_date_filter(
        user_id,
        start_date,
        end_date
    )

    # =====================================================
    # REPORT TITLE
    # =====================================================

    report_title = "Financial Report"

    if filter_type == "today":
        report_title = "Today's Financial Report"
    elif filter_type == "week":
        report_title = "This Week Report"
    elif filter_type == "month":
        report_title = "Monthly Financial Report"
    elif filter_type == "last_month":
        report_title = "Last Month Report"
    elif filter_type == "3months":
        report_title = "Last 3 Months Report"
    elif filter_type == "year":
        report_title = "Yearly Financial Report"
    elif filter_type == "custom":
        report_title = "Custom Financial Report"

    # =====================================================
    # REPORT PERIOD
    # =====================================================

    if not start_date or not end_date:
        report_period = "All Records"
    else:
        report_period = f"{start_date} to {end_date}"

    # =====================================================
    # TOTAL INCOME / EXPENSE / SAVINGS
    # =====================================================

    cursor.execute(f"""
        SELECT IFNULL(SUM(amount),0) AS total
        FROM income
        WHERE user_id=? {income_condition}
    """, income_params)
    total_income = cursor.fetchone()["total"]

    cursor.execute(f"""
        SELECT IFNULL(SUM(amount),0) AS total
        FROM expense
        WHERE user_id=? {expense_condition}
    """, expense_params)
    total_expense = cursor.fetchone()["total"]

    savings = total_income - total_expense

    if total_income > 0:
        saving_rate = round((savings / total_income) * 100, 2)
    else:
        saving_rate = 0

    # =====================================================
    # TRANSACTION REPORT
    # =====================================================

    transaction_params = [user_id]
    transaction_condition = ""
    if start_date and end_date:
        transaction_condition = "AND date BETWEEN ? AND ?"
        transaction_params.extend([start_date, end_date])

    cursor.execute(f"""
        SELECT
            date,
            'Income' AS transaction_type,
            source AS category,
            source AS description,
            amount
        FROM income
        WHERE user_id=? {transaction_condition}
        UNION ALL
        SELECT
            date,
            'Expense' AS transaction_type,
            category,
            description,
            amount
        FROM expense
        WHERE user_id=? {transaction_condition}
        ORDER BY date DESC
    """, transaction_params + transaction_params)

    transactions = cursor.fetchall()

    conn.close()

    # =====================================================
    # CREATE WORKBOOK
    # =====================================================

    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    title_font = Font(bold=True, size=14)

    # ---- Title ----
    ws["A1"] = "Finance Analytics Platform"
    ws["A1"].font = title_font
    ws["A2"] = report_title
    ws["A2"].font = Font(bold=True, size=12)
    ws["A3"] = f"Report Period : {report_period}"
    ws["A4"] = f"Saving Rate : {saving_rate}%"

    # ---- Income Section ----
    ws["A6"] = "Income"
    ws["A6"].font = Font(bold=True, size=12)
    ws.append(["Date", "Source", "Amount"])

    cursor_income_col = 7
    for row in transactions:
        if row["transaction_type"] == "Income":
            ws.append([row["date"], row["category"], row["amount"]])

    # ---- Expense Section ----
    start_row = ws.max_row + 2
    ws[f"A{start_row}"] = "Expenses"
    ws[f"A{start_row}"].font = Font(bold=True, size=12)
    ws.append(["Date", "Category", "Description", "Amount"])

    for row in transactions:
        if row["transaction_type"] == "Expense":
            ws.append([row["date"], row["category"], row["description"], row["amount"]])

    # ---- Summary Section ----
    summary_row = ws.max_row + 2
    ws[f"A{summary_row}"] = "Summary"
    ws[f"A{summary_row}"].font = Font(bold=True, size=12)
    ws.append(["Metric", "Value"])
    ws.append(["Total Income", total_income])
    ws.append(["Total Expense", total_expense])
    ws.append(["Savings", savings])
    ws.append(["Saving Rate", f"{saving_rate}%"])

    # ---- Style headers ----
    for col in range(1, 6):
        cell = ws.cell(row=6, column=col)
        if cell.value:
            cell.font = header_font
            cell.fill = header_fill

    # =====================================================
    # SAVE FILE
    # =====================================================

    excel_path = os.path.join("exports", "Financial_Report.xlsx")
    os.makedirs("exports", exist_ok=True)
    wb.save(excel_path)

    return send_file(
        excel_path,
        as_attachment=True
    )


# ============================================================
# EXPORT REPORT AS PDF
# ============================================================

@app.route("/export_pdf")
def export_pdf():

    # =====================================================
    # LOGIN CHECK
    # =====================================================

    if "user_id" not in session:
        return redirect(url_for("login"))

    # =====================================================
    # REPORT FILTER  (same filter as the reports page)
    # =====================================================

    filter_type, start_date, end_date = get_report_date_filter()

    user_id = session["user_id"]

    conn = sqlite3.connect("finance.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    income_condition, income_params = build_date_filter(
        user_id,
        start_date,
        end_date
    )

    expense_condition, expense_params = build_date_filter(
        user_id,
        start_date,
        end_date
    )

    # =====================================================
    # REPORT TITLE
    # =====================================================

    report_title = "Financial Report"

    if filter_type == "today":
        report_title = "Today's Financial Report"
    elif filter_type == "week":
        report_title = "This Week Report"
    elif filter_type == "month":
        report_title = "Monthly Financial Report"
    elif filter_type == "last_month":
        report_title = "Last Month Report"
    elif filter_type == "3months":
        report_title = "Last 3 Months Report"
    elif filter_type == "year":
        report_title = "Yearly Financial Report"
    elif filter_type == "custom":
        report_title = "Custom Financial Report"

    # =====================================================
    # REPORT PERIOD
    # =====================================================

    if not start_date or not end_date:
        report_period = "All Records"
    else:
        report_period = f"{start_date}  →  {end_date}"

    # =====================================================
    # TOTAL INCOME / EXPENSE / SAVINGS
    # =====================================================

    cursor.execute(f"""
        SELECT IFNULL(SUM(amount),0) AS total
        FROM income
        WHERE user_id=? {income_condition}
    """, income_params)
    total_income = cursor.fetchone()["total"]

    cursor.execute(f"""
        SELECT IFNULL(SUM(amount),0) AS total
        FROM expense
        WHERE user_id=? {expense_condition}
    """, expense_params)
    total_expense = cursor.fetchone()["total"]

    savings = total_income - total_expense

    # Saving rate / expense ratio
    if total_income > 0:
        saving_rate = round((savings / total_income) * 100, 2)
        expense_ratio = round((total_expense / total_income) * 100, 2)
    else:
        saving_rate = 0
        expense_ratio = 0

    # =====================================================
    # CATEGORY ANALYSIS
    # =====================================================

    cursor.execute(f"""
        SELECT category, SUM(amount) AS total
        FROM expense
        WHERE user_id=? {expense_condition}
        GROUP BY category
        ORDER BY total DESC
    """, expense_params)
    category_rows = cursor.fetchall()

    if len(category_rows) > 0:
        top_category = category_rows[0]["category"]
        top_category_amount = category_rows[0]["total"]
    else:
        top_category = "N/A"
        top_category_amount = 0

    # =====================================================
    # MONTHLY BUDGET
    # =====================================================

    cursor.execute("""
        SELECT monthly_budget
        FROM budget
        WHERE user_id=?
    """, (user_id,))
    budget_row = cursor.fetchone()
    monthly_budget = budget_row["monthly_budget"] if budget_row else 0

    if monthly_budget > 0:
        budget_used = round((total_expense / monthly_budget) * 100, 2)
        budget_remaining = round(monthly_budget - total_expense, 2)
    else:
        budget_used = 0
        budget_remaining = 0

    if budget_used >= 100:
        budget_status = "Budget Exceeded"
    elif budget_used >= 90:
        budget_status = "Budget Almost Full"
    elif budget_used >= 70:
        budget_status = "Budget Healthy"
    else:
        budget_status = "Excellent"

    # =====================================================
    # INVESTMENT SUMMARY
    # =====================================================

    cursor.execute("""
        SELECT
            COALESCE(SUM(invested_amount),0) AS invested,
            COALESCE(SUM(current_value),0) AS current
        FROM investments
        WHERE user_id=?
    """, (user_id,))
    investment = cursor.fetchone()
    total_invested = investment["invested"]
    current_value = investment["current"]
    investment_profit = round(current_value - total_invested, 2)

    if total_invested > 0:
        roi = round((investment_profit / total_invested) * 100, 2)
    else:
        roi = 0

    cursor.execute("""
        SELECT asset_type
        FROM investments
        WHERE user_id=?
        ORDER BY current_value DESC
        LIMIT 1
    """, (user_id,))
    best_row = cursor.fetchone()
    best_asset = best_row["asset_type"] if best_row else "N/A"

    # =====================================================
    # FINANCIAL GOALS
    # =====================================================

    cursor.execute("""
        SELECT
            goal_name,
            goal_type,
            target_amount,
            saved_amount,
            target_date
        FROM financial_goals
        WHERE user_id=?
        ORDER BY target_date
    """, (user_id,))
    goals = cursor.fetchall()

    goal_rows = []
    completed_goals = 0
    for goal in goals:
        target = goal["target_amount"]
        saved = goal["saved_amount"]
        progress = round((saved / target) * 100, 1) if target > 0 else 0
        if progress >= 100:
            status = "Achieved"
            completed_goals += 1
        elif progress >= 75:
            status = "Almost Complete"
        elif progress >= 50:
            status = "Halfway"
        else:
            status = "In Progress"
        goal_rows.append({
            "goal_name": goal["goal_name"],
            "target_amount": target,
            "saved_amount": saved,
            "target_date": goal["target_date"],
            "progress": progress,
            "status": status
        })

    # =====================================================
    # AI RECOMMENDATIONS
    # =====================================================

    recommendations = []
    if saving_rate < 20:
        recommendations.append("Try to save at least 20% of your monthly income.")
    if budget_used > 90:
        recommendations.append("Your monthly budget is almost exhausted.")
    if expense_ratio > 80:
        recommendations.append("Your expenses are consuming a large portion of your income.")
    if roi < 5:
        recommendations.append("Consider reviewing or diversifying your investment portfolio.")
    if completed_goals == len(goal_rows) and len(goal_rows) > 0:
        recommendations.append("Congratulations! All your financial goals have been achieved.")
    if len(recommendations) == 0:
        recommendations.append("Your financial report looks healthy. Keep up the good work!")

    # =====================================================
    # TRANSACTION REPORT
    # =====================================================

    transaction_params = [user_id]
    transaction_condition = ""
    if start_date and end_date:
        transaction_condition = "AND date BETWEEN ? AND ?"
        transaction_params.extend([start_date, end_date])

    cursor.execute(f"""
        SELECT
            date,
            'Income' AS transaction_type,
            source AS category,
            source AS description,
            amount
        FROM income
        WHERE user_id=? {transaction_condition}
        UNION ALL
        SELECT
            date,
            'Expense' AS transaction_type,
            category,
            description,
            amount
        FROM expense
        WHERE user_id=? {transaction_condition}
        ORDER BY date DESC
    """, transaction_params + transaction_params)

    transactions = cursor.fetchall()

    conn.close()

    # =====================================================
    # HEALTH SCORE
    # =====================================================

    health_score = 100
    if expense_ratio > 90:
        health_score -= 30
    elif expense_ratio > 75:
        health_score -= 20
    elif expense_ratio > 60:
        health_score -= 10
    if saving_rate < 10:
        health_score -= 25
    elif saving_rate < 20:
        health_score -= 15
    elif saving_rate < 30:
        health_score -= 5
    if budget_used > 100:
        health_score -= 20
    elif budget_used > 90:
        health_score -= 10
    if roi < 0:
        health_score -= 15
    elif roi < 5:
        health_score -= 5
    health_score = max(0, min(100, health_score))

    if health_score >= 90:
        health_status = "Excellent"
    elif health_score >= 75:
        health_status = "Good"
    elif health_score >= 60:
        health_status = "Average"
    else:
        health_status = "Needs Improvement"

    # =====================================================
    # BUILD PDF
    # =====================================================

    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    styles = getSampleStyleSheet()

    title_style = styles["Title"]
    title_style.alignment = TA_CENTER

    heading_style = styles["Heading2"]
    heading_style.textColor = colors.HexColor("#1f2937")

    normal_style = styles["Normal"]

    story = []

    # ---- Header ----
    story.append(Paragraph("Finance Analytics Platform", title_style))
    story.append(Paragraph(report_title, heading_style))
    story.append(Paragraph(
        f"Report Period : {report_period}",
        normal_style
    ))
    story.append(Paragraph(
        "Generated on : " + datetime.now().strftime("%d-%m-%Y %I:%M %p"),
        normal_style
    ))
    story.append(Spacer(1, 18))

    # ---- Financial Summary ----
    story.append(Paragraph("Financial Summary", heading_style))

    summary_data = [
        ["Metric", "Value"],
        ["Total Income", f"Rs. {total_income:,.2f}"],
        ["Total Expense", f"Rs. {total_expense:,.2f}"],
        ["Savings", f"Rs. {savings:,.2f}"],
        ["Saving Rate", f"{saving_rate}%"],
        ["Expense Ratio", f"{expense_ratio}%"],
        ["Financial Health", f"{health_score}/100 ({health_status})"],
        ["Investment ROI", f"{roi}%"],
        ["Budget Used", f"{budget_used}% ({budget_status})"]
    ]

    summary_table = Table(summary_data, colWidths=[220, 300])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f3f4f6")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("PADDING", (0, 0), (-1, -1), 7),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 18))

    # ---- Budget ----
    story.append(Paragraph("Budget Status", heading_style))
    budget_data = [
        ["Monthly Budget", "Used", "Remaining", "Status"],
        [
            f"Rs. {monthly_budget:,.2f}",
            f"{budget_used}%",
            f"Rs. {budget_remaining:,.2f}",
            budget_status
        ]
    ]
    budget_table = Table(budget_data, colWidths=[140, 100, 140, 140])
    budget_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f59e0b")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("PADDING", (0, 0), (-1, -1), 7),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
    ]))
    story.append(budget_table)
    story.append(Spacer(1, 18))

    # ---- Category Analysis ----
    story.append(Paragraph("Expense Category Analysis", heading_style))
    category_data = [["#", "Category", "Amount"]]
    for idx, row in enumerate(category_rows, start=1):
        category_data.append([
            str(idx),
            row["category"],
            f"Rs. {row['total']:,.2f}"
        ])
    category_data.append(["", "Top Category", top_category])
    category_data.append(["", "Top Amount", f"Rs. {top_category_amount:,.2f}"])
    category_table = Table(category_data, colWidths=[50, 220, 250])
    category_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dc3545")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("PADDING", (0, 0), (-1, -1), 6),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    story.append(category_table)
    story.append(Spacer(1, 18))

    # ---- Investment Summary ----
    story.append(Paragraph("Investment Summary", heading_style))
    investment_data = [
        ["Metric", "Value"],
        ["Total Invested", f"Rs. {total_invested:,.2f}"],
        ["Current Value", f"Rs. {current_value:,.2f}"],
        ["Profit / Loss", f"Rs. {investment_profit:,.2f}"],
        ["ROI", f"{roi}%"],
        ["Best Asset", best_asset]
    ]
    investment_table = Table(investment_data, colWidths=[220, 300])
    investment_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#146ef5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("PADDING", (0, 0), (-1, -1), 7),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
    ]))
    story.append(investment_table)
    story.append(Spacer(1, 18))

    # ---- Financial Goals ----
    story.append(Paragraph("Financial Goals", heading_style))
    goal_table_data = [["Goal", "Target", "Saved", "Progress", "Status"]]
    for goal in goal_rows:
        goal_table_data.append([
            goal["goal_name"],
            f"Rs. {goal['target_amount']:,.2f}",
            f"Rs. {goal['saved_amount']:,.2f}",
            f"{goal['progress']}%",
            goal["status"]
        ])
    if len(goal_table_data) == 1:
        goal_table_data.append(["No goals found", "", "", "", ""])
    goal_table = Table(goal_table_data, colWidths=[120, 100, 100, 80, 120])
    goal_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16a34a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("PADDING", (0, 0), (-1, -1), 6),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    story.append(goal_table)
    story.append(Spacer(1, 18))

    # ---- AI Recommendations ----
    story.append(Paragraph("AI Financial Recommendations", heading_style))
    for rec in recommendations:
        story.append(Paragraph(f"• {rec}", normal_style))
        story.append(Spacer(1, 4))
    story.append(Spacer(1, 12))

    # ---- Transaction Report ----
    story.append(Paragraph(
        f"Transaction Report ({len(transactions)} transactions)",
        heading_style
    ))

    transaction_table_data = [["#", "Date", "Type", "Category", "Description", "Amount"]]
    for idx, row in enumerate(transactions, start=1):
        transaction_table_data.append([
            str(idx),
            row["date"],
            row["transaction_type"],
            row["category"],
            str(row["description"]),
            f"Rs. {row['amount']:,.2f}"
        ])

    if len(transaction_table_data) == 1:
        transaction_table_data.append(["", "No records", "", "", "", ""])

    transaction_table = Table(
        transaction_table_data,
        colWidths=[30, 70, 60, 100, 150, 110],
        repeatRows=1
    )
    transaction_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#343a40")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("PADDING", (0, 0), (-1, -1), 5),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    story.append(transaction_table)
    story.append(Spacer(1, 18))

    # ---- Footer ----
    story.append(Paragraph(
        "Finance Analytics Platform | Generated Report",
        styles["Italic"]
    ))

    document.build(story)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="Smart_Finance_Report.pdf",
        mimetype="application/pdf"
    )



@app.route("/budget", methods=["GET", "POST"])
def budget():

    if "user_id" not in session:
        return redirect(url_for("login"))

    conn = sqlite3.connect("finance.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    # ==========================================
    # SAVE MONTHLY BUDGET
    # ==========================================

    if request.method == "POST":

        monthly_budget = float(request.form.get("monthly_budget", 0))

        cursor.execute("""

            SELECT id

            FROM budget

            WHERE user_id=?

        """,(session["user_id"],))

        existing = cursor.fetchone()

        if existing:

            cursor.execute("""

                UPDATE budget

                SET monthly_budget=?

                WHERE user_id=?

            """,(monthly_budget, session["user_id"]))

        else:

            cursor.execute("""

                INSERT INTO budget

                (user_id,monthly_budget)

                VALUES(?,?)

            """,(session["user_id"], monthly_budget))

        conn.commit()

    # ==========================================
    # CURRENT MONTHLY BUDGET
    # ==========================================

    cursor.execute("""

        SELECT monthly_budget

        FROM budget

        WHERE user_id=?

    """,(session["user_id"],))

    row = cursor.fetchone()

    if row:

        current_budget = float(row["monthly_budget"])

    else:

        current_budget = 0

    # ==========================================
    # TOTAL EXPENSE
    # ==========================================

    cursor.execute("""

        SELECT IFNULL(SUM(amount),0)

        FROM expense

        WHERE user_id=?

    """,(session["user_id"],))

    total_expense = float(cursor.fetchone()[0])

    # ==========================================
    # REMAINING BUDGET
    # ==========================================

    remaining_budget = max(0, current_budget - total_expense)

    # ==========================================
    # BUDGET USED %
    # ==========================================

    if current_budget > 0:

        budget_used = round(

            (total_expense/current_budget)*100,

            1

        )

    else:

        budget_used = 0

        # ==========================================
    # CATEGORY BUDGETS
    # ==========================================

    cursor.execute("""

        SELECT *

        FROM category_budget

        WHERE user_id=?

        ORDER BY category ASC

    """,(session["user_id"],))

    category_budgets = cursor.fetchall()

    # ==========================================
    # CATEGORY ANALYSIS
    # ==========================================

    category_analysis = []

    for row in category_budgets:

        category = row["category"]

        budget = float(row["budget"])

        cursor.execute("""

            SELECT IFNULL(SUM(amount),0)

            FROM expense

            WHERE user_id=?

            AND category=?

        """,(session["user_id"],category))

        spent = float(cursor.fetchone()[0])

        remaining = budget - spent

        if spent > budget:

            status = "Over Budget"

        elif spent >= budget*0.80:

            status = "Warning"

        else:

            status = "Healthy"

        category_analysis.append({

            "category":category,

            "budget":budget,

            "spent":spent,

            "remaining":remaining,

            "status":status

        })

    # ==========================================
    # AI RECOMMENDATIONS
    # ==========================================

    recommendations=[]

    for item in category_analysis:

        if item["status"]=="Over Budget":

            recommendations.append(

                f"⚠ You have exceeded the budget for {item['category']}."

            )

        elif item["status"]=="Warning":

            recommendations.append(

                f"🟡 {item['category']} is close to its budget limit."

            )

        else:

            recommendations.append(

                f"✅ {item['category']} spending is healthy."

            )

    # ==========================================
    # PIE CHART DATA
    # ==========================================

    chart_labels=[]

    chart_budget=[]

    chart_spent=[]

    for item in category_analysis:

        chart_labels.append(item["category"])

        chart_budget.append(item["budget"])

        chart_spent.append(item["spent"])

        # ==========================================
    # FINANCIAL HEALTH SCORE
    # ==========================================

    health_score = 100

    if budget_used > 100:

        health_score -= 30

    elif budget_used > 90:

        health_score -= 15

    elif budget_used > 80:

        health_score -= 5

    over_budget_count = 0

    for item in category_analysis:

        if item["status"] == "Over Budget":

            over_budget_count += 1

    health_score -= over_budget_count * 10

    if health_score < 0:

        health_score = 0

    if health_score > 100:

        health_score = 100

    if health_score >= 90:

        health_status = "🟢 Excellent"

    elif health_score >= 75:

        health_status = "🟢 Good"

    elif health_score >= 60:

        health_status = "🟡 Average"

    elif health_score >= 40:

        health_status = "🟠 Poor"

    else:

        health_status = "🔴 Critical"

    # ==========================================
    # MONTHLY FORECAST
    # ==========================================

    from datetime import datetime
    import calendar

    today = datetime.today()

    days_passed = today.day

    total_days = calendar.monthrange(
        today.year,
        today.month
    )[1]

    if days_passed > 0:

        predicted_expense = round(

            (total_expense / days_passed) * total_days,

            2

        )

    else:

        predicted_expense = total_expense

    predicted_savings = round(

        current_budget - predicted_expense,

        2

    )

    if predicted_expense > current_budget:

        forecast_status = "🔴 Budget Likely To Exceed"

    elif predicted_expense >= current_budget * 0.90:

        forecast_status = "🟠 Close To Budget Limit"

    else:

        forecast_status = "🟢 Spending Under Control"

    # ==========================================
    # AI SPENDING PERSONALITY
    # ==========================================

    if budget_used <= 50:

        spending_personality = "💎 Smart Saver"

        personality_message = (

            "Excellent! You save more than you spend."

        )

    elif budget_used <= 80:

        spending_personality = "🙂 Balanced Spender"

        personality_message = (

            "Your spending is well managed."

        )

    elif budget_used <= 100:

        spending_personality = "😐 Careful Spender"

        personality_message = (

            "Monitor your spending closely."

        )

    else:

        spending_personality = "🔥 High Risk Spender"

        personality_message = (

            "Your expenses are exceeding your planned budget."

        )

    # ==========================================
    # DAILY FINANCE QUOTE
    # ==========================================

    import random

    finance_quotes = [

        "Save before you spend.",

        "A budget is telling your money where to go.",

        "Small savings today become big wealth tomorrow.",

        "Never spend more than you earn.",

        "Track every expense to achieve financial freedom.",

        "Financial discipline creates financial independence.",

        "Invest in your future before buying luxuries."

    ]

    todays_quote = random.choice(finance_quotes)

    conn.close()

    return render_template(

        "budget.html",

        current_budget=current_budget,
        total_expense=total_expense,
        remaining_budget=remaining_budget,
        budget_used=budget_used,

        category_budgets=category_budgets,
        category_analysis=category_analysis,

        recommendations=recommendations,

        health_score=health_score,
        health_status=health_status,

        predicted_expense=predicted_expense,
        predicted_savings=predicted_savings,
        forecast_status=forecast_status,

        spending_personality=spending_personality,
        personality_message=personality_message,

        todays_quote=todays_quote,

        chart_labels=chart_labels,
        chart_budget=chart_budget,
        chart_spent=chart_spent

    )


@app.route("/save_category_budget", methods=["POST"])
def save_category_budget():

    if "user_id" not in session:
        return redirect(url_for("login"))

    category = request.form["category"]
    budget = request.form["budget"]

    conn = sqlite3.connect("finance.db")
    cursor = conn.cursor()

    # Check if this category already exists
    cursor.execute("""
        SELECT id
        FROM category_budget
        WHERE user_id=?
        AND category=?
    """,(session["user_id"],category))

    existing = cursor.fetchone()

    if existing:

        cursor.execute("""
            UPDATE category_budget
            SET budget=?
            WHERE user_id=?
            AND category=?
        """,(budget,
             session["user_id"],
             category))

    else:

        cursor.execute("""
            INSERT INTO category_budget

            (user_id,category,budget)

            VALUES(?,?,?)
        """,(session["user_id"],
             category,
             budget))

    conn.commit()

    conn.close()

    return redirect(url_for("budget"))



@app.route("/export_csv")
def export_csv():

    if "user_id" not in session:
        return redirect(url_for("login"))

    conn = sqlite3.connect("finance.db")
    cursor = conn.cursor()

    cursor.execute("""

        SELECT category,
               amount,
               description,
               date

        FROM expense

        WHERE user_id=?

    """,(session["user_id"],))

    rows=cursor.fetchall()

    conn.close()

    filename="expenses.csv"

    with open(filename,"w",newline="",encoding="utf-8") as file:

        writer=csv.writer(file)

        writer.writerow([
            "Category",
            "Amount",
            "Description",
            "Date"
        ])

        writer.writerows(rows)

    return send_file(
        filename,
        as_attachment=True
    )




@app.route("/delete_transactions")
def delete_transactions():

    if "user_id" not in session:
        return redirect(url_for("login"))

    conn=sqlite3.connect("finance.db")

    cursor=conn.cursor()

    cursor.execute("""

        DELETE FROM expense

        WHERE user_id=?

    """,(session["user_id"],))

    conn.commit()

    conn.close()

    flash("All transactions deleted successfully.","success")

    return redirect(url_for("settings"))


@app.route("/reset_budget")
def reset_budget():

    if "user_id" not in session:
        return redirect(url_for("login"))

    conn=sqlite3.connect("finance.db")

    cursor=conn.cursor()

    cursor.execute("""

        UPDATE budget

        SET monthly_budget=0

        WHERE user_id=?

    """,(session["user_id"],))

    cursor.execute("""

        DELETE FROM category_budget

        WHERE user_id=?

    """,(session["user_id"],))

    conn.commit()

    conn.close()

    flash("Budget reset successfully.","success")

    return redirect(url_for("settings"))


@app.route("/delete_account", methods=["POST"])
def delete_account():

    if "user_id" not in session:
        return redirect(url_for("login"))

    uid=session["user_id"]

    conn=sqlite3.connect("finance.db")

    cursor=conn.cursor()

    cursor.execute(

        "DELETE FROM expense WHERE user_id=?",

        (uid,)
    )

    cursor.execute(

        "DELETE FROM income WHERE user_id=?",

        (uid,)
    )

    cursor.execute(

        "DELETE FROM budget WHERE user_id=?",

        (uid,)
    )

    cursor.execute(

        "DELETE FROM category_budget WHERE user_id=?",

        (uid,)
    )

    cursor.execute(

        "DELETE FROM user_settings WHERE user_id=?",

        (uid,)
    )

    cursor.execute(

        "DELETE FROM users WHERE id=?",

        (uid,)
    )

    conn.commit()

    conn.close()

    session.clear()

    return redirect(url_for("register"))



@app.route("/reset_preferences")
def reset_preferences():

    if "user_id" not in session:
        return redirect(url_for("login"))

    conn = sqlite3.connect("finance.db")
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE user_settings
        SET
            theme='light',
            currency='INR',
            remember_login=0,
            two_factor=0,
            email_alert=1,
            notifications=1,
            ai_level='Balanced'
        WHERE user_id=?
    """, (session["user_id"],))

    conn.commit()
    conn.close()

    flash("Preferences reset successfully.", "success")

    return redirect(url_for("settings"))




@app.route("/investment")
def investment():

    investments = get_all_investments()

    return render_template(
        "investment.html",
        investments=investments
    )


@app.route("/add-investment", methods=["POST"])
def add_new_investment():

    asset_type = request.form["asset_type"]

    investment_name = request.form["investment_name"]

    invested_amount = float(request.form["invested_amount"])

    current_value = float(request.form["current_value"])

    purchase_date = request.form["purchase_date"]

    notes = request.form["notes"]

    add_investment(

        asset_type,

        investment_name,

        invested_amount,

        current_value,

        purchase_date,

        notes

    )

    create_notification(
        session["user_id"],
        "Investment Added",
        f"₹{invested_amount} invested in {investment_name}.",
        "Investment",
        "Medium"
    )

    flash("Investment Added Successfully!", "success")

    return redirect("/investment")


@app.route("/delete-investment/<int:id>")
def remove_investment(id):

    delete_investment(id)

    flash("Investment Deleted Successfully!", "danger")

    return redirect("/investment")


@app.route("/edit-investment/<int:id>", methods=["POST"])
def edit_investment(id):

    asset_type = request.form["asset_type"]

    investment_name = request.form["investment_name"]

    invested_amount = float(
        request.form["invested_amount"]
    )

    current_value = float(
        request.form["current_value"]
    )

    purchase_date = request.form.get(
        "purchase_date",
        ""
    )

    notes = request.form.get(
        "notes",
        ""
    )

    update_investment(
        id,
        asset_type,
        investment_name,
        invested_amount,
        current_value,
        purchase_date,
        notes
    )

    flash(
        "Investment Updated Successfully!",
        "success"
    )

    return redirect("/investment")



@app.route("/portfolio-analytics")
def portfolio_analytics():

    if "user_id" not in session:
        return redirect(url_for("login"))

    conn = sqlite3.connect("finance.db")
    cursor = conn.cursor()

    # =====================================================
    # INVESTMENT DATA
    # =====================================================

    cursor.execute("""
        SELECT
            id,
            asset_type,
            investment_name,
            invested_amount,
            current_value,
            purchase_date,
            notes,
            created_at
        FROM investments
        ORDER BY id DESC
    """)

    investments = cursor.fetchall()

    # =====================================================
    # PORTFOLIO SUMMARY
    # =====================================================

    total_investment = sum(
        float(investment[3] or 0)
        for investment in investments
    )

    current_portfolio_value = sum(
        float(investment[4] or 0)
        for investment in investments
    )

    overall_profit_loss = (
        current_portfolio_value - total_investment
    )

    overall_roi = 0

    if total_investment > 0:
        overall_roi = (
            overall_profit_loss / total_investment
        ) * 100

    # =====================================================
    # ASSET ALLOCATION
    # =====================================================

    asset_allocation = defaultdict(float)

    for investment in investments:

        asset_type = investment[1]

        invested_amount = float(
            investment[3] or 0
        )

        asset_allocation[asset_type] += invested_amount

    asset_labels = list(asset_allocation.keys())

    asset_values = list(asset_allocation.values())

    # =====================================================
    # INVESTMENT PERFORMANCE
    # =====================================================

    investment_performance = []

    for investment in investments:

        invested = float(investment[3] or 0)

        current = float(investment[4] or 0)

        profit_loss = current - invested

        roi = 0

        if invested > 0:
            roi = (
                profit_loss / invested
            ) * 100

        if roi >= 20:
            performance = "Excellent"

        elif roi >= 10:
            performance = "Good"

        elif roi >= 0:
            performance = "Average"

        else:
            performance = "Loss"

        investment_performance.append({

            "id": investment[0],

            "asset_type": investment[1],

            "investment_name": investment[2],

            "invested_amount": invested,

            "current_value": current,

            "profit_loss": profit_loss,

            "roi": roi,

            "performance": performance

        })

    # =====================================================
    # TOP AND LOWEST PERFORMING ASSET
    # =====================================================

    top_asset = None

    lowest_asset = None

    if investment_performance:

        top_asset = max(
            investment_performance,
            key=lambda x: x["roi"]
        )

        lowest_asset = min(
            investment_performance,
            key=lambda x: x["roi"]
        )

    # =====================================================
    # DIVERSIFICATION ANALYSIS
    # =====================================================

    category_count = len(asset_allocation)

    diversification_score = min(
        category_count * 15,
        100
    )

    if diversification_score >= 70:

        diversification_status = "Well Diversified"

        diversification_message = (
            "Your investments are distributed across "
            "multiple asset categories."
        )

    elif diversification_score >= 40:

        diversification_status = "Moderate Diversification"

        diversification_message = (
            "Your portfolio has moderate diversification. "
            "Consider adding more asset classes."
        )

    else:

        diversification_status = "Low Diversification"

        diversification_message = (
            "Your portfolio is concentrated in very few "
            "asset classes. Consider diversifying."
        )

    # =====================================================
    # PORTFOLIO RISK ANALYSIS
    # =====================================================

    risk_weights = {

        "Stocks": 80,

        "Mutual Funds": 55,

        "Fixed Deposit": 15,

        "Gold": 35,

        "Bonds": 25,

        "Real Estate": 45,

        "Cryptocurrency": 100

    }

    weighted_risk = 0

    if total_investment > 0:

        for asset_type, amount in asset_allocation.items():

            allocation_percentage = (
                amount / total_investment
            )

            risk_weight = risk_weights.get(
                asset_type,
                50
            )

            weighted_risk += (
                allocation_percentage * risk_weight
            )

    risk_score = round(weighted_risk, 2)

    if risk_score >= 70:

        risk_level = "High"

        risk_message = (
            "Your portfolio contains a high proportion "
            "of high-risk investments."
        )

    elif risk_score >= 40:

        risk_level = "Moderate"

        risk_message = (
            "Your portfolio has a balanced level of "
            "investment risk."
        )

    else:

        risk_level = "Low"

        risk_message = (
            "Your portfolio currently has relatively "
            "low investment risk."
        )

    # =====================================================
    # INVESTMENT DISTRIBUTION
    # =====================================================

    largest_allocation = 0

    if total_investment > 0 and asset_values:

        largest_allocation = (
            max(asset_values) / total_investment
        ) * 100

    if largest_allocation >= 70:

        distribution_status = "Highly Concentrated"

    elif largest_allocation >= 40:

        distribution_status = "Moderately Concentrated"

    else:

        distribution_status = "Balanced Distribution"

    # =====================================================
    # FINANCIAL GOAL ANALYTICS
    # =====================================================

    cursor.execute("""
        SELECT
            id,
            goal_type,
            goal_name,
            target_amount,
            saved_amount,
            target_date
        FROM financial_goals
        WHERE user_id=?
        ORDER BY id DESC
    """, (session["user_id"],))

    goals = cursor.fetchall()

    total_goals = len(goals)

    goals_achieved = 0

    goals_on_track = 0

    goals_behind = 0

    total_goal_target = 0

    total_goal_saved = 0

    for goal in goals:

        target_amount = float(goal[3] or 0)

        saved_amount = float(goal[4] or 0)

        total_goal_target += target_amount

        total_goal_saved += saved_amount

        if saved_amount >= target_amount:

            goals_achieved += 1

        else:

            progress = 0

            if target_amount > 0:

                progress = (
                    saved_amount / target_amount
                ) * 100

            if progress >= 50:

                goals_on_track += 1

            else:

                goals_behind += 1

    overall_goal_progress = 0

    if total_goal_target > 0:

        overall_goal_progress = (
            total_goal_saved / total_goal_target
        ) * 100

    overall_goal_progress = min(
        overall_goal_progress,
        100
    )

    remaining_savings = max(
        total_goal_target - total_goal_saved,
        0
    )

    # =====================================================
    # MONTHLY PORTFOLIO GROWTH
    # =====================================================

    monthly_data = defaultdict(float)

    for investment in investments:

        date_value = (
            investment[5]
            or investment[7]
        )

        if date_value:

            try:

                clean_date = str(date_value)[:10]

                parsed_date = datetime.strptime(
                    clean_date,
                    "%Y-%m-%d"
                )

                month_label = parsed_date.strftime(
                    "%b %Y"
                )

                monthly_data[month_label] += float(
                    investment[4] or 0
                )

            except ValueError:

                pass

    growth_labels = list(monthly_data.keys())

    monthly_values = list(monthly_data.values())

    growth_values = []

    running_total = 0

    for value in monthly_values:

        running_total += value

        growth_values.append(running_total)

    # =====================================================
    # MONTHLY INVESTMENT REPORT
    # =====================================================

    current_month = datetime.now().strftime("%Y-%m")

    monthly_investments = []

    for investment in investments:

        purchase_date = investment[5]

        created_at = investment[7]

        date_value = purchase_date or created_at

        if date_value:

            if str(date_value).startswith(current_month):

                monthly_investments.append(investment)

    monthly_new_investments = len(
        monthly_investments
    )

    monthly_amount_invested = sum(

        float(investment[3] or 0)

        for investment in monthly_investments

    )

    monthly_current_value = sum(

        float(investment[4] or 0)

        for investment in monthly_investments

    )

    monthly_profit = (
        monthly_current_value
        - monthly_amount_invested
    )

    monthly_roi = 0

    if monthly_amount_invested > 0:

        monthly_roi = (
            monthly_profit
            / monthly_amount_invested
        ) * 100

    report_month = datetime.now().strftime(
        "%B %Y"
    )

    conn.close()

    # =====================================================
    # SEND DATA TO HTML
    # =====================================================

    return render_template(

        "portfolio_analytics.html",

        investments=investment_performance,

        total_investment=total_investment,

        current_portfolio_value=current_portfolio_value,

        overall_profit_loss=overall_profit_loss,

        overall_roi=overall_roi,

        asset_labels=asset_labels,

        asset_values=asset_values,

        diversification_score=diversification_score,

        diversification_status=diversification_status,

        diversification_message=diversification_message,

        top_asset=top_asset,

        lowest_asset=lowest_asset,

        risk_score=risk_score,

        risk_level=risk_level,

        risk_message=risk_message,

        distribution_status=distribution_status,

        total_goals=total_goals,

        goals_achieved=goals_achieved,

        goals_on_track=goals_on_track,

        goals_behind=goals_behind,

        overall_goal_progress=overall_goal_progress,

        remaining_savings=remaining_savings,

        growth_labels=growth_labels,

        growth_values=growth_values,

        monthly_new_investments=monthly_new_investments,

        monthly_amount_invested=monthly_amount_invested,

        monthly_current_value=monthly_current_value,

        monthly_profit=monthly_profit,

        monthly_roi=monthly_roi,

        report_month=report_month

    )

@app.route("/download-portfolio-pdf")
def download_portfolio_pdf():

    if "user_id" not in session:
        return redirect(url_for("login"))

    conn = sqlite3.connect("finance.db")
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            asset_type,
            investment_name,
            invested_amount,
            current_value
        FROM investments
        ORDER BY id DESC
    """)

    investments = cursor.fetchall()

    cursor.execute("""
        SELECT
            goal_name,
            target_amount,
            saved_amount,
            target_date
        FROM financial_goals
        WHERE user_id=?
        ORDER BY id DESC
    """, (session["user_id"],))

    goals = cursor.fetchall()

    conn.close()

    total_investment = sum(
        float(row[2] or 0)
        for row in investments
    )

    current_value = sum(
        float(row[3] or 0)
        for row in investments
    )

    profit_loss = (
        current_value - total_investment
    )

    roi = 0

    if total_investment > 0:

        roi = (
            profit_loss / total_investment
        ) * 100

    buffer = BytesIO()

    document = SimpleDocTemplate(

        buffer,

        pagesize=A4,

        rightMargin=35,

        leftMargin=35,

        topMargin=35,

        bottomMargin=35

    )

    styles = getSampleStyleSheet()

    title_style = styles["Title"]

    title_style.alignment = TA_CENTER

    story = []

    story.append(

        Paragraph(
            "Finance Analytics Platform",
            title_style
        )

    )

    story.append(

        Paragraph(
            "Portfolio Analytics Report",
            styles["Heading2"]
        )

    )

    story.append(

        Paragraph(

            "Generated on: "
            + datetime.now().strftime(
                "%d %B %Y"
            ),

            styles["Normal"]

        )

    )

    story.append(Spacer(1, 20))

    story.append(

        Paragraph(
            "Portfolio Summary",
            styles["Heading2"]
        )

    )

    summary_data = [

        ["Metric", "Value"],

        [
            "Total Investment",
            f"Rs. {total_investment:,.2f}"
        ],

        [
            "Current Portfolio Value",
            f"Rs. {current_value:,.2f}"
        ],

        [
            "Overall Profit / Loss",
            f"Rs. {profit_loss:,.2f}"
        ],

        [
            "Overall ROI",
            f"{roi:.2f}%"
        ]

    ]

    summary_table = Table(

        summary_data,

        colWidths=[250, 200]

    )

    summary_table.setStyle(

        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#1f2937")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                1,
                colors.grey
            ),

            (
                "PADDING",
                (0, 0),
                (-1, -1),
                8
            )

        ])

    )

    story.append(summary_table)

    story.append(Spacer(1, 25))

    story.append(

        Paragraph(
            "Investment Performance",
            styles["Heading2"]
        )

    )

    investment_data = [[

        "Asset",

        "Investment",

        "Invested",

        "Current",

        "ROI"

    ]]

    for row in investments:

        invested = float(row[2] or 0)

        current = float(row[3] or 0)

        investment_roi = 0

        if invested > 0:

            investment_roi = (
                (current - invested)
                / invested
            ) * 100

        investment_data.append([

            row[0],

            row[1],

            f"Rs. {invested:,.2f}",

            f"Rs. {current:,.2f}",

            f"{investment_roi:.2f}%"

        ])

    investment_table = Table(

        investment_data,

        repeatRows=1

    )

    investment_table.setStyle(

        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#146ef5")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                8
            ),

            (
                "PADDING",
                (0, 0),
                (-1, -1),
                6
            )

        ])

    )

    story.append(investment_table)

    story.append(Spacer(1, 25))

    story.append(

        Paragraph(
            "Financial Goal Summary",
            styles["Heading2"]
        )

    )

    goal_data = [[

        "Goal",

        "Target",

        "Saved",

        "Progress",

        "Target Date"

    ]]

    for goal in financial_goals:

        target = float(goal[1] or 0)

        saved = float(goal[2] or 0)

        progress = 0

        if target > 0:

            progress = (
                saved / target
            ) * 100

        goal_data.append([

            goal[0],

            f"Rs. {target:,.2f}",

            f"Rs. {saved:,.2f}",

            f"{min(progress, 100):.2f}%",

            goal[3]

        ])

    goal_table = Table(

        goal_data,

        repeatRows=1

    )

    goal_table.setStyle(

        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#16a34a")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                8
            ),

            (
                "PADDING",
                (0, 0),
                (-1, -1),
                6
            )

        ])

    )

    story.append(goal_table)

    story.append(Spacer(1, 25))

    story.append(

        Paragraph(

            "Finance Analytics Platform | "
            "Portfolio Analytics Report",

            styles["Normal"]

        )

    )

    document.build(story)

    buffer.seek(0)

    return send_file(

        buffer,

        as_attachment=True,

        download_name="Smart_Finance_Portfolio_Report.pdf",

        mimetype="application/pdf"

    )


@app.route("/download-portfolio-excel")
def download_portfolio_excel():

    if "user_id" not in session:
        return redirect(url_for("login"))

    conn = sqlite3.connect("finance.db")
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            asset_type,
            investment_name,
            invested_amount,
            current_value,
            purchase_date
        FROM investments
        ORDER BY id DESC
    """)

    investments = cursor.fetchall()

    cursor.execute("""
        SELECT
            goal_type,
            goal_name,
            target_amount,
            saved_amount,
            target_date
        FROM financial_goals
        WHERE user_id=?
        ORDER BY id DESC
    """, (session["user_id"],))

    goals = cursor.fetchall()

    conn.close()

    workbook = Workbook()

    # =====================================================
    # PORTFOLIO SUMMARY SHEET
    # =====================================================

    summary_sheet = workbook.active

    summary_sheet.title = "Portfolio Summary"

    total_investment = sum(

        float(row[2] or 0)

        for row in investments

    )

    current_value = sum(

        float(row[3] or 0)

        for row in investments

    )

    profit_loss = (
        current_value - total_investment
    )

    roi = 0

    if total_investment > 0:

        roi = (
            profit_loss / total_investment
        ) * 100

    summary_sheet.append([

        "Finance Analytics Platform"

    ])

    summary_sheet.append([

        "Portfolio Analytics Report"

    ])

    summary_sheet.append([])

    summary_sheet.append([

        "Metric",

        "Value"

    ])

    summary_sheet.append([

        "Total Investment",

        total_investment

    ])

    summary_sheet.append([

        "Current Portfolio Value",

        current_value

    ])

    summary_sheet.append([

        "Overall Profit / Loss",

        profit_loss

    ])

    summary_sheet.append([

        "Overall ROI",

        roi

    ])


    # =====================================================
    # INVESTMENTS SHEET
    # =====================================================

    investment_sheet = workbook.create_sheet(

        "Investments"

    )

    investment_sheet.append([

        "Asset Type",

        "Investment Name",

        "Invested Amount",

        "Current Value",

        "Profit / Loss",

        "ROI",

        "Purchase Date"

    ])

    for row in investments:

        invested = float(row[2] or 0)

        current = float(row[3] or 0)

        profit = current - invested

        investment_roi = 0

        if invested > 0:

            investment_roi = (
                profit / invested
            ) * 100

        investment_sheet.append([

            row[0],

            row[1],

            invested,

            current,

            profit,

            investment_roi,

            row[4]

        ])


    # =====================================================
    # FINANCIAL GOALS SHEET
    # =====================================================

    goal_sheet = workbook.create_sheet(

        "Financial Goals"

    )

    goal_sheet.append([

        "Goal Type",

        "Goal Name",

        "Target Amount",

        "Saved Amount",

        "Remaining Amount",

        "Progress",

        "Target Date"

    ])

    for goal in financial_goals:

        target = float(goal[2] or 0)

        saved = float(goal[3] or 0)

        remaining = max(
            target - saved,
            0
        )

        progress = 0

        if target > 0:

            progress = (
                saved / target
            ) * 100

        goal_sheet.append([

            goal[0],

            goal[1],

            target,

            saved,

            remaining,

            min(progress, 100),

            goal[4]

        ])


    # =====================================================
    # STYLE ALL SHEETS
    # =====================================================

    for sheet in workbook.worksheets:

        for cell in sheet[1]:

            cell.font = Font(

                bold=True,

                color="FFFFFF"

            )

            cell.fill = PatternFill(

                start_color="1F4E78",

                end_color="1F4E78",

                fill_type="solid"

            )

            cell.alignment = Alignment(

                horizontal="center"

            )

        for column in sheet.columns:

            max_length = 0

            column_letter = column[0].column_letter

            for cell in column:

                try:

                    if cell.value:

                        max_length = max(

                            max_length,

                            len(str(cell.value))

                        )

                except:

                    pass

            sheet.column_dimensions[
                column_letter
            ].width = max_length + 5


    # =====================================================
    # SAVE EXCEL IN MEMORY
    # =====================================================

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return send_file(

        output,

        as_attachment=True,

        download_name="Smart_Finance_Portfolio_Report.xlsx",

        mimetype=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )

    )



@app.route("/asset-allocation")
def asset_allocation():

    if "user_id" not in session:
        return redirect(url_for("login"))


    conn = sqlite3.connect("finance.db")

    conn.row_factory = sqlite3.Row

    cursor = conn.cursor()


    # =====================================================
    # FETCH ALL INVESTMENTS
    # =====================================================

    cursor.execute("""

        SELECT
            id,
            asset_type,
            investment_name,
            invested_amount,
            current_value,
            purchase_date,
            notes

        FROM investments

        ORDER BY id DESC

    """)


    investments = cursor.fetchall()


    # =====================================================
    # TOTAL INVESTMENT
    # =====================================================

    total_investment = sum(

        float(investment["invested_amount"])

        for investment in investments

    )


    # =====================================================
    # CURRENT PORTFOLIO VALUE
    # =====================================================

    current_value = sum(

        float(investment["current_value"])

        for investment in investments

    )


    # =====================================================
    # TOTAL PROFIT / LOSS
    # =====================================================

    total_profit = current_value - total_investment


    # =====================================================
    # OVERALL ROI
    # =====================================================

    overall_roi = 0

    if total_investment > 0:

        overall_roi = round(

            (total_profit / total_investment) * 100,

            2

        )


    # =====================================================
    # INDIVIDUAL INVESTMENT RETURN CALCULATION
    # =====================================================

    investment_returns = []


    for investment in investments:

        invested_amount = float(

            investment["invested_amount"]

        )

        investment_current_value = float(

            investment["current_value"]

        )


        profit = (

            investment_current_value

            - invested_amount

        )


        roi = 0


        if invested_amount > 0:

            roi = round(

                (profit / invested_amount) * 100,

                2

            )


        investment_returns.append({

            "asset_type":
                investment["asset_type"],

            "investment_name":
                investment["investment_name"],

            "invested_amount":
                invested_amount,

            "current_value":
                investment_current_value,

            "profit":
                profit,

            "roi":
                roi

        })


    # =====================================================
    # ASSET ALLOCATION
    # =====================================================

    asset_allocation_data = {}


    for investment in investments:

        asset_type = investment["asset_type"]

        invested_amount = float(

            investment["invested_amount"]

        )


        if asset_type in asset_allocation_data:

            asset_allocation_data[asset_type] += invested_amount

        else:

            asset_allocation_data[asset_type] = invested_amount


    # =====================================================
    # ASSET SUMMARY
    # =====================================================

    asset_summary = []


    for asset_type, amount in asset_allocation_data.items():

        percentage = 0


        if total_investment > 0:

            percentage = round(

                (amount / total_investment) * 100,

                2

            )


        asset_summary.append({

            "asset_type":
                asset_type,

            "amount":
                amount,

            "percentage":
                percentage

        })


    # =====================================================
    # CHART COLORS
    # =====================================================

    chart_colors = [

        "#1265e8",

        "#16a34a",

        "#f59e0b",

        "#7c3aed",

        "#ef4444",

        "#06b6d4",

        "#ec4899"

    ]


    # =====================================================
    # CHART DATA
    # =====================================================

    asset_chart_data = {

        "labels":
            list(asset_allocation_data.keys()),

        "values":
            list(asset_allocation_data.values()),

        "colors":
            chart_colors[
                :len(asset_allocation_data)
            ]

    }


    conn.close()


    # =====================================================
    # RENDER MODULE 2 PAGE
    # =====================================================

    return render_template(

        "asset_allocation.html",

        total_investment=total_investment,

        current_value=current_value,

        total_profit=total_profit,

        overall_roi=overall_roi,

        investment_returns=investment_returns,

        asset_summary=asset_summary,

        asset_chart_data=asset_chart_data,

        chart_colors=chart_colors

    )



@app.route("/analytics")
def spending_analysis():

    # -------------------------------
    # Check Login
    # -------------------------------
    if "user_id" not in session:
        flash("Please login first.", "warning")
        return redirect(url_for("login"))

    user_id = session["user_id"]

    # -------------------------------
    # Database Connection
    # -------------------------------
    conn = sqlite3.connect("finance.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    # -------------------------------
    # Total Expenses
    # -------------------------------
    cursor.execute("""
        SELECT COALESCE(SUM(amount),0) AS total_expense
        FROM expense
        WHERE user_id = ?
    """, (user_id,))

    total_expense = cursor.fetchone()["total_expense"]

    # -------------------------------
    # Monthly Budget
    # -------------------------------
    cursor.execute("""
        SELECT monthly_budget
        FROM budget
        WHERE user_id = ?
        LIMIT 1
    """, (user_id,))

    budget = cursor.fetchone()

    if budget:
        monthly_budget = budget["monthly_budget"]
    else:
        monthly_budget = 0

    remaining_budget = monthly_budget - total_expense

    # -------------------------------
    # Category Wise Spending
    # -------------------------------
    cursor.execute("""
        SELECT
            category,
            SUM(amount) AS total
        FROM expense
        WHERE user_id = ?
        GROUP BY category
        ORDER BY total DESC
    """, (user_id,))

    category_rows = cursor.fetchall()

    category_data = []

    labels = []
    amounts = []

    highest_category = "N/A"
    highest_percentage = 0

    for row in category_rows:

        category = row["category"]
        amount = float(row["total"])

        percentage = 0

        if total_expense > 0:
            percentage = round((amount / total_expense) * 100, 1)

        if percentage >= 30:
            status = "High"
        elif percentage >= 15:
            status = "Medium"
        else:
            status = "Low"

        category_data.append({
            "category": category,
            "amount": amount,
            "percentage": percentage,
            "status": status
        })

        labels.append(category)
        amounts.append(amount)

    if len(category_data) > 0:
        highest_category = category_data[0]["category"]
        highest_percentage = category_data[0]["percentage"]

    # -------------------------------
    # Monthly Expense Trend
    # -------------------------------
    cursor.execute("""
        SELECT
            strftime('%m', date) AS month,
            SUM(amount) AS total
        FROM expense
        WHERE user_id = ?
        GROUP BY month
        ORDER BY month
    """, (user_id,))

    month_rows = cursor.fetchall()

    month_names = {
        "01": "Jan",
        "02": "Feb",
        "03": "Mar",
        "04": "Apr",
        "05": "May",
        "06": "Jun",
        "07": "Jul",
        "08": "Aug",
        "09": "Sep",
        "10": "Oct",
        "11": "Nov",
        "12": "Dec"
    }

    monthly_labels = []
    monthly_values = []

    for row in month_rows:

        monthly_labels.append(
            month_names.get(row["month"], row["month"])
        )

        monthly_values.append(float(row["total"]))

    # -------------------------------
    # Budget Utilization
    # -------------------------------
    if monthly_budget > 0:
        budget_utilization = round(
            (total_expense / monthly_budget) * 100,
            1
        )
    else:
        budget_utilization = 0

    # -------------------------------
    # Recommendations
    # -------------------------------
    recommendations = []

    if highest_category != "N/A":
        recommendations.append(
            f"You spend the most on {highest_category} ({highest_percentage}%)."
        )

    if budget_utilization >= 90:
        recommendations.append(
            "Your budget utilization is above 90%. Consider reducing discretionary expenses."
        )
    elif budget_utilization >= 75:
        recommendations.append(
            "You have used most of your monthly budget."
        )
    else:
        recommendations.append(
            "Your spending is currently within your planned budget."
        )

    if remaining_budget < 0:
        recommendations.append(
            "Warning: Your expenses have exceeded your monthly budget."
        )
    elif remaining_budget > 0:
        recommendations.append(
            f"You still have ₹{remaining_budget:,.2f} remaining this month."
        )

    conn.close()

    return render_template(
        "analytics.html",

        total_expense=total_expense,

        monthly_budget=monthly_budget,

        remaining_budget=remaining_budget,

        highest_category=highest_category,

        highest_percentage=highest_percentage,

        category_data=category_data,

        labels=labels,

        amounts=amounts,

        monthly_labels=monthly_labels,

        monthly_values=monthly_values,

        budget_utilization=budget_utilization,

        recommendations=recommendations
    )



@app.route("/financial_health")
def financial_health():
        if "user_id" not in session:
            return redirect(url_for("login"))

        conn = sqlite3.connect("finance.db")
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        user_id = session["user_id"]

        # ==========================================
        # TOTAL INCOME
        # ==========================================

        cursor.execute("""
            SELECT IFNULL(SUM(amount),0)
            FROM income
            WHERE user_id=?
        """, (user_id,))

        total_income = cursor.fetchone()[0]

        # ==========================================
        # TOTAL EXPENSE
        # ==========================================

        cursor.execute("""
            SELECT IFNULL(SUM(amount),0)
            FROM expense
            WHERE user_id=?
        """, (user_id,))

        total_expense = cursor.fetchone()[0]

        # ==========================================
        # TOTAL SAVINGS
        # ==========================================

        total_savings = total_income - total_expense

        # ==========================================
        # TOTAL INVESTMENT
        # ==========================================

        cursor.execute("""
            SELECT IFNULL(SUM(current_value),0)
            FROM investments
            WHERE user_id=?
        """, (user_id,))

        total_investment = cursor.fetchone()[0]

        # ==========================================
        # FINANCIAL GOALS
        # ==========================================

        cursor.execute("""
            SELECT
                IFNULL(SUM(target_amount),0) AS target_amount,
                IFNULL(SUM(saved_amount),0) AS saved_amount
            FROM financial_goals
            WHERE user_id=?
        """, (user_id,))

        goal = cursor.fetchone()

        goal_target = goal["target_amount"] if goal else 0
        goal_saved = goal["saved_amount"] if goal else 0

        # ==========================================
        # MONTHLY BUDGET
        # ==========================================

        cursor.execute("""
            SELECT monthly_budget
            FROM budget
            WHERE user_id=?
            ORDER BY id DESC
            LIMIT 1
        """, (user_id,))

        row = cursor.fetchone()

        if row:
            monthly_budget = row["monthly_budget"]
        else:
            monthly_budget = 0

        # ==========================================
        # LAST INCOME
        # ==========================================

        cursor.execute("""
            SELECT amount
            FROM income
            WHERE user_id=?
            ORDER BY id DESC
            LIMIT 1
        """, (user_id,))

        row = cursor.fetchone()

        last_income = row["amount"] if row else 0

        # ==========================================
        # LAST EXPENSE
        # ==========================================

        cursor.execute("""
            SELECT amount
            FROM expense
            WHERE user_id=?
            ORDER BY id DESC
            LIMIT 1
        """, (user_id,))

        row = cursor.fetchone()

        last_expense = row["amount"] if row else 0

        # ==========================================
        # DEBT (OPTIONAL)
        # ==========================================

        debt = 0

        try:

            cursor.execute("""
                SELECT IFNULL(SUM(outstanding_amount),0)
                FROM loans
                WHERE user_id=?
            """, (user_id,))

            debt = cursor.fetchone()[0]

        except Exception:

            debt = 0
                # ==========================================
        # SAVINGS RATIO
        # ==========================================

        if total_income > 0:

            savings_ratio = round(
                (total_savings / total_income) * 100,
                1
            )

            expense_ratio = round(
                (total_expense / total_income) * 100,
                1
            )

            investment_ratio = round(
                (total_investment / total_income) * 100,
                1
            )

            debt_ratio = round(
                (debt / total_income) * 100,
                1
            )

        else:

            savings_ratio = 0
            expense_ratio = 0
            investment_ratio = 0
            debt_ratio = 0

        # ==========================================
        # FINANCIAL HEALTH SCORE
        # ==========================================

        health_score = 100

        # Savings Score

        if savings_ratio >= 40:
            pass

        elif savings_ratio >= 25:
            health_score -= 5

        elif savings_ratio >= 15:
            health_score -= 15

        elif savings_ratio >= 5:
            health_score -= 25

        else:
            health_score -= 40

        # Expense Score

        if expense_ratio >= 90:
            health_score -= 25

        elif expense_ratio >= 80:
            health_score -= 15

        elif expense_ratio >= 70:
            health_score -= 8

        # Debt Score

        if debt_ratio >= 50:
            health_score -= 20

        elif debt_ratio >= 30:
            health_score -= 10

        # Investment Bonus

        if investment_ratio >= 30:
            health_score += 5

        elif investment_ratio >= 15:
            health_score += 2

        health_score = max(0, min(100, round(health_score)))

        # ==========================================
        # HEALTH STATUS
        # ==========================================

        if health_score >= 90:

            health_status = "Excellent"

        elif health_score >= 75:

            health_status = "Good"

        elif health_score >= 60:

            health_status = "Fair"

        else:

            health_status = "Poor"

        # ==========================================
        # AI RECOMMENDATIONS
        # ==========================================

        recommendations = []

        if savings_ratio < 20:

            recommendations.append(
                "Increase your monthly savings by reducing unnecessary expenses."
            )

        if expense_ratio > 80:

            recommendations.append(
                "Your expenses are high. Review discretionary spending."
            )

        if investment_ratio < 15:

            recommendations.append(
                "Consider investing more to build long-term wealth."
            )

        if debt_ratio > 30:

            recommendations.append(
                "Reduce outstanding debt to improve financial health."
            )

        if goal_target > 0 and goal_saved < goal_target:

            remaining_goal = goal_target - goal_saved

            recommendations.append(
                f"Save another ₹{remaining_goal:,.2f} to achieve your financial goals."
            )

        if len(recommendations) == 0:

            recommendations.append(
                "Excellent financial health! Keep maintaining your current habits."
            )

                # ==========================================
        # EXPENSE CATEGORY CHART
        # ==========================================

        cursor.execute("""
            SELECT
                category,
                IFNULL(SUM(amount),0) AS total
            FROM expense
            WHERE user_id=?
            GROUP BY category
            ORDER BY total DESC
        """, (user_id,))

        expense_rows = cursor.fetchall()

        category_labels = []
        category_values = []

        for row in expense_rows:

            category_labels.append(row["category"])
            category_values.append(float(row["total"]))

        # ==========================================
        # LAST 6 MONTH SAVINGS TREND
        # ==========================================

        from datetime import datetime
        import calendar

        month_labels = []
        month_values = []

        today = datetime.today()

        for i in range(5, -1, -1):

            month = today.month - i
            year = today.year

            while month <= 0:
                month += 12
                year -= 1

            month_key = f"{year}-{month:02d}"

            cursor.execute("""
                SELECT IFNULL(SUM(amount),0)
                FROM income
                WHERE user_id=?
                AND substr(date,1,7)=?
            """,(user_id, month_key))

            income_month = cursor.fetchone()[0]

            cursor.execute("""
                SELECT IFNULL(SUM(amount),0)
                FROM expense
                WHERE user_id=?
                AND substr(date,1,7)=?
            """,(user_id, month_key))

            expense_month = cursor.fetchone()[0]

            month_labels.append(
                calendar.month_abbr[month]
            )

            month_values.append(
                round(income_month - expense_month,2)
            )

        # ==========================================
        # INVESTMENT TREND
        # ==========================================

        cursor.execute("""
            SELECT
                purchase_date,
                current_value
            FROM investments
            WHERE user_id=?
            ORDER BY purchase_date
        """, (user_id,))

        investment_rows = cursor.fetchall()

        investment_values = []

        running_total = 0

        for row in investment_rows:

            running_total += float(
                row["current_value"] or 0
            )

            investment_values.append(
                round(running_total,2)
            )

        if len(investment_values) == 0:

            investment_values = [0]

        # ==========================================
        # MATCH LABEL COUNT
        # ==========================================

        while len(month_labels) < len(investment_values):

            month_labels.append(
                f"M{len(month_labels)+1}"
            )

        while len(investment_values) < len(month_labels):

            if investment_values:

                investment_values.append(
                    investment_values[-1]
                )

            else:

                investment_values.append(0)

        # ==========================================
        # DATABASE CLOSE
        # ==========================================

        conn.close()

        # ==========================================
        # RENDER TEMPLATE
        # ==========================================

        return render_template(

            "financial_health.html",

            # ==========================
            # Summary Cards
            # ==========================

            total_income=total_income,
            total_expense=total_expense,
            total_savings=total_savings,
            total_investment=total_investment,

            # ==========================
            # Health Score
            # ==========================

            health_score=health_score,
            health_status=health_status,

            # ==========================
            # Financial Ratios
            # ==========================

            savings_ratio=savings_ratio,
            expense_ratio=expense_ratio,
            investment_ratio=investment_ratio,
            debt_ratio=debt_ratio,

            # ==========================
            # Goal Details
            # ==========================

            goal_target=goal_target,
            goal_saved=goal_saved,
            debt=debt,

            # ==========================
            # Latest Transactions
            # ==========================

            last_income=last_income,
            last_expense=last_expense,

            # ==========================
            # Charts
            # ==========================

            category_labels=category_labels,
            category_values=category_values,

            month_labels=month_labels,
            month_values=month_values,

            investment_values=investment_values,

            # ==========================
            # AI Recommendations
            # ==========================

            recommendations=recommendations

        )  



@app.route('/notification-center')
def notification_center():

    if 'user_id' not in session:
        return redirect(url_for('login'))

    user_id = session['user_id']

    conn = sqlite3.connect("finance.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    # ==========================================
    # Notifications
    # ==========================================

    cursor.execute("""
        SELECT *
        FROM notifications
        WHERE user_id=?
        ORDER BY created_at DESC
    """, (user_id,))

    notifications = cursor.fetchall()

    recent_notifications = notifications[:5]

    total_notifications = len(notifications)

    unread_notifications = sum(
        1 for n in notifications if n["is_read"] == 0
    )

    high_priority = sum(
        1 for n in notifications if n["priority"] == "High"
    )

    pending_notifications = sum(
        1 for n in notifications if n["status"] == "Pending"
    )

    completed_notifications = sum(
        1 for n in notifications if n["status"] == "Completed"
    )

    # ==========================================
    # Budget
    # ==========================================

    cursor.execute("""
        SELECT monthly_budget
        FROM budget
        WHERE user_id=?
        ORDER BY id DESC
        LIMIT 1
    """, (user_id,))

    row = cursor.fetchone()

    monthly_budget = row["monthly_budget"] if row else 0

    cursor.execute("""
        SELECT IFNULL(SUM(amount),0)
        FROM expense
        WHERE user_id=?
    """, (user_id,))

    monthly_expense = cursor.fetchone()[0]

    budget_remaining = max(
        0,
        monthly_budget - monthly_expense
    )

    budget_usage = 0

    if monthly_budget > 0:

        budget_usage = round(
            (monthly_expense / monthly_budget) * 100,
            1
        )

    # ==========================================
    # Financial Goals
    # ==========================================

    cursor.execute("""
        SELECT COUNT(*)
        FROM financial_goals
        WHERE user_id=?
        AND saved_amount < target_amount
    """, (user_id,))

    goal_reminders = cursor.fetchone()[0]

    # ==========================================
    # Investments
    # ==========================================

    cursor.execute("""
        SELECT COUNT(*)
        FROM investments
        WHERE user_id=?
    """, (user_id,))

    investment_alerts = cursor.fetchone()[0]

    # ==========================================
    # Bills
    # ==========================================

    cursor.execute("""
        SELECT
            bill_name,
            amount,
            due_date,
            status
        FROM bills
        WHERE user_id = ?
        AND status = 'Pending'
        ORDER BY due_date ASC
        LIMIT 5
    """, (user_id,))

    bills = cursor.fetchall()

    today = datetime.today().date()

    for bill in bills:

        due = datetime.strptime(
            bill["due_date"],
            "%Y-%m-%d"
        ).date()

        days = (due - today).days

        if days == 1:

            create_notification(
                session["user_id"],
                "Bill Due Tomorrow",
                f"{bill['bill_name']} is due tomorrow.",
                "Bill Reminder",
                "High"
            )

        elif days == 0:

            create_notification(
                session["user_id"],
                "Bill Due Today",
                f"{bill['bill_name']} is due today.",
                "Bill Reminder",
                "High"
            )

        upcoming_bills = len(bills)

    # ==========================================
    # Doughnut Chart
    # ==========================================

    cursor.execute("""
        SELECT type,COUNT(*)
        FROM notifications
        WHERE user_id=?
        GROUP BY type
    """, (user_id,))

    chart = cursor.fetchall()

    notification_labels = [
        x["type"] for x in chart
    ]

    notification_values = [
        x[1] for x in chart
    ]

    # ==========================================
    # Monthly Chart
    # ==========================================

    month_labels = [
        "Jan","Feb","Mar",
        "Apr","May","Jun"
    ]

    monthly_notification_values = [
        4,6,8,5,7,total_notifications
    ]

    # ==========================================
    # AI Insights
    # ==========================================

    ai_insights=[]

    if budget_usage>=80:

        ai_insights.append(
            "Your monthly budget utilization is high."
        )

    if goal_reminders>0:

        ai_insights.append(
            "Keep saving to achieve your financial goals."
        )

    if investment_alerts>0:

        ai_insights.append(
            "Review your investment portfolio regularly."
        )

    if unread_notifications>0:

        ai_insights.append(
            f"You have {unread_notifications} unread notifications."
        )

    if len(ai_insights)==0:

        ai_insights.append(
            "Everything looks great! No urgent financial alerts."
        )

    conn.close()

    return render_template(

        "notification.html",

        notifications=notifications,

        recent_notifications=recent_notifications,

        total_notifications=total_notifications,

        unread_notifications=unread_notifications,

        high_priority=high_priority,

        pending_notifications=pending_notifications,

        completed_notifications=completed_notifications,

        monthly_budget=monthly_budget,

        monthly_expense=monthly_expense,

        budget_remaining=budget_remaining,

        budget_usage=budget_usage,

        bills=bills,

        upcoming_bills=upcoming_bills,

        goal_reminders=goal_reminders,

        investment_alerts=investment_alerts,

        notification_labels=notification_labels,

        notification_values=notification_values,

        month_labels=month_labels,

        monthly_notification_values=monthly_notification_values,

        ai_insights=ai_insights

    )



@app.route('/bills')
def bills():

    if 'user_id' not in session:
        return redirect(url_for('login'))

    conn = sqlite3.connect("finance.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        SELECT *
        FROM bills
        WHERE user_id=?
        ORDER BY due_date ASC
    """,(session['user_id'],))

    bills = cursor.fetchall()

    conn.close()

    return render_template(
        "bills.html",
        bills=bills
    )


@app.route('/add_bill', methods=['POST'])
def add_bill():

    if 'user_id' not in session:
        return redirect(url_for('login'))

    bill_name = request.form['bill_name']
    category = request.form['category']
    amount = request.form['amount']
    due_date = request.form['due_date']
    frequency = request.form['frequency']
    reminder_days = request.form['reminder_days']
    notes = request.form['notes']

    conn = sqlite3.connect("finance.db")
    cursor = conn.cursor()

    cursor.execute("""

        INSERT INTO bills(

            user_id,

            bill_name,

            category,

            amount,

            due_date,

            frequency,

            reminder_days,

            notes

        )

        VALUES(?,?,?,?,?,?,?,?)

    """,(

        session['user_id'],

        bill_name,

        category,

        amount,

        due_date,

        frequency,

        reminder_days,

        notes

    ))

    conn.commit()

    create_notification(
        session["user_id"],
        "Bill Added",
        f"{bill_name} bill of ₹{amount} is due on {due_date}.",
        "Bill Reminder",
        "High"
    )

    conn.close()

    flash("Bill added successfully.","success")

    return redirect(url_for('bills'))


@app.route('/delete_bill/<int:id>')
def delete_bill(id):

    conn = sqlite3.connect("finance.db")
    cursor = conn.cursor()

    cursor.execute("""

        DELETE FROM bills

        WHERE id=?

    """,(id,))

    conn.commit()

    conn.close()

    flash("Bill deleted successfully.","success")

    return redirect(url_for('bills'))


@app.route('/mark_bill_paid/<int:id>')
def mark_bill_paid(id):

    conn = sqlite3.connect("finance.db")
    cursor = conn.cursor()

    cursor.execute("""

        UPDATE bills

        SET status='Paid'

        WHERE id=?

    """,(id,))

    conn.commit()

    create_notification(
    session["user_id"],
    "Bill Paid",
    "One of your bills has been marked as paid.",
    "Bill Reminder",
    "Low"
)

    conn.close()

    flash("Bill marked as paid.","success")

    return redirect(url_for('bills'))


@app.route('/edit_bill/<int:id>')
def edit_bill(id):

    conn = sqlite3.connect("finance.db")
    conn.row_factory = sqlite3.Row

    cursor = conn.cursor()

    cursor.execute("""

        SELECT *

        FROM bills

        WHERE id=?

    """,(id,))

    bill = cursor.fetchone()

    conn.close()

    return render_template(

        "edit_bill.html",

        bill=bill

    )


@app.route('/update_bill/<int:id>', methods=['POST'])
def update_bill(id):

    conn = sqlite3.connect("finance.db")
    cursor = conn.cursor()

    cursor.execute("""

        UPDATE bills

        SET

            bill_name=?,

            category=?,

            amount=?,

            due_date=?,

            frequency=?,

            reminder_days=?,

            notes=?

        WHERE id=?

    """,(

        request.form['bill_name'],

        request.form['category'],

        request.form['amount'],

        request.form['due_date'],

        request.form['frequency'],

        request.form['reminder_days'],

        request.form['notes'],

        id

    ))

    conn.commit()

    conn.close()

    flash("Bill updated successfully.","success")

    return redirect(url_for('bills'))


def generate_ai_insights(
        income,
        expense,
        savings,
        investment,
        budget_usage,
        goal_progress,
        category_expenses,
        monthly_data
):

    insights=[]

    # Cash Flow
    if income>0:
        expense_ratio=(expense/income)*100

        if expense_ratio>90:
            insights.append({
                "type":"danger",
                "title":"Critical Cash Flow",
                "message":"You are spending over 90% of your income. Reduce discretionary expenses."
            })
        elif expense_ratio>75:
            insights.append({
                "type":"warning",
                "title":"High Spending",
                "message":"Your expenses are consuming most of your income."
            })
        else:
            insights.append({
                "type":"success",
                "title":"Healthy Cash Flow",
                "message":"Your income comfortably covers your expenses."
            })

    # Savings
    if income>0:
        saving_rate=(savings/income)*100

        if saving_rate<20:
            insights.append({
                "type":"warning",
                "title":"Increase Savings",
                "message":f"Current savings rate is {saving_rate:.1f}%. Aim for at least 20%."
            })
        else:
            insights.append({
                "type":"success",
                "title":"Strong Savings",
                "message":"Excellent savings discipline."
            })

    # Investment
    if investment==0:
        insights.append({
            "type":"primary",
            "title":"Start Investing",
            "message":"Begin a SIP or Mutual Fund investment for long-term wealth creation."
        })
    elif investment<savings:
        insights.append({
            "type":"info",
            "title":"Grow Investments",
            "message":"Consider investing a larger portion of your savings."
        })

    # Budget
    if budget_usage>100:
        insights.append({
            "type":"danger",
            "title":"Budget Exceeded",
            "message":"You exceeded your monthly budget."
        })
    elif budget_usage>90:
        insights.append({
            "type":"warning",
            "title":"Budget Alert",
            "message":"You have used over 90% of your budget."
        })
    else:
        insights.append({
            "type":"success",
            "title":"Budget Under Control",
            "message":"Your spending remains within budget."
        })

    # Goals
    if goal_progress>=100:
        insights.append({
            "type":"success",
            "title":"Goal Achieved",
            "message":"Congratulations! Your financial goal has been achieved."
        })
    elif goal_progress>=80:
        insights.append({
            "type":"info",
            "title":"Almost There",
            "message":"You're close to completing your financial goal."
        })

    # Highest spending category
    if category_expenses:

        highest=max(category_expenses,key=lambda x:x["amount"])

        insights.append({
            "type":"info",
            "title":"Highest Spending Category",
            "message":f"You spent the most on {highest['category']} (₹{highest['amount']:,.2f})."
        })

    # Spending Trend
    if len(monthly_data)>=2:

        previous=monthly_data[-2]["expense"]
        current=monthly_data[-1]["expense"]

        if current>previous:
            insights.append({
                "type":"warning",
                "title":"Expenses Increased",
                "message":"Your spending increased compared to last month."
            })
        elif current<previous:
            insights.append({
                "type":"success",
                "title":"Expenses Reduced",
                "message":"Great! Your spending decreased this month."
            })

    return insights



@app.route("/ai-insights")
def ai_insights():

    if "user_id" not in session:
        return redirect(url_for("login"))

    conn = sqlite3.connect("finance.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    user = session["user_id"]

    # ==============================
    # TOTAL INCOME
    # ==============================

    cursor.execute("""
        SELECT COALESCE(SUM(amount),0)
        FROM income
        WHERE user_id=?
    """, (user,))
    total_income = cursor.fetchone()[0]

    # ==============================
    # TOTAL EXPENSE
    # ==============================

    cursor.execute("""
        SELECT COALESCE(SUM(amount),0)
        FROM expense
        WHERE user_id=?
    """, (user,))
    total_expense = cursor.fetchone()[0]

    # ==============================
    # MONTHLY BUDGET
    # ==============================

    cursor.execute("""
        SELECT COALESCE(monthly_budget,0)
        FROM budget
        WHERE user_id=?
        ORDER BY id DESC
        LIMIT 1
    """, (user,))

    row = cursor.fetchone()

    budget = row[0] if row else 0

    # ==============================
    # INVESTMENTS
    # ==============================

    cursor.execute("""
        SELECT COALESCE(SUM(current_value),0)
        FROM investments
        WHERE user_id=?
    """, (user,))

    investment = cursor.fetchone()[0]

    # ==============================
    # GOALS
    # ==============================

    cursor.execute("""
        SELECT
            COALESCE(SUM(target_amount),0),
            COALESCE(SUM(saved_amount),0)
        FROM financial_goals
        WHERE user_id=?
    """, (user,))

    goal = cursor.fetchone()

    total_target = goal[0]
    saved_amount = goal[1]

    goal_progress = (
        (saved_amount / total_target) * 100
        if total_target > 0 else 0
    )

    # ==============================
    # CATEGORY ANALYSIS
    # ==============================

    cursor.execute("""
        SELECT
            category,
            SUM(amount)
        FROM expense
        WHERE user_id=?
        GROUP BY category
        ORDER BY SUM(amount) DESC
    """, (user,))

    category_expenses = []

    for row in cursor.fetchall():
        category_expenses.append({
            "category": row[0],
            "amount": row[1]
        })

    # ==============================
    # MONTHLY TREND
    # ==============================

    cursor.execute("""
        SELECT
            strftime('%Y-%m', date) AS month,
            SUM(amount) AS expense
        FROM expense
        WHERE user_id=?
        GROUP BY month
        ORDER BY month
        LIMIT 6
    """, (user,))

    monthly_data = []

    for row in cursor.fetchall():
        monthly_data.append({
            "month": row[0],
            "expense": row[1]
        })

    # ==============================
    # CALCULATIONS
    # ==============================

    savings = total_income - total_expense

    budget_usage = (
        (total_expense / budget) * 100
        if budget > 0 else 0
    )

    health_score = calculate_health_score(
        total_income,
        total_expense,
        savings,
        investment,
        budget_usage
    )

    ai_cards = generate_ai_insights(
        total_income,
        total_expense,
        savings,
        investment,
        budget_usage,
        goal_progress,
        category_expenses,
        monthly_data
    )

    conn.close()

    return render_template(
        "ai_insights.html",
        total_income=total_income,
        total_expense=total_expense,
        savings=savings,
        investment=investment,
        budget=budget,
        budget_usage=round(budget_usage,2),
        health_score=health_score,
        goal_progress=round(goal_progress,2),
        ai_cards=ai_cards,
        category_expenses=category_expenses,
        monthly_data=monthly_data
    )



@app.route("/intelligence-dashboard")
def intelligence_dashboard():

    if "user_id" not in session:
        return redirect(url_for("login"))

    conn = sqlite3.connect("finance.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    user = session["user_id"]

    # ===========================================
    # TOTAL INCOME
    # ===========================================

    cursor.execute("""
        SELECT COALESCE(SUM(amount),0) AS total
        FROM income
        WHERE user_id=?
    """, (user,))

    total_income = float(cursor.fetchone()["total"])

    # ===========================================
    # TOTAL EXPENSE
    # ===========================================

    cursor.execute("""
        SELECT COALESCE(SUM(amount),0) AS total
        FROM expense
        WHERE user_id=?
    """, (user,))

    total_expense = float(cursor.fetchone()["total"])

    # ===========================================
    # MONTHLY BUDGET
    # ===========================================

    cursor.execute("""
        SELECT monthly_budget
        FROM budget
        WHERE user_id=?
        ORDER BY id DESC
        LIMIT 1
    """, (user,))

    row = cursor.fetchone()

    monthly_budget = float(row["monthly_budget"]) if row else 0

    # ===========================================
    # TOTAL INVESTMENT
    # ===========================================

    cursor.execute("""
        SELECT COALESCE(SUM(current_value),0) AS total
        FROM investments
        WHERE user_id=?
    """, (user,))

    total_investment = float(cursor.fetchone()["total"])

    # ===========================================
    # GOAL PROGRESS
    # ===========================================

    cursor.execute("""
        SELECT
            COALESCE(SUM(target_amount),0) AS target,
            COALESCE(SUM(saved_amount),0) AS saved
        FROM financial_goals
        WHERE user_id=?
    """, (user,))

    goal = cursor.fetchone()

    target = float(goal["target"])
    saved = float(goal["saved"])

    if target > 0:
        goal_progress = round((saved / target) * 100, 2)
    else:
        goal_progress = 0

    # ===========================================
    # SAVINGS
    # ===========================================

    savings = total_income - total_expense

    # ===========================================
    # BUDGET UTILIZATION
    # ===========================================

    if monthly_budget > 0:
        budget_utilization = round(
            (total_expense / monthly_budget) * 100, 2
        )
    else:
        budget_utilization = 0

    # ===========================================
    # FINANCIAL HEALTH SCORE
    # ===========================================

    health_score = calculate_health_score(
        total_income,
        total_expense,
        savings,
        total_investment,
        budget_utilization
    )

    # ===========================================
    # EXPENSE CATEGORY CHART
    # ===========================================

    cursor.execute("""
        SELECT
            category,
            SUM(amount) AS total
        FROM expense
        WHERE user_id=?
        GROUP BY category
        ORDER BY total DESC
    """, (user,))

    expense_categories = []

    for row in cursor.fetchall():

        expense_categories.append({

            "category": row["category"],

            "amount": float(row["total"])

        })

    # ===========================================
    # MONTHLY TREND
    # ===========================================

    cursor.execute("""
        SELECT
            strftime('%Y-%m', date) AS month,
            SUM(amount) AS total
        FROM expense
        WHERE user_id=?
        GROUP BY strftime('%Y-%m', date)
        ORDER BY month
    """, (user,))

    monthly_trend = []

    for row in cursor.fetchall():

        monthly_trend.append({

            "month": row["month"],

            "expense": float(row["total"])

        })

    # ===========================================
    # NOTIFICATIONS
    # ===========================================

    cursor.execute("""
        SELECT *
        FROM notifications
        WHERE user_id=?
        ORDER BY created_at DESC
        LIMIT 5
    """, (user,))

    notifications = cursor.fetchall()

    # ===========================================
    # BILLS
    # ===========================================

    cursor.execute("""
        SELECT *
        FROM bills
        WHERE user_id=?
        AND status='Pending'
        ORDER BY due_date
        LIMIT 5
    """, (user,))

    bills = cursor.fetchall()

    # ===========================================
    # AI RECOMMENDATIONS
    # ===========================================

    ai_cards = generate_ai_insights(

        total_income,

        total_expense,

        savings,

        total_investment,

        budget_utilization,

        goal_progress,

        expense_categories,

        monthly_trend

    )

    conn.close()

    return render_template(

        "intelligence_dashboard.html",

        total_income=total_income,

        total_expense=total_expense,

        savings=savings,

        total_investment=total_investment,

        monthly_budget=monthly_budget,

        budget_utilization=budget_utilization,

        health_score=health_score,

        goal_progress=goal_progress,

        expense_categories=expense_categories,

        monthly_trend=monthly_trend,

        notifications=notifications,

        bills=bills,

        ai_cards=ai_cards

    )

    

@app.route("/ai-chat")
def ai_chat():

    if "user_id" not in session:
        return redirect(url_for("login"))

    return render_template("chat_assistant.html")



# ============================================================
# AI DOCUMENT ANALYSIS
# ============================================================

@app.route("/api/document", methods=["POST"])
def api_document():

    if "user_id" not in session:
        return jsonify({"success": False}), 401

    uploaded_file = request.files.get("file")

    if not uploaded_file:
        return jsonify({
            "success": False,
            "message": "No file uploaded."
        }), 400

    result = finance_ai.process_document(uploaded_file)

    return jsonify(result)




# ============================================================
# SMART AI CHAT
# ============================================================

@app.route("/api/chat", methods=["POST"])
def api_chat():

    if "user_id" not in session:
        return jsonify({"success": False}), 401

    message = request.form.get("message", "")

    uploaded_file = request.files.get("file")

    document = None

    # ==========================================
    # Process Uploaded Document
    # ==========================================

    if uploaded_file:

        document = finance_ai.process_document(uploaded_file)

        if document.get("success"):

            document = document["data"]

            print("\n========== DOCUMENT DATA ==========")
            print(document)
            print("===================================\n")

    # ==========================================
    # Send to AI
    # ==========================================

    result = finance_ai.smart_chat(

        user_id=session["user_id"],

        message=message,

        document_data=document

    )

    # ==========================================
    # Response
    # ==========================================

    return jsonify({

        "success": result.get("success", False),

        "response": result.get("data", ""),

        "metadata": result.get("metadata"),

        "suggestions": result.get("suggestions", [])

    })
    


# ============================================================
# AI CHART GENERATOR
# ============================================================

@app.route("/api/chart", methods=["POST"])
def api_chart():

    if "user_id" not in session:
        return jsonify({"success": False}), 401

    data = request.get_json()

    result = finance_ai.generate_chart(
        user_id=session["user_id"],
        entity=data.get("entity"),
        chart_type=data.get("chart_type", "pie"),
        period=data.get("period", "all")
    )

    chart = result.get("data", {})

    return jsonify({

        "success": result.get("success", False),

        "chart": chart,

        "message": result.get("message")

    })



# ============================================================
# AI FINANCIAL REPORT
# ============================================================

@app.route("/api/report", methods=["POST"])
def api_report():

    if "user_id" not in session:
        return jsonify({"success": False}), 401

    data = request.get_json()

    result = finance_ai.generate_report(
        user_id=session["user_id"],
        report_type=data.get("report_type", "financial")
    )

    return jsonify(result)


# ============================================================
# AI QUICK SUMMARY
# ============================================================

@app.route("/api/dashboard-summary")
def dashboard_summary():

    return jsonify(
        finance_ai.dashboard_summary(
            session["user_id"]
        )
    )




# ============================================================
# AI RESET MEMORY
# ============================================================

@app.route("/api/reset", methods=["POST"])
def api_reset():

    return jsonify(
        finance_ai.reset_session(
            session["user_id"]
        )
    )




# ============================================================
# AI CHAT EXPORT
# ============================================================

@app.route("/export-chat", methods=["POST"])
def export_chat():

    if "user_id" not in session:

        return jsonify({

            "success": False,

            "message": "Please login first."

        }), 401

    try:

        data = request.get_json()

        conversation = data.get("conversation", [])

        filename = f"chat_{session['user_id']}.txt"

        filepath = os.path.join(

            app.config["UPLOAD_FOLDER"],

            filename

        )

        with open(filepath, "w", encoding="utf-8") as file:

            file.write("Finance AI Assistant CHAT\n")

            file.write("=" * 50)

            file.write("\n\n")

            for item in conversation:

                role = item.get("role", "User")

                message = item.get("message", "")

                file.write(f"{role}: {message}\n\n")

        return send_file(

            filepath,

            as_attachment=True,

            download_name="SmartFinanceChat.txt"

        )

    except Exception as e:

        return jsonify({

            "success": False,

            "message": str(e)

        }), 500



# ============================================================
# AI STATUS
# ============================================================

@app.route("/api/status")
def api_status():

    return jsonify(
        finance_ai.system_status()
    )




# ============================================================
# CHAT HISTORY
# ============================================================

@app.route("/api/history")
def api_history():

    if "user_id" not in session:
        return jsonify({
            "success": False
        }), 401

    return jsonify(
        finance_ai.chat_history(
            session["user_id"]
        )
    )



# ==========================================
# FINANCIAL CALENDAR
# ==========================================
print("Financial Calendar Route Loaded")
@app.route("/financial-calendar")
def financial_calendar():

    if "user_id" not in session:

        flash(

            "Please login first.",

            "warning"

        )

        return redirect(

            url_for("login")

        )

    return render_template(

        "financial_calendar.html"

    )


# ==========================================
# CALENDAR SUMMARY API
# ==========================================

@app.route("/api/calendar-summary")
def calendar_summary():

    if "user_id" not in session:
        return jsonify({})

    conn = sqlite3.connect("finance.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    user_id = session["user_id"]

    cursor.execute("SELECT COALESCE(SUM(amount),0) total FROM income WHERE user_id=?", (user_id,))
    total_income = cursor.fetchone()["total"]

    cursor.execute("SELECT COALESCE(SUM(amount),0) total FROM expense WHERE user_id=?", (user_id,))
    total_expense = cursor.fetchone()["total"]

    cursor.execute("SELECT COALESCE(SUM(invested_amount),0) total FROM investments WHERE user_id=?", (user_id,))
    total_investment = cursor.fetchone()["total"]

    cursor.execute("SELECT COUNT(*) total FROM financial_goals WHERE user_id=?", (user_id,))
    total_goals = cursor.fetchone()["total"]

    cursor.execute("SELECT monthly_budget FROM budget WHERE user_id=?", (user_id,))
    row = cursor.fetchone()

    monthly_budget = row["monthly_budget"] if row else 0

    savings = total_income - total_expense

    budget_percent = 0

    if monthly_budget > 0:
        budget_percent = round((total_expense / monthly_budget) * 100, 1)

    conn.close()

    return jsonify({

        "income": total_income,

        "expense": total_expense,

        "investment": total_investment,

        "goals": total_goals,

        "savings": savings,

        "budget": monthly_budget,

        "budget_percent": budget_percent

    })

# ==========================================
# CALENDAR EVENTS API
# ==========================================

@app.route("/api/calendar-events")
def calendar_events():

    if "user_id" not in session:
        return jsonify([])

    conn = sqlite3.connect("finance.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    user_id = session["user_id"]

    events = []

    # ==========================================
    # INCOME EVENTS
    # ==========================================

    cursor.execute("""
        SELECT
            source,
            amount,
            date
        FROM income
        WHERE user_id=?
    """, (user_id,))

    for row in cursor.fetchall():

        events.append({

            "title": f"💰 {row['source']} ₹{row['amount']}",

            "start": row["date"],

            "color": "#22c55e"

        })

    # ==========================================
    # EXPENSE EVENTS
    # ==========================================

    cursor.execute("""
        SELECT
            category,
            amount,
            date
        FROM expense
        WHERE user_id=?
    """, (user_id,))

    for row in cursor.fetchall():

        events.append({

            "title": f"💸 {row['category']} ₹{row['amount']}",

            "start": row["date"],

            "color": "#ef4444"

        })

    # ==========================================
    # INVESTMENT EVENTS
    # ==========================================

    cursor.execute("""
        SELECT
            investment_name,
            invested_amount,
            purchase_date
        FROM investments
        WHERE user_id=?
    """, (user_id,))

    for row in cursor.fetchall():

        if row["purchase_date"]:

            events.append({

                "title": f"📈 {row['investment_name']} ₹{row['invested_amount']}",

                "start": row["purchase_date"],

                "color": "#3b82f6"

            })

    # ==========================================
    # FINANCIAL GOAL EVENTS
    # ==========================================

    cursor.execute("""
        SELECT
            goal_name,
            target_amount,
            target_date
        FROM financial_goals
        WHERE user_id=?
    """, (user_id,))

    for row in cursor.fetchall():

        if row["target_date"]:

            events.append({

                "title": f"🎯 {row['goal_name']} ₹{row['target_amount']}",

                "start": row["target_date"],

                "color": "#a855f7"

            })

        # ==========================================
        # BUDGET ALERT
        # ==========================================

        cursor.execute("""
        SELECT
            monthly_budget
        FROM budget
        WHERE user_id=?
        """, (user_id,))

        budget = cursor.fetchone()

        cursor.execute("""
        SELECT
            SUM(amount)
        FROM expense
        WHERE user_id=?
        """, (user_id,))

        spent = cursor.fetchone()[0] or 0

        if budget:

            budget_amount = budget["monthly_budget"]

            if spent >= budget_amount * 0.80:

                from datetime import date

                events.append({

                    "title": f"⚠️ Budget Used {round((spent/budget_amount)*100)}%",

                    "start": str(date.today()),

                    "color": "#f59e0b"

                })

            if spent >= budget_amount:

                from datetime import date

                events.append({

                    "title": "🚨 Monthly Budget Exceeded",

                    "start": str(date.today()),

                    "color": "#dc2626"

                })

        # ==========================================
        # GOAL COMPLETION
        # ==========================================

        cursor.execute("""

        SELECT

        goal_name,

        target_amount,

        saved_amount,

        target_date

        FROM financial_goals

        WHERE user_id=?

        """,(user_id,))

        for goal in cursor.fetchall():

            if goal["saved_amount"] >= goal["target_amount"]:

                events.append({

                    "title": f"🎉 Goal Achieved : {goal['goal_name']}",

                    "start": goal["target_date"],

                    "color":"#10b981"

                })

        # ==========================================
        # LOW SAVINGS ALERT
        # ==========================================

        cursor.execute("""

        SELECT SUM(amount)

        FROM income

        WHERE user_id=?

        """,(user_id,))

        income = cursor.fetchone()[0] or 0

        cursor.execute("""

        SELECT SUM(amount)

        FROM expense

        WHERE user_id=?

        """,(user_id,))

        expense = cursor.fetchone()[0] or 0

        saving = income-expense

        if income>0:

            percentage=(saving/income)*100

            if percentage<20:

                from datetime import date

                events.append({

                    "title":"💳 Savings below 20%",

                    "start":str(date.today()),

                    "color":"#ef4444"

                })     

        # ==========================================
        # INVESTMENT REMINDER
        # ==========================================

        cursor.execute("""

        SELECT

        investment_name,

        purchase_date

        FROM investments

        WHERE user_id=?

        """,(user_id,))

        for item in cursor.fetchall():

            if item["purchase_date"]:

                events.append({

                    "title":f"📈 Review {item['investment_name']}",

                    "start":item["purchase_date"],

                    "color":"#2563eb"

                })
    conn.close()
    return jsonify(events)


# ==========================================
# SUBMIT FEEDBACK
# ==========================================

@app.route("/submit-feedback", methods=["POST"])
def submit_feedback():

    if "user_id" not in session:

        flash("Please login first.")

        return redirect(url_for("login"))

    rating = request.form.get("rating")

    experience = request.form.get("experience")

    feedback = request.form.get("feedback")

    email = request.form.get("email")

    conn = sqlite3.connect("finance.db")

    cursor = conn.cursor()

    cursor.execute("""

        INSERT INTO feedback(

            user_id,

            rating,

            experience,

            feedback,

            email

        )

        VALUES(?,?,?,?,?)

    """,(

        session["user_id"],

        rating,

        experience,

        feedback,

        email

    ))

    conn.commit()

    conn.close()

    flash(

        "Thank you for your valuable feedback!",

        "success"

    )

    return redirect(url_for("dashboard"))

# ==========================================
# VIEW FEEDBACK (Developer)
# ==========================================

@app.route("/feedbacks")
def feedbacks():

    conn = sqlite3.connect("finance.db")

    conn.row_factory = sqlite3.Row

    cursor = conn.cursor()

    cursor.execute("""

        SELECT

            users.name,

            feedback.*

        FROM feedback

        LEFT JOIN users

        ON feedback.user_id = users.id

        ORDER BY created_at DESC

    """)

    feedbacks = cursor.fetchall()

    conn.close()

    return render_template(

        "feedbacks.html",

        feedbacks=feedbacks

    )


@app.route("/dataset_management", methods=["GET", "POST"])
def dataset_management():

    if "user_id" not in session:
        return redirect(url_for("login"))

    user_id = session["user_id"]

    # --------------------------------------------
    # Get current dataset info
    # --------------------------------------------

    dataset = get_dataset_info(user_id)

    history = get_dataset_history(user_id)

    preview = None
    columns = []

    uploaded_filename = None
    total_rows = 0
    total_columns = 0

    # --------------------------------------------
    # Upload CSV
    # --------------------------------------------

    if request.method == "POST":

        if "dataset" not in request.files:

            flash(
                "Please choose a CSV file.",
                "warning"
            )

            return redirect(
                url_for("dataset_management")
            )

        file = request.files["dataset"]

        if file.filename == "":

            flash(
                "No file selected.",
                "warning"
            )

            return redirect(
                url_for("dataset_management")
            )

        if not allowed_dataset(file.filename):

            flash(
                "Only CSV files are allowed.",
                "danger"
            )

            return redirect(
                url_for("dataset_management")
            )

        filename = secure_filename(file.filename)

        upload_folder = app.config["DATASET_UPLOAD_FOLDER"]

        if not os.path.exists(upload_folder):
            os.makedirs(upload_folder)

        filepath = os.path.join(
            upload_folder,
            filename
        )

        file.save(filepath)

        try:

            df = pd.read_csv(filepath)

        except Exception as e:

            flash(
                f"Unable to read CSV file : {e}",
                "danger"
            )

            return redirect(
                url_for("dataset_management")
            )

        # ----------------------------------------
        # Clean column names
        # ----------------------------------------

        df.columns = df.columns.str.strip()

        preview = df.head(10).values.tolist()

        columns = df.columns.tolist()

        uploaded_filename = filename

        total_rows = len(df)

        total_columns = len(df.columns)

        # ----------------------------------------
        # Save in session
        # ----------------------------------------

        session["dataset_path"] = filepath

        session["dataset_name"] = filename

        session["uploaded_filename"] = filename

        flash(
            "Dataset uploaded successfully. Please review the preview before importing.",
            "success"
        )

    # --------------------------------------------
    # If dataset already uploaded but not imported
    # --------------------------------------------

    elif "dataset_path" in session:

        filepath = session["dataset_path"]

        if os.path.exists(filepath):

            try:

                df = pd.read_csv(filepath)

                df.columns = df.columns.str.strip()

                preview = df.head(10).values.tolist()

                columns = df.columns.tolist()

                uploaded_filename = session.get(
                    "uploaded_filename"
                )

                total_rows = len(df)

                total_columns = len(df.columns)

            except:

                pass

    # --------------------------------------------
    # Render Page
    # --------------------------------------------

    return render_template(

        "dataset_management.html",

        dataset=dataset,

        history=history,

        preview=preview,

        columns=columns,

        uploaded_filename=uploaded_filename,

        total_rows=total_rows,

        total_columns=total_columns

    )


# ==========================================================
# IMPORT UPLOADED DATASET
# ==========================================================

@app.route("/import_uploaded_dataset", methods=["POST"])
def import_uploaded_dataset():

    # ============================================
    # USER LOGIN CHECK
    # ============================================

    if "user_id" not in session:
        return redirect(url_for("login"))
    user_id = session["user_id"]

    conn = sqlite3.connect("finance.db")
    cursor = conn.cursor()


    conn.commit()

    # ============================================
    # DATASET CHECK
    # ============================================

    if "dataset_path" not in session:

        flash(
            "Please upload a dataset first.",
            "warning"
        )

        return redirect(
            url_for("dataset_management")
        )


    dataset_path = session["dataset_path"]

    dataset_name = session["dataset_name"]

    print("=" * 70)
    print("IMPORT DATASET ROUTE STARTED")
    print("User ID :", user_id)
    print("Dataset :", dataset_name)
    print("Path :", dataset_path)
    print("=" * 70)

    # ============================================
    # CHECK FILE EXISTS
    # ============================================

    if not os.path.exists(dataset_path):

        flash(
            "Uploaded CSV file not found.",
            "danger"
        )

        return redirect(
            url_for("dataset_management")
        )

    # ============================================
    # CREATE FILE HASH
    # ============================================

    with open(dataset_path, "rb") as file:

        file_hash = hashlib.md5(
            file.read()
        ).hexdigest()

    print("FILE HASH :", file_hash)

    # ============================================
    # DATABASE
    # ============================================

    conn = sqlite3.connect("finance.db")

    cursor = conn.cursor()

    # ============================================
    # DUPLICATE CHECK
    # ============================================

    cursor.execute("""

        SELECT COUNT(*)

        FROM dataset_info

        WHERE user_id=?

        AND file_hash=?

    """,

    (

        user_id,

        file_hash

    ))

    duplicate = cursor.fetchone()[0]

    print("Duplicate :", duplicate)

    if duplicate > 0:

        conn.close()

        flash(

            "This dataset has already been imported.",

            "warning"

        )

        return redirect(
            url_for("dataset_management")
        )

    
        # ============================================
    # READ CSV
    # ============================================

    try:

        df = pd.read_csv(dataset_path)

    except Exception as e:

        conn.close()

        flash(

            f"Unable to read CSV file : {e}",

            "danger"

        )

        return redirect(

            url_for("dataset_management")

        )

    # ============================================
    # CLEAN COLUMN NAMES
    # ============================================

    df.columns = df.columns.str.strip()

    print("=" * 70)
    print("CSV COLUMNS")
    print(df.columns.tolist())
    print("=" * 70)

    # ============================================
    # REQUIRED COLUMNS
    # ============================================

    required_columns = [

        "Date",

        "Type",

        "Category",

        "Amount"

    ]

    missing_columns = []

    for column in required_columns:

        if column not in df.columns:

            missing_columns.append(column)

    if len(missing_columns) > 0:

        conn.close()

        flash(

            "Missing Columns : " +

            ", ".join(missing_columns),

            "danger"

        )

        return redirect(

            url_for("dataset_management")

        )

    # ============================================
    # DESCRIPTION COLUMN
    # ============================================

    description_column = None

    possible_columns = [

        "Transaction",

        "Description",

        "Details",

        "Transaction Description"

    ]

    for col in possible_columns:

        if col in df.columns:

            description_column = col

            break

    print("Description Column :", description_column)

    print("Total Rows :", len(df))

    # ============================================
    # START DATABASE TRANSACTION
    # ============================================

    conn.execute("BEGIN")

    income_count = 0

    expense_count = 0

    total_records = 0

    print("=" * 70)
    print("STARTING IMPORT")
    print("=" * 70)



        # ============================================
    # IMPORT RECORDS
    # ============================================

    try:

        for index, row in df.iterrows():

            # -----------------------------
            # VALIDATION
            # -----------------------------

            if pd.isna(row["Amount"]):
                continue

            if pd.isna(row["Category"]):
                continue

            try:

                amount = float(row["Amount"])

            except:

                continue

            if amount <= 0:
                continue

            transaction_type = str(
                row["Type"]
            ).strip().lower()

            transaction_type = transaction_type.replace("credit", "income")
            transaction_type = transaction_type.replace("debit", "expense")
            transaction_type = transaction_type.replace("cr", "income")
            transaction_type = transaction_type.replace("dr", "expense")

            category = str(
                row["Category"]
            ).strip()

            date = str(
                row["Date"]
            )

            # -----------------------------
            # DESCRIPTION
            # -----------------------------

            if description_column:

                if pd.isna(row[description_column]):

                    description = "Imported Dataset"

                else:

                    description = str(

                        row[description_column]

                    )

            else:

                description = "Imported Dataset"

            # -----------------------------
            # DEBUG
            # -----------------------------

            print(

                f"Row {index+1} :",

                transaction_type,

                category,

                amount

            )

            # -----------------------------
            # INCOME
            # -----------------------------

            if transaction_type == "income":

                cursor.execute("""

                    INSERT INTO income

                    (

                        user_id,

                        source,

                        amount,

                        date,

                        imported

                    )

                    VALUES

                    (

                        ?,

                        ?,

                        ?,

                        ?,

                        ?

                    )

                """,

                (

                    user_id,

                    category,

                    amount,

                    date,

                    1

                ))

                income_count += 1

            # -----------------------------
            # EXPENSE
            # -----------------------------

            elif transaction_type == "expense":

                cursor.execute("""

                    INSERT INTO expense

                    (

                        user_id,

                        category,

                        amount,

                        description,

                        date,

                        imported

                    )

                    VALUES

                    (

                        ?,

                        ?,

                        ?,

                        ?,

                        ?,

                        ?

                    )

                """,

                (

                    user_id,

                    category,

                    amount,

                    description,

                    date,

                    1

                ))

                expense_count += 1

            else:

                print(

                    "Skipped Unknown Type :",

                    transaction_type

                )

                continue

            total_records += 1

        print("=" * 60)
        print("IMPORT COMPLETED")
        print("Income :", income_count)
        print("Expense :", expense_count)
        print("Total :", total_records)
        print("=" * 60)



        # ============================================
        # SAVE DATASET INFORMATION
        # ============================================

        cursor.execute("""

            DELETE FROM dataset_info

            WHERE user_id=?

        """,

        (

            user_id,

        ))

        cursor.execute("""

            INSERT INTO dataset_info
            (
                user_id,
                dataset_name,
                total_rows,
                income_rows,
                expense_rows,
                upload_date,
                status,
                file_hash
            )

            VALUES
            (
                ?,
                ?,
                ?,
                ?,
                ?,
                datetime('now'),
                ?,
                ?
            )

        """,

        (
            user_id,
            dataset_name,
            total_records,
            income_count,
            expense_count,
            "Imported",
            file_hash
        ))

        print("✓ dataset_info updated")

        # ============================================
        # SAVE IMPORT HISTORY
        # ============================================

        cursor.execute("""

            INSERT INTO dataset_history

            (

                user_id,

                dataset_name,

                action,

                action_date

            )

            VALUES

            (

                ?,

                ?,

                'Imported',

                datetime('now')

            )

        """,

        (

            user_id,

            dataset_name

        ))

        print("✓ dataset_history updated")

        # ============================================
        # COMMIT DATABASE
        # ============================================

        conn.commit()

        print("✓ Database Commit Successful")

        # ============================================
        # CLEAR SESSION
        # ============================================

        session.pop("dataset_path", None)
        session.pop("dataset_name", None)
        session.pop("uploaded_filename", None)

        # ============================================
        # SUCCESS MESSAGE
        # ============================================

        flash(

            f"""
        Dataset Imported Successfully!

        Dataset Name : {dataset_name}

        Income Imported : {income_count}

        Expense Imported : {expense_count}

        Total Imported : {total_records}
        """,

                    "success"

                )

    except Exception as e:

        import traceback

        conn.rollback()

        print("\n" + "=" * 80)
        print("❌ IMPORT FAILED")
        print("=" * 80)

        print("ERROR TYPE :", type(e).__name__)

        print("ERROR MESSAGE :", str(e))

        print("\nFULL TRACEBACK:\n")

        traceback.print_exc()

        print("=" * 80)

        flash(

            f"Import Failed : {str(e)}",

            "danger"

        )

    finally:

        conn.close()

    return redirect(

        url_for("dataset_management")

    )



# ==========================================================
# DELETE IMPORTED DATASET
# ==========================================================

@app.route("/delete_imported_dataset", methods=["POST"])
def delete_imported_dataset():

    if "user_id" not in session:
        return redirect(url_for("login"))

    user_id = session["user_id"]

    conn = sqlite3.connect("finance.db")
    cursor = conn.cursor()
    try:

        # -----------------------------
        # Delete imported income only
        # -----------------------------

        cursor.execute("""

            DELETE FROM income

            WHERE user_id=?

            AND imported=1

        """, (user_id,))

        income_deleted = cursor.rowcount

        # -----------------------------
        # Delete imported expense only
        # -----------------------------

        cursor.execute("""

            DELETE FROM expense

            WHERE user_id=?

            AND imported=1

        """, (user_id,))

        expense_deleted = cursor.rowcount

        # -----------------------------
        # Get dataset name
        # -----------------------------

        cursor.execute("""

            SELECT dataset_name

            FROM dataset_info

            WHERE user_id=?

        """, (user_id,))

        row = cursor.fetchone()

        if row:

            dataset_name = row[0]

        else:

            dataset_name = "Unknown Dataset"

        # -----------------------------
        # Delete dataset information
        # -----------------------------

        cursor.execute("""

            DELETE FROM dataset_info

            WHERE user_id=?

        """, (user_id,))

        # -----------------------------
        # Save history
        # -----------------------------

        cursor.execute("""

            INSERT INTO dataset_history
            (
                user_id,
                dataset_name,
                action,
                action_date
            )

            VALUES
            (
                ?,
                ?,
                'Deleted',
                datetime('now')
            )

        """,

        (

            user_id,

            dataset_name

        ))

        # -----------------------------
        # Commit Changes
        # -----------------------------
        print("\nSaving dataset_info...")

        cursor.execute("SELECT COUNT(*) FROM income WHERE imported=1")
        print("Imported Income :", cursor.fetchone()[0])

        cursor.execute("SELECT COUNT(*) FROM expense WHERE imported=1")
        print("Imported Expense :", cursor.fetchone()[0])

        print("Committing Database...")

        conn.commit()

        flash(

            f"""
Imported Dataset Deleted Successfully!

Income Records Deleted : {income_deleted}

Expense Records Deleted : {expense_deleted}
""",

            "success"

        )

    except Exception as e:

        conn.rollback()

        flash(

            f"Delete Failed : {e}",

            "danger"

        )

    finally:

        conn.close()

    return redirect(

        url_for("dataset_management")

    )














if __name__=="__main__":
    app.run(debug=True)
